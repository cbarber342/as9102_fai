import openpyxl
import copy
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.cell.cell import MergedCell
from typing import List
import logging

from as9102_fai.parsers.chr_parser import FaiCharacteristic


logger = logging.getLogger(__name__)


class FaiGenerator:
    def __init__(self, template_path: str):
        self.template_path = template_path
        self.wb = None
        self.sheet = None

    def generate_report(
        self,
        characteristics: List[FaiCharacteristic],
        output_path: str,
        notes: str = "",
    ) -> bool:
        """Populate the Form 3 sheet in the Excel template.

        Column mapping (based on the user's template):

        - Col B (2): 5. Char No.                  -> sequential index (1, 2, 3, ...)
        - Col C (3): Op #                         -> left blank for now
        - Col D (4): 6. Reference Location        -> `char.group1` when present
        - Col E (5): Bubble Number                -> characteristic id
        - Col F (6): 7. Characteristic Designator -> use `char.type` when present
        - Col G (7): Description / Note text      -> `char.feature_name`
        - Col H (8): Specification                -> formatted tolerance string (`char.description`)
        - Col I (9): GDT Callout                  -> not populated yet
        - Col J (10): Unit of measurement         -> `char.unit` when present
        - Col K (11): Bonus Tolerance             -> left blank for now
        - Col L (12): 9. Results                  -> `char.actual`
        """

        try:
            self.wb = openpyxl.load_workbook(self.template_path)
        except Exception as e:
            logger.exception("Error loading template")
            return False

        # Find Form 3 sheet
        sheet_name = None
        for name in self.wb.sheetnames:
            if "Form 3" in name:
                sheet_name = name
                break

        if not sheet_name:
            logger.error("Form 3 sheet not found")
            return False

        self.sheet = self.wb[sheet_name]

        def get_unmerged_cell(row: int, column: int):
            """Return a writable cell at (row, column), unmerging if needed.

            Some templates merge header and body cells vertically. When we
            write body data, we want individual rows, so if the target cell
            is part of a merged range, we first unmerge that range.
            """

            cell = self.sheet.cell(row=row, column=column)
            if not isinstance(cell, MergedCell):
                return cell

            # Unmerge the range that contains this cell, then return the
            # newly created regular cell at the requested coordinates.
            for merged_range in list(self.sheet.merged_cells.ranges):
                if cell.coordinate in merged_range:
                    # Use openpyxl's API to unmerge so the underlying
                    # cells become normal, writable Cell objects.
                    self.sheet.unmerge_cells(str(merged_range))
                    break

            return self.sheet.cell(row=row, column=column)

        def set_cell_value(row: int, column: int, value):
            cell = get_unmerged_cell(row, column)
            cell.value = value

        def has_visible_border(row: int, column: int) -> bool:
            cell = self.sheet.cell(row=row, column=column)
            border = getattr(cell, "border", None)
            if border is None:
                return False
            for side_name in ("top", "bottom", "left", "right"):
                side = getattr(border, side_name, None)
                if side is not None and getattr(side, "style", None) not in (None, "", "none"):
                    return True
            return False

        def find_template_table_bottom_row(start_row: int) -> int:
            sheet_max = int(getattr(self.sheet, "max_row", 0) or 0)
            max_scan = min(max(sheet_max, start_row + 60), start_row + 250)
            last_seen = start_row
            empty_run = 0
            found_any = False
            max_col = min(int(getattr(self.sheet, "max_column", 0) or 0), 26)
            for row in range(start_row, max_scan + 1):
                row_has_structure = False
                for col in range(2, max_col + 1):
                    cell = self.sheet.cell(row=row, column=col)
                    if cell.value is not None and str(cell.value).strip() != "":
                        row_has_structure = True
                        break
                    if has_visible_border(row, col):
                        row_has_structure = True
                        break
                if row_has_structure:
                    found_any = True
                    last_seen = row
                    empty_run = 0
                elif found_any:
                    empty_run += 1
                    if empty_run >= 15:
                        break
            # Use ws.max_row as fallback — Excel Table styles on empty rows
            # aren't visible to openpyxl as cell-level borders.
            return max(start_row, last_seen, sheet_max)

        def ensure_form3_title_merge() -> None:
            try:
                title_cell = self.sheet.cell(row=2, column=2)
                title_text = str(title_cell.value or "").strip()
                if not title_text or "form 3" not in title_text.lower():
                    return

                desired = (2, 2, 2, 12)  # B2:L2
                overlapping: list[str] = []
                already_merged = False
                for mr in list(self.sheet.merged_cells.ranges):
                    if mr.max_row < desired[0] or mr.min_row > desired[2] or mr.max_col < desired[1] or mr.min_col > desired[3]:
                        continue
                    if (mr.min_row, mr.min_col, mr.max_row, mr.max_col) == desired:
                        already_merged = True
                        break
                    overlapping.append(str(mr))

                if not already_merged:
                    for rng in overlapping:
                        self.sheet.unmerge_cells(rng)
                    self.sheet.merge_cells(start_row=2, start_column=2, end_row=2, end_column=12)

                try:
                    title_cell.alignment = (title_cell.alignment or Alignment()).copy(horizontal="left", vertical="center")
                except Exception:
                    title_cell.alignment = Alignment(horizontal="left", vertical="center")
            except Exception:
                pass

        def apply_form3_border_tweaks(last_row: int) -> None:
            _thin = Side(style="thin")
            _none = Side()
            max_col = min(max(int(getattr(self.sheet, "max_column", 0) or 0), 16), 26)

            def visible(side) -> bool:
                try:
                    return bool(side and getattr(side, "style", None) not in (None, "", "none"))
                except Exception:
                    return False

            def rebuild_border(cell, *, left=None, right=None, top=None, bottom=None):
                old = getattr(cell, "border", None) or Border()
                return Border(
                    left=left if left is not None else old.left,
                    right=right if right is not None else old.right,
                    top=top if top is not None else old.top,
                    bottom=bottom if bottom is not None else old.bottom,
                    diagonal=old.diagonal,
                )

            def edge_border(cell, *, force_left=False, force_right=False, top=None, bottom=None):
                old = getattr(cell, "border", None) or Border()
                old_left = old.left if visible(getattr(old, "left", None)) else _thin
                old_right = old.right if visible(getattr(old, "right", None)) else _thin
                old_top = old.top if visible(getattr(old, "top", None)) else _thin
                old_bottom = old.bottom if visible(getattr(old, "bottom", None)) else _thin
                return Border(
                    left=_thin if force_left else old_left,
                    right=_thin if force_right else old_right,
                    top=top if top is not None else old_top,
                    bottom=bottom if bottom is not None else old_bottom,
                    diagonal=old.diagonal,
                )

            # Rows 1-3: truly borderless.
            for row in range(1, min(int(getattr(self.sheet, "max_row", 0) or 0), 3) + 1):
                for col in range(1, max_col + 1):
                    cell = self.sheet.cell(row=row, column=col)
                    if isinstance(cell, MergedCell):
                        continue
                    cell.border = Border()

            # Column M(13)/N(14): add horizontal borders.  Column P(16): hide horizontal borders.
            for row in range(4, max(4, int(last_row or 4)) + 1):
                for cc, force_left, force_right in ((2, True, False),):
                    cell = self.sheet.cell(row=row, column=cc)
                    if isinstance(cell, MergedCell):
                        continue
                    cell.border = edge_border(cell, force_left=force_left, force_right=force_right)
                for cc in (13, 14):
                    cell = self.sheet.cell(row=row, column=cc)
                    if isinstance(cell, MergedCell):
                        continue
                    cell.border = rebuild_border(cell, top=_thin, bottom=_thin)
                cell = self.sheet.cell(row=row, column=16)
                if not isinstance(cell, MergedCell):
                    cell.border = edge_border(cell, force_left=False, force_right=False, top=_none, bottom=_none)
                for cc in (21, 22):
                    cell = self.sheet.cell(row=row, column=cc)
                    if isinstance(cell, MergedCell):
                        continue
                    cell.border = Border()

        # Determine start row (look for header containing "Char No.")
        # The data should start at row 6 based on the template structure
        start_row = 6  # Fixed start row based on template
        for row in range(1, 20):
            cell_val = self.sheet.cell(row=row, column=2).value
            # Use "Char No." to be more specific and avoid matching the report title
            if cell_val and "Char No." in str(cell_val):
                start_row = row + 1
                break

        # Data starts immediately after the header row
        current_row = start_row
        template_table_bottom_row = find_template_table_bottom_row(start_row)
        ensure_form3_title_merge()

        # Styles for result cell
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        row_num = 0  # Track actual row number for Char No / Bubble Number
        for i, char in enumerate(characteristics, start=1):
            # Skip rows without a valid description (empty, whitespace only, or contains "nan")
            if not char.description or not char.description.strip():
                continue
            if "nan" in char.description.lower():
                continue

            row_num += 1  # Increment only for rows we actually write

            # Column B (2): Char No. - sequential number for rows with data
            set_cell_value(current_row, 2, row_num)

            # Column D (4): 6. Reference Location
            group1_val = getattr(char, "group1", "")
            if group1_val is not None and str(group1_val).strip():
                existing = get_unmerged_cell(current_row, 4).value
                if existing is None or str(existing).strip() == "":
                    set_cell_value(current_row, 4, str(group1_val).strip())
            
            # Column E (5): Bubble Number - same as Char No.
            set_cell_value(current_row, 5, row_num)
            
            # Column E (5): left blank
            # Column F (6): left blank (7. Characteristic Designator - not used)
            
            # Column G (7): Description / Note text - feature name + ID info
            description_text = f"{char.id}"
            desc_cell = get_unmerged_cell(current_row, 7)
            desc_cell.value = description_text
            try:
                desc_cell.alignment = (desc_cell.alignment or Alignment()).copy(wrap_text=True)
            except Exception:
                desc_cell.alignment = Alignment(wrap_text=True)
            
            # Column H (8): Specification - the tolerance/requirement string
            set_cell_value(current_row, 8, char.description)

            # Detect BASIC in Description/Note text and related fields.
            basic_text = " ".join(
                [
                    str(description_text or ""),
                    str(getattr(char, "description", "") or ""),
                    str(getattr(char, "comment", "") or ""),
                    str(getattr(char, "feature_name", "") or ""),
                ]
            )
            is_basic = "basic" in basic_text.lower()

            # Column I (9): GD&T Callout - force blank (and clear any template formula).
            try:
                set_cell_value(current_row, 9, None)
            except Exception:
                pass

            # Column J (10): Unit of measurement
            unit_val = getattr(char, "unit", "")
            if unit_val is not None and str(unit_val).strip():
                set_cell_value(current_row, 10, str(unit_val).strip())

            # Result (9. Results column) - round to 4 decimal places
            result_cell = get_unmerged_cell(current_row, 12)
            try:
                if is_basic:
                    result_cell.value = "NA"
                else:
                    result_val = float(char.actual)
                    result_cell.value = round(result_val, 4)
            except (ValueError, TypeError):
                result_cell.value = "NA" if is_basic else char.actual  # Keep as-is if not numeric
            
            # Conditional formatting / basic pass-fail coloring
            is_pass = True
            if getattr(char, "is_attribute", False):
                if is_basic:
                    pass
                elif str(char.actual).lower() == "pass":
                    result_cell.fill = green_fill
                elif not char.actual:
                    result_cell.fill = red_fill
                    is_pass = False
                else:
                    result_cell.fill = red_fill
                    is_pass = False
            else:
                # Numeric check
                try:
                    if is_basic:
                        raise ValueError("basic")
                    val = float(char.actual)
                    nom = float(char.nominal)
                    up = float(char.upper_tol) if char.upper_tol else 0.0
                    low = float(char.lower_tol) if char.lower_tol else 0.0

                    limit_high = nom + up
                    limit_low = nom + low

                    # Treat very large upper tolerance as BASIC (no pass/fail)
                    if abs(up) >= 990:
                        pass
                    elif val > limit_high + 1e-6 or val < limit_low - 1e-6:
                        result_cell.fill = red_fill
                        is_pass = False
                    else:
                        # In-tolerance: leave default fill (or set green if desired)
                        pass
                except (ValueError, TypeError):
                    if not char.actual:
                        result_cell.fill = red_fill

            current_row += 1

        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )

        target_bottom = max(template_table_bottom_row, current_row - 1)
        for row in range(start_row, target_bottom + 1):
            for col in range(2, min(int(getattr(self.sheet, 'max_column', 0) or 0), 26) + 1):
                cell = self.sheet.cell(row=row, column=col)
                if not isinstance(cell, MergedCell):
                    if has_visible_border(start_row, col):
                        cell.border = copy.copy(self.sheet.cell(row=start_row, column=col).border)
                    else:
                        cell.border = thin_border

        apply_form3_border_tweaks(target_bottom)

        # Append notes (e.g. extracted from PDF)
        if notes:
            current_row += 2
            set_cell_value(current_row, 1, "NOTES:")
            self.sheet.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1

            for line in notes.split("\n"):
                if line.strip():
                    set_cell_value(current_row, 1, line.strip())
                    current_row += 1

        try:
            self.wb.save(output_path)
            return True
        except Exception as e:
            logger.exception("Error saving report")
            return False
