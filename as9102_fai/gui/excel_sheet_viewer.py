from __future__ import annotations

from dataclasses import dataclass
from typing import Any, Callable, Dict, Optional, Tuple

import json
import re
import logging

from PySide6.QtCore import Qt, QEvent, QPoint, QTimer, Signal, QSignalBlocker, QItemSelection, QItemSelectionModel, QRect
from PySide6.QtGui import QColor, QFont, QPen, QGuiApplication, QKeySequence, QShortcut, QPalette, QTextOption, QFontMetrics
from PySide6.QtWidgets import QAbstractItemView, QAbstractItemDelegate, QStyledItemDelegate, QTableWidget, QTableWidgetItem, QWidget, QVBoxLayout, QMenu, QStyleOptionViewItem, QStyle, QTableWidgetSelectionRange, QTextEdit

from openpyxl.styles import PatternFill

logger = logging.getLogger(__name__)


BorderKey = Tuple[int, int]  # (row, col) 1-based


@dataclass(frozen=True)
class CellOverride:
    value: Any = None
    background: Optional[QColor] = None


class _ExcelBorderDelegate(QStyledItemDelegate):
    """Paint per-cell borders based on data stored on the item."""

    BORDER_ROLE = Qt.ItemDataRole.UserRole + 101

    # Optional visual overrides set by the owning ExcelSheetViewer.
    # These are intentionally palette/constant driven so we don't have to
    # re-encode Excel's original border colors.
    border_color_override: Optional[QColor] = None
    border_width_scale: float = 1.0

    def paint(self, painter, option, index):
        super().paint(painter, option, index)

        borders = index.data(self.BORDER_ROLE)
        if not borders:
            return

        rect = option.rect

        def _draw_side(side: str, pen: QPen):
            painter.setPen(pen)
            if side == "top":
                painter.drawLine(rect.topLeft(), rect.topRight())
            elif side == "bottom":
                painter.drawLine(rect.bottomLeft(), rect.bottomRight())
            elif side == "left":
                painter.drawLine(rect.topLeft(), rect.bottomLeft())
            elif side == "right":
                painter.drawLine(rect.topRight(), rect.bottomRight())

        for side_name, spec in borders.items():
            width = spec.get("width", 0)
            if width <= 0:
                continue

            try:
                scale = float(getattr(self, "border_width_scale", 1.0) or 1.0)
            except Exception:
                scale = 1.0

            try:
                override = getattr(self, "border_color_override", None)
            except Exception:
                override = None
            color = override if override is not None else (spec.get("color") or QColor(0, 0, 0))

            pen = QPen(color)
            # Use floating widths so Form 3 can be subtly thinner.
            try:
                pen.setWidthF(max(0.6, float(width) * max(0.1, scale)))
            except Exception:
                pen.setWidth(int(width))
            _draw_side(side_name, pen)


class ExcelSheetViewer(QWidget):
    """Renders an openpyxl Worksheet into a QTableWidget (layout-focused)."""

    scaleChanged = Signal(float)  # effective scale factor (fit * user)
    paintModeCleared = Signal()
    # Request to insert a worksheet row at a given 1-based row index.
    # Payload: (row_1based, where) where where in {'above','below'}.
    rowInsertRequested = Signal(int, str)
    # Request to delete a worksheet row at a given 1-based row index.
    rowDeleteRequested = Signal(int)
    # Request to delete multiple worksheet rows (1-based row indices).
    # Payload: list[int]
    rowDeleteManyRequested = Signal(object)
    modified = Signal()

    # We clamp the *effective* scale (fit * zoom) to keep behavior consistent
    # across sheets of different physical sizes.
    MIN_EFFECTIVE_SCALE = 0.50
    MAX_EFFECTIVE_SCALE = 2.00
    MIN_USER_ZOOM = 0.10
    MAX_USER_ZOOM = 2000.0

    BASE_FONT_SIZE_ROLE = Qt.ItemDataRole.UserRole + 201

    CELL_COORD_ROLE = Qt.ItemDataRole.UserRole + 1  # (row, col) 1-based
    WRAP_ROLE = Qt.ItemDataRole.UserRole + 202

    def __init__(self, parent: Optional[QWidget] = None):
        super().__init__(parent)
        self._ws = None
        self._overrides: Dict[BorderKey, CellOverride] = {}
        self._covered_cells: set[BorderKey] = set()
        self._validation_lists: Dict[BorderKey, list[str]] = {}
        # Map cell -> {display_text: stored_value} for validations where we want
        # different display text than the stored worksheet value.
        self._validation_display_to_value: Dict[BorderKey, dict[str, str]] = {}
        # Map cell -> kind of mapped validation (e.g. 'supplier_code', 'supplier_directory')
        self._validation_mapped_kind: Dict[BorderKey, str] = {}
        self._hidden_columns_1based: set[int] = set()

        # Optional: enable row-insert context menu (used by Form 3).
        self._row_insert_menu_enabled: bool = False
        self._row_insert_min_row_1based: int = 6

        # Optional: allow caller to add per-cell context menu actions.
        # Signature: provider(row_1based, col_1based, menu)
        self._cell_context_menu_provider: Optional[Callable[[int, int, QMenu], None]] = None

        # Optional: allow caller to intercept wheel events for custom navigation.
        # Signature: handler(row_1based, col_1based, delta_y) -> bool (True = consumed)
        self._wheel_navigation_handler: Optional[Callable[[int, int, int], bool]] = None

        # Optional: allow caller to intercept key navigation for custom behavior.
        # Signature: handler(row_1based, col_1based, direction) -> bool (True = consumed)
        # direction: +1 = down, -1 = up
        self._key_navigation_handler: Optional[Callable[[int, int, int], bool]] = None

        # Optional: click-to-paint background fills.
        # When set, clicking a cell applies the fill to the underlying worksheet.
        # Use None to clear fill (no color).
        self._click_paint_fill_rgb: Optional[str] = None

        self._undo_stack: list[dict[BorderKey, tuple[Any, Any]]] = []
        self._redo_stack: list[dict[BorderKey, tuple[Any, Any]]] = []
        self._in_programmatic_change = False
        self._custom_undo_handler: Optional[Callable[[], bool]] = None
        self._fit_mode: str = "none"  # 'none' | 'width' | 'both'
        self._rendered = False
        self._base_col_widths: list[int] = []
        self._base_row_heights: list[int] = []
        self._last_effective_scale: float = 1.0
        self._fit_scale: float = 1.0
        self._user_zoom: float = 1.0
        # Additional multiplier applied to fonts only (does not affect row/col sizes).
        self._font_scale_multiplier: float = 1.0
        self._pending_effective_scale: Optional[float] = None
        self._lock_effective_scale: bool = False
        self._locked_effective_scale: Optional[float] = None

        # Optional: auto-fit row heights to wrapped content for specified columns.
        # Store as 0-based column indexes.
        self._auto_fit_row_height_cols: set[int] = set()
        self._suppress_resize_handlers: bool = False

        # Optional: draw selection as an outline (Excel-like) instead of a filled background.
        self._selection_outline_only: bool = False
        self._selection_outline_base_stylesheet: Optional[str] = None
        self._selection_outline_base_palette: Optional[QPalette] = None

        # Excel-like navigation: remember whether Enter/Shift+Enter was pressed
        # while editing, so we can move after the editor commits/closes.
        self._pending_enter_nav_dir: Optional[int] = None

        # Header selection anchors (Excel-like Shift selection).
        self._row_header_anchor0: Optional[int] = None
        self._col_header_anchor0: Optional[int] = None

        # Optional persistence for user-resized row/column base sizes.
        self._settings = None
        self._settings_key: Optional[str] = None
        self._persist_timer = QTimer(self)
        self._persist_timer.setSingleShot(True)
        self._persist_timer.setInterval(250)
        self._persist_timer.timeout.connect(self._persist_base_sizes)

        self._apply_timer = QTimer(self)
        self._apply_timer.setSingleShot(True)
        self._apply_timer.setInterval(30)
        self._apply_timer.timeout.connect(self._apply_scale)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)

        self.table = QTableWidget()
        self.table.setEditTriggers(
            QTableWidget.EditTrigger.DoubleClicked
            | QTableWidget.EditTrigger.EditKeyPressed
            | QTableWidget.EditTrigger.AnyKeyPressed
        )
        self.table.setSelectionMode(QTableWidget.SelectionMode.ExtendedSelection)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectItems)
        self.table.setFocusPolicy(Qt.FocusPolicy.StrongFocus)
        # Show headers to enable Excel-like resizing and a "freeze panes" feel
        # (headers stay visible while scrolling).
        self.table.verticalHeader().setVisible(True)
        self.table.horizontalHeader().setVisible(True)
        self.table.horizontalHeader().setSectionResizeMode(self.table.horizontalHeader().ResizeMode.Interactive)
        self.table.verticalHeader().setSectionResizeMode(self.table.verticalHeader().ResizeMode.Interactive)
        self.table.setShowGrid(False)
        # Wrap is handled per-cell via delegate using WRAP_ROLE.
        self.table.setWordWrap(False)
        # Context menu for optional row insertion actions.
        try:
            self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
            self.table.customContextMenuRequested.connect(self._on_table_context_menu)
        except Exception:
            pass
        self.table.viewport().installEventFilter(self)
        self.table.installEventFilter(self)
        try:
            self.table.verticalHeader().installEventFilter(self)
            self.table.horizontalHeader().installEventFilter(self)
        except Exception:
            pass

        self._border_delegate = _ExcelBorderDelegate(self.table)
        # Default: make form borders a light gray for a cleaner look.
        try:
            self._border_delegate.border_color_override = QColor(200, 200, 200)
        except Exception:
            pass
        try:
            self._border_delegate.border_width_scale = 1.0
        except Exception:
            pass
        self._edit_delegate = _ExcelEditDelegate(self)
        # Chain delegates by using the edit delegate and reusing the border painting.
        self._edit_delegate.border_delegate = self._border_delegate
        self.table.setItemDelegate(self._edit_delegate)

        # Move selection after editing closes when Enter/Shift+Enter is used.
        try:
            self._edit_delegate.closeEditor.connect(self._on_close_editor)
        except Exception:
            pass

        # When scrolling, only the visible set of items changes; fonts for newly
        # visible cells should be updated without re-scaling rows/columns.
        self.table.verticalScrollBar().valueChanged.connect(self._apply_font_scale_visible)
        self.table.horizontalScrollBar().valueChanged.connect(self._apply_font_scale_visible)

        # If auto-fit row heights is enabled, keep newly visible rows readable.
        self.table.verticalScrollBar().valueChanged.connect(self._apply_row_height_visible)
        self.table.horizontalScrollBar().valueChanged.connect(self._apply_row_height_visible)

        # Persist edits back to the worksheet.
        self.table.itemChanged.connect(self._on_item_changed)

        # Optional click paint behavior (used by Form 3 color selectors).
        self.table.cellClicked.connect(self._on_cell_clicked)

        # Ensure Ctrl+Z / Ctrl+Y works even if eventFilter misses key events.
        # Use WidgetWithChildrenShortcut so it triggers while the table/viewport has focus.
        try:
            undo_sc = QShortcut(QKeySequence.Undo, self.table)
            undo_sc.setContext(Qt.ShortcutContext.WidgetWithChildrenShortcut)
            undo_sc.activated.connect(self._on_undo_shortcut)
            redo_sc = QShortcut(QKeySequence.Redo, self.table)
            redo_sc.setContext(Qt.ShortcutContext.WidgetWithChildrenShortcut)
            redo_sc.activated.connect(self.redo)
            self._undo_shortcut = undo_sc
            self._redo_shortcut = redo_sc
        except Exception:
            self._undo_shortcut = None
            self._redo_shortcut = None

        # When user resizes rows/cols, treat that as a base-size change.
        self.table.horizontalHeader().sectionResized.connect(self._on_column_resized)
        self.table.verticalHeader().sectionResized.connect(self._on_row_resized)

        layout.addWidget(self.table)

    def set_border_width_scale(self, scale: float) -> None:
        """Scale border thickness (used to make Form 3 borders thinner)."""
        try:
            s = float(scale)
        except Exception:
            s = 1.0
        s = max(0.1, min(3.0, s))
        try:
            self._border_delegate.border_width_scale = float(s)
        except Exception:
            pass
        try:
            self.table.viewport().update()
        except Exception:
            pass

    def set_selection_outline_only(self, enabled: bool) -> None:
        """Render selected cells with an outline (no selection fill)."""
        try:
            self._selection_outline_only = bool(enabled)
        except Exception:
            self._selection_outline_only = False

        # Save/restore palette so native-style selection highlight doesn't bleed
        # through (some styles ignore item background overrides).
        try:
            if self._selection_outline_base_palette is None:
                self._selection_outline_base_palette = QPalette(self.table.palette())
        except Exception:
            self._selection_outline_base_palette = self._selection_outline_base_palette

        try:
            if self._selection_outline_only:
                pal = QPalette(self.table.palette())
                pal.setColor(QPalette.ColorRole.Highlight, QColor(0, 0, 0, 0))
                pal.setColor(QPalette.ColorRole.HighlightedText, pal.color(QPalette.ColorRole.Text))
                self.table.setPalette(pal)
                try:
                    self.table.viewport().setPalette(pal)
                except Exception:
                    pass
            else:
                if self._selection_outline_base_palette is not None:
                    self.table.setPalette(self._selection_outline_base_palette)
                    try:
                        self.table.viewport().setPalette(self._selection_outline_base_palette)
                    except Exception:
                        pass
        except Exception:
            pass

        # Qt may paint selected row/item panels behind the delegate, resulting in
        # theme-colored selection fills (blue/brown). Force selection background
        # to transparent while outline-only mode is enabled.
        try:
            if self._selection_outline_base_stylesheet is None:
                self._selection_outline_base_stylesheet = self.table.styleSheet() or ""
        except Exception:
            self._selection_outline_base_stylesheet = self._selection_outline_base_stylesheet or ""

        try:
            base = self._selection_outline_base_stylesheet or ""
            if self._selection_outline_only:
                override = (
                    "\n"
                    "QTableWidget, QTableView {"
                    " selection-background-color: rgba(0,0,0,0);"
                    " selection-color: palette(text);"
                    "}\n"
                    "QTableWidget::item:selected {"
                    " background-color: rgba(0,0,0,0);"
                    " color: palette(text);"
                    "}\n"
                    "QTableView::item:selected {"
                    " background-color: rgba(0,0,0,0);"
                    " color: palette(text);"
                    "}\n"
                    "QTableWidget::item:selected:active {"
                    " background-color: rgba(0,0,0,0);"
                    " color: palette(text);"
                    "}\n"
                    "QTableView::item:selected:active {"
                    " background-color: rgba(0,0,0,0);"
                    " color: palette(text);"
                    "}\n"
                    "QTableWidget::item:selected:!active {"
                    " background-color: rgba(0,0,0,0);"
                    " color: palette(text);"
                    "}\n"
                    "QTableView::item:selected:!active {"
                    " background-color: rgba(0,0,0,0);"
                    " color: palette(text);"
                    "}\n"
                    "QTableWidget::item:focus, QTableView::item:focus { outline: none; }\n"
                )
                self.table.setStyleSheet(base + override)
            else:
                self.table.setStyleSheet(base)
        except Exception:
            pass
        try:
            self.table.viewport().update()
        except Exception:
            pass

    def set_custom_undo_handler(self, handler: Optional[Callable[[], bool]]) -> None:
        """Allow parent to intercept Ctrl+Z; return True if handled."""
        try:
            self._custom_undo_handler = handler
        except Exception:
            self._custom_undo_handler = None

    def _on_undo_shortcut(self) -> None:
        try:
            handler = getattr(self, "_custom_undo_handler", None)
            if callable(handler):
                try:
                    if bool(handler()):
                        try:
                            print("Form3 Ctrl+Z handled by custom undo")
                        except Exception:
                            pass
                        try:
                            logger.debug("ExcelSheetViewer: custom undo handler consumed Ctrl+Z")
                        except Exception:
                            pass
                        return
                except Exception:
                    pass
        except Exception:
            pass
        try:
            try:
                print("Ctrl+Z fallback to internal undo")
            except Exception:
                pass
            logger.debug("ExcelSheetViewer: falling back to internal undo")
        except Exception:
            pass
        self.undo()

    def enable_row_insert_context_menu(self, enabled: bool, *, min_row_1based: int = 6) -> None:
        """Enable right-click menu for inserting rows (below a minimum row)."""
        try:
            self._row_insert_menu_enabled = bool(enabled)
        except Exception:
            self._row_insert_menu_enabled = False
        try:
            self._row_insert_min_row_1based = max(1, int(min_row_1based))
        except Exception:
            self._row_insert_min_row_1based = 6

    def set_cell_context_menu_provider(self, provider: Optional[Callable[[int, int, QMenu], None]]) -> None:
        """Set an optional callback to add per-cell context menu actions."""
        try:
            self._cell_context_menu_provider = provider
        except Exception:
            self._cell_context_menu_provider = None

    def set_wheel_navigation_handler(self, handler: Optional[Callable[[int, int, int], bool]]) -> None:
        """Set an optional handler for mouse wheel navigation within the table."""
        try:
            self._wheel_navigation_handler = handler
        except Exception:
            self._wheel_navigation_handler = None

    def set_key_navigation_handler(self, handler: Optional[Callable[[int, int, int], bool]]) -> None:
        """Set an optional handler for Up/Down key navigation within the table."""
        try:
            self._key_navigation_handler = handler
        except Exception:
            self._key_navigation_handler = None

    def _on_table_context_menu(self, pos: QPoint) -> None:
        try:
            idx = self.table.indexAt(pos)
        except Exception:
            idx = None
        if idx is None or not getattr(idx, "isValid", lambda: False)():
            return

        row_1based = int(idx.row()) + 1
        col_1based = int(idx.column()) + 1

        row_insert_enabled = bool(getattr(self, "_row_insert_menu_enabled", False))
        provider = getattr(self, "_cell_context_menu_provider", None)
        if (not row_insert_enabled) and (provider is None):
            return

        menu = QMenu(self.table)

        # Custom per-cell actions (e.g., Form 3 "Find bubble").
        if provider is not None:
            try:
                provider(int(row_1based), int(col_1based), menu)
            except Exception:
                pass

        act_above = None
        act_below = None
        act_delete = None
        act_delete_selected = None
        if row_insert_enabled:
            try:
                min_row = int(getattr(self, "_row_insert_min_row_1based", 6) or 6)
            except Exception:
                min_row = 6
            if int(row_1based) >= int(min_row):
                try:
                    if menu.actions():
                        menu.addSeparator()
                except Exception:
                    pass
                act_above = menu.addAction("Insert row above")
                act_below = menu.addAction("Insert row below")
                try:
                    act_delete = menu.addAction("Delete row")
                except Exception:
                    act_delete = None

                # Multi-row delete (if multiple distinct rows are selected and the
                # right-click occurred within the selection).
                try:
                    sel_rows = {int(ix.row()) + 1 for ix in (self.table.selectedIndexes() or [])}
                except Exception:
                    sel_rows = set()
                try:
                    if sel_rows and int(row_1based) in sel_rows and len(sel_rows) > 1:
                        eligible = [r for r in sel_rows if int(r) >= int(min_row)]
                        if len(eligible) > 1:
                            act_delete_selected = menu.addAction(f"Delete selected rows ({len(eligible)})")
                except Exception:
                    act_delete_selected = None

        # If no actions ended up being added, don't show an empty menu.
        try:
            if not list(menu.actions() or []):
                return
        except Exception:
            pass

        chosen = menu.exec(self.table.viewport().mapToGlobal(pos))
        if chosen is None:
            return
        if act_above is not None and chosen == act_above:
            try:
                self.rowInsertRequested.emit(int(row_1based), "above")
            except Exception:
                pass
        elif act_below is not None and chosen == act_below:
            try:
                self.rowInsertRequested.emit(int(row_1based), "below")
            except Exception:
                pass
        elif act_delete is not None and chosen == act_delete:
            try:
                self.rowDeleteRequested.emit(int(row_1based))
            except Exception:
                pass
        elif act_delete_selected is not None and chosen == act_delete_selected:
            try:
                min_row = int(getattr(self, "_row_insert_min_row_1based", 6) or 6)
            except Exception:
                min_row = 6
            try:
                rows = sorted({int(ix.row()) + 1 for ix in (self.table.selectedIndexes() or []) if (int(ix.row()) + 1) >= int(min_row)})
            except Exception:
                rows = []
            if len(rows) > 1:
                try:
                    self.rowDeleteManyRequested.emit(list(rows))
                except Exception:
                    pass

    def set_click_paint_fill_rgb(self, rgb: Optional[str]) -> None:
        """Enable click-to-paint behavior using an Excel RGB hex string.

        Examples: "FFC7CE" (light red), "FFEB9C" (light orange/yellow).
        Use "" (empty string) to clear fill on click.
        Use None to disable paint-on-click.
        """
        if rgb is None:
            self._click_paint_fill_rgb = None
            return
        s = str(rgb).strip().lstrip("#")
        # Empty string means "clear fill" while staying in paint mode.
        if not s:
            self._click_paint_fill_rgb = ""
            return
        self._click_paint_fill_rgb = s.upper()

    def clear_click_paint_mode(self) -> None:
        self._click_paint_fill_rgb = None
        try:
            self.paintModeCleared.emit()
        except Exception:
            pass

    def apply_fill_to_selection(self, rgb: Optional[str], *, include_current_if_none: bool = True) -> bool:
        """Apply a background fill to the selected cells (Excel-like).

        - If no cells are selected and include_current_if_none is True, applies to the current cell.
        - rgb can be a 6-char RGB string (e.g. "FFC7CE"), "" to clear fill, or None (treated as "" for clearing).
        Returns True if at least one cell was updated.
        """
        if self._ws is None:
            return False

        # Gather targets as item-backed coordinates so we respect merged/covered cells.
        selected: list[tuple[int, int]] = []
        try:
            for ix in (self.table.selectedIndexes() or []):
                try:
                    selected.append((int(ix.row()), int(ix.column())))
                except Exception:
                    continue
        except Exception:
            selected = []

        if not selected and include_current_if_none:
            try:
                r0 = int(self.table.currentRow())
                c0 = int(self.table.currentColumn())
                if r0 >= 0 and c0 >= 0:
                    selected = [(r0, c0)]
            except Exception:
                selected = []

        if not selected:
            return False

        applied = False
        cmd: dict[BorderKey, tuple[Any, Any]] = {}

        # Normalize rgb: None => clear fill.
        s = "" if rgb is None else str(rgb)

        for r0, c0 in sorted(set(selected)):
            it = None
            try:
                it = self.table.item(int(r0), int(c0))
            except Exception:
                it = None
            if it is None:
                continue
            try:
                coord = it.data(Qt.ItemDataRole.UserRole + 1)
            except Exception:
                coord = None
            if not coord:
                continue
            try:
                r1, c1 = int(coord[0]), int(coord[1])
            except Exception:
                continue
            if (r1, c1) in self._covered_cells:
                continue
            before = None
            try:
                before = self._cell_fill_rgb(self._ws.cell(row=r1, column=c1))
            except Exception:
                before = None
            self._apply_cell_fill(int(r1), int(c1), s, push_undo=False)
            after = None
            try:
                after = self._cell_fill_rgb(self._ws.cell(row=r1, column=c1))
            except Exception:
                after = None
            if (before or "") != (after or ""):
                applied = True
                cmd[(int(r1), int(c1))] = ({"fill_rgb": before}, {"fill_rgb": (after or None)})

        if cmd:
            try:
                self._undo_stack.append(cmd)
                self._redo_stack.clear()
            except Exception:
                pass

        if applied:
            try:
                self.modified.emit()
            except Exception:
                pass
        return applied

    def _cell_fill_rgb(self, cell) -> Optional[str]:
        """Return 6-char RGB hex (no alpha) if cell has a solid fill."""
        try:
            fill = getattr(cell, "fill", None)
            if not fill or getattr(fill, "patternType", None) != "solid":
                return None
            fg = getattr(fill, "fgColor", None)
            rgb = getattr(fg, "rgb", None) if fg is not None else None
            if not rgb:
                return None
            rgb = str(rgb).strip().lstrip("#")
            # openpyxl may store ARGB
            if len(rgb) == 8:
                rgb = rgb[2:]
            if len(rgb) != 6:
                return None
            return rgb.upper()
        except Exception:
            return None

    def _apply_cell_fill(self, r1: int, c1: int, rgb: Optional[str], *, push_undo: bool = True) -> None:
        if self._ws is None:
            return

        try:
            r1 = int(r1)
            c1 = int(c1)
        except Exception:
            return

        try:
            cell = self._ws.cell(row=r1, column=c1)
        except Exception:
            return

        old_rgb = self._cell_fill_rgb(cell)
        new_rgb = str(rgb or "").strip().lstrip("#")
        if new_rgb:
            new_rgb = new_rgb.upper()
            if len(new_rgb) == 8:
                new_rgb = new_rgb[2:]
            if len(new_rgb) != 6:
                new_rgb = ""

        # No-op
        if (old_rgb or "") == (new_rgb or ""):
            return

        def _qcolor_from_rgb(s: str) -> Optional[QColor]:
            s = str(s or "").strip().lstrip("#")
            if len(s) == 8:
                s = s[2:]
            if len(s) != 6:
                return None
            try:
                return QColor(int(s[0:2], 16), int(s[2:4], 16), int(s[4:6], 16))
            except Exception:
                return None

        def _ideal_text_color(bg: Optional[QColor]) -> Optional[QColor]:
            if bg is None:
                return None
            return QColor(0, 0, 0) if bg.lightness() >= 140 else QColor(255, 255, 255)

        # Apply to worksheet.
        if not new_rgb:
            try:
                cell.fill = PatternFill()
            except Exception:
                pass
        else:
            try:
                cell.fill = PatternFill(start_color=new_rgb, end_color=new_rgb, fill_type="solid")
            except Exception:
                pass

        # Update visible item if present.
        it = self.table.item(r1 - 1, c1 - 1)
        if it is not None:
            if not new_rgb:
                try:
                    it.setData(Qt.ItemDataRole.BackgroundRole, None)
                    it.setData(Qt.ItemDataRole.ForegroundRole, None)
                except Exception:
                    pass
            else:
                bg = _qcolor_from_rgb(new_rgb)
                if bg is not None:
                    try:
                        it.setBackground(bg)
                    except Exception:
                        pass
                    fg = _ideal_text_color(bg)
                    if fg is not None:
                        try:
                            it.setForeground(fg)
                        except Exception:
                            pass

        if push_undo:
            self._undo_stack.append({(r1, c1): ({"fill_rgb": old_rgb}, {"fill_rgb": (new_rgb or None)})})
            self._redo_stack.clear()

    def _apply_cell_value(self, r1: int, c1: int, value: Any, *, push_undo: bool = True) -> None:
        if self._ws is None:
            return
        try:
            cell = self._ws.cell(row=r1, column=c1)
            old_val = cell.value
            cell.value = value
            
            it = self.table.item(r1 - 1, c1 - 1)
            if it is not None:
                it.setText(str(value) if value is not None else "")
                
            if push_undo:
                # Simple undo support for value changes could be added here
                pass
        except Exception:
            pass

    def _on_cell_clicked(self, row0: int, col0: int) -> None:
        if self._ws is None:
            return

        item = self.table.item(row0, col0)
        if item is None:
            return

        try:
            coord = item.data(self.CELL_COORD_ROLE)
        except Exception:
            coord = None

        if not coord or not isinstance(coord, (tuple, list)) or len(coord) != 2:
            return

        try:
            r1 = int(coord[0])
            c1 = int(coord[1])
        except Exception:
            return

        # Apply fill to the worksheet cell and update the visible item.
        # None => paint mode off. "" => clear fill.
        if self._click_paint_fill_rgb is None:
            return
        rgb = self._click_paint_fill_rgb
        self._apply_cell_fill(r1, c1, rgb if rgb != "" else None, push_undo=True)

    def set_persistence(self, settings, key: str) -> None:
        """Enable persistence of user-resized row/column sizes.

        `key` should be unique per viewer instance purpose (e.g. forms/1).
        """
        self._settings = settings
        self._settings_key = str(key or "").strip() or None

        # If we already rendered, apply immediately.
        if self._rendered:
            self._restore_base_sizes()
            if not self._lock_effective_scale:
                self._recompute_fit_scale()
            self._schedule_apply()

    def _schedule_persist(self) -> None:
        if self._settings is None or not self._settings_key:
            return
        if self._suppress_resize_handlers:
            return
        self._persist_timer.start()

    def _persist_base_sizes(self) -> None:
        if self._settings is None or not self._settings_key:
            return
        if not self._base_col_widths or not self._base_row_heights:
            return
        try:
            payload = {
                "base_col_widths": [int(x) for x in self._base_col_widths],
                "base_row_heights": [int(x) for x in self._base_row_heights],
            }
            self._settings.setValue(f"{self._settings_key}/sizes", json.dumps(payload))
        except Exception:
            return

    def _restore_base_sizes(self) -> None:
        if self._settings is None or not self._settings_key:
            return
        try:
            raw = self._settings.value(f"{self._settings_key}/sizes", "", type=str)
        except Exception:
            raw = ""
        if not raw:
            return
        try:
            payload = json.loads(str(raw))
        except Exception:
            return

        cols = payload.get("base_col_widths") if isinstance(payload, dict) else None
        rows = payload.get("base_row_heights") if isinstance(payload, dict) else None
        if not isinstance(cols, list) or not isinstance(rows, list):
            return

        # Normalize lengths to current sheet size.
        if self.table.columnCount() <= 0 or self.table.rowCount() <= 0:
            return
        col_count = int(self.table.columnCount())
        row_count = int(self.table.rowCount())

        def _to_int_list(xs: list, n: int, fallback: list[int]) -> list[int]:
            out: list[int] = []
            for i in range(n):
                try:
                    v = int(xs[i])
                except Exception:
                    v = int(fallback[i]) if 0 <= i < len(fallback) else 1
                out.append(max(v, 1))
            return out

        self._base_col_widths = _to_int_list(cols, col_count, self._base_col_widths)
        self._base_row_heights = _to_int_list(rows, row_count, self._base_row_heights)

    def set_auto_fit(self, enabled: bool) -> None:
        # Back-compat: enable == fit both, disable == none
        self.set_fit_mode("both" if enabled else "none")

    def set_fit_mode(self, mode: str) -> None:
        mode = (mode or "").strip().lower()
        if mode not in {"none", "width", "both"}:
            mode = "none"
        self._fit_mode = mode
        if self._rendered:
            self._recompute_fit_scale()
            self._schedule_apply()

    def effective_scale(self) -> float:
        return float(self._fit_scale * self._user_zoom)

    def set_effective_scale(self, effective_scale: float) -> None:
        """Set the overall effective scale (fit * zoom).

        If called before render/layout, the requested scale is applied after
        the first fit pass.
        """
        try:
            effective_scale = float(effective_scale)
        except Exception:
            return
        effective_scale = max(self.MIN_EFFECTIVE_SCALE, min(effective_scale, self.MAX_EFFECTIVE_SCALE))

        # When the user explicitly sets a starting scale, we lock the effective
        # scale so fit-to-width correction and fit recomputation don't override it.
        self._lock_effective_scale = True
        self._locked_effective_scale = effective_scale

        # If we can't apply yet, store for the post-render fit.
        if (not self._rendered) or self._fit_scale <= 0:
            self._pending_effective_scale = effective_scale
            return

        desired_zoom = effective_scale / self._fit_scale
        self.set_zoom(desired_zoom)

    def reset_auto_scale(self) -> None:
        """Return to auto scaling based on fit mode (no user starting scale)."""
        self._pending_effective_scale = None
        self._lock_effective_scale = False
        self._locked_effective_scale = None
        self._user_zoom = 1.0
        if self._rendered:
            self._recompute_fit_scale()
            self._schedule_apply()

    def set_zoom(self, zoom: float) -> None:
        try:
            zoom = float(zoom)
        except Exception:
            return
        # Clamp zoom dynamically so that (fit_scale * user_zoom) stays within
        # [MIN_EFFECTIVE_SCALE, MAX_EFFECTIVE_SCALE].
        if self._rendered and self._fit_scale > 0:
            min_zoom = self.MIN_EFFECTIVE_SCALE / self._fit_scale
            max_zoom = self.MAX_EFFECTIVE_SCALE / self._fit_scale
            min_zoom = max(min_zoom, self.MIN_USER_ZOOM)
            max_zoom = min(max_zoom, self.MAX_USER_ZOOM)
            if min_zoom > max_zoom:
                # Extremely pathological fit_scale; fall back to a safe clamp.
                min_zoom, max_zoom = self.MIN_USER_ZOOM, self.MAX_USER_ZOOM
            zoom = max(min_zoom, min(zoom, max_zoom))
        else:
            zoom = max(self.MIN_USER_ZOOM, min(zoom, self.MAX_USER_ZOOM))
        if abs(zoom - self._user_zoom) < 1e-6:
            return
        self._user_zoom = zoom
        if self._rendered:
            self._schedule_apply()

    def font_scale_multiplier(self) -> float:
        try:
            return float(self._font_scale_multiplier)
        except Exception:
            return 1.0

    def set_font_scale_multiplier(self, multiplier: float) -> None:
        """Scale fonts without changing row/column sizes.

        multiplier is a factor (e.g. 1.0 = default, 1.15 = +15%).
        """
        try:
            multiplier = float(multiplier)
        except Exception:
            return
        multiplier = max(0.50, min(multiplier, 2.50))
        if abs(multiplier - float(getattr(self, "_font_scale_multiplier", 1.0))) < 1e-6:
            return
        self._font_scale_multiplier = float(multiplier)
        if self._rendered:
            try:
                self._apply_font_scale_visible()
            except Exception:
                pass
            try:
                self._apply_row_height_visible()
            except Exception:
                pass
            try:
                self.table.viewport().update()
            except Exception:
                pass

    def zoom_in(self) -> None:
        self.set_zoom(self._user_zoom * 1.1)

    def zoom_out(self) -> None:
        self.set_zoom(self._user_zoom / 1.1)

    def zoom_reset(self) -> None:
        # Reset to 100% *effective* scale (like Excel), not 1.0 multiplier.
        if self._rendered and self._fit_scale > 0:
            self.set_zoom(1.0 / self._fit_scale)
        else:
            self.set_zoom(1.0)

    def set_worksheet(self, ws) -> None:
        self._ws = ws
        if ws is not None:
            self._validation_lists, self._validation_display_to_value, self._validation_mapped_kind = self._build_validation_list_map(ws)
        else:
            self._validation_lists = {}
            self._validation_display_to_value = {}
            self._validation_mapped_kind = {}

    def refresh_validations(self) -> None:
        """Rebuild validation maps from the current worksheet."""
        if self._ws is None:
            self._validation_lists = {}
            self._validation_display_to_value = {}
            self._validation_mapped_kind = {}
            return
        self._validation_lists, self._validation_display_to_value, self._validation_mapped_kind = self._build_validation_list_map(self._ws)

    def set_overrides(self, overrides: Dict[BorderKey, CellOverride]) -> None:
        self._overrides = dict(overrides or {})

    def set_hidden_columns(self, cols_1based: list[int] | set[int] | tuple[int, ...]) -> None:
        self._hidden_columns_1based = {int(c) for c in (cols_1based or []) if int(c) > 0}
        if self._rendered:
            self._apply_hidden_columns()
            self._recompute_fit_scale()
            self._schedule_apply()

    def set_auto_fit_row_height_columns(self, cols_1based: list[int] | set[int] | tuple[int, ...]) -> None:
        """Auto-resize visible row heights to fit wrapped text in these columns.

        Columns are 1-based (Excel-style). Pass an empty list to disable.
        """
        cols = set()
        for c in (cols_1based or []):
            try:
                c = int(c)
            except Exception:
                continue
            if c > 0:
                cols.add(c - 1)
        self._auto_fit_row_height_cols = cols
        if self._rendered:
            self._apply_row_height_visible()

    def fit_columns_to_wrapped_text(self, min_col_1based: int = 1, max_col_1based: int = 20) -> None:
        """Auto-fit column widths based on wrapped text between columns.

        Only considers cells with WRAP_ROLE True (Excel Wrap Text).
        Columns are 1-based (Excel-style).
        """
        if not self._rendered:
            return
        if self.table.columnCount() <= 0 or self.table.rowCount() <= 0:
            return

        try:
            min_c = int(min_col_1based)
            max_c = int(max_col_1based)
        except Exception:
            return
        if min_c <= 0 or max_c <= 0:
            return

        min_c = max(1, min(min_c, self.table.columnCount()))
        max_c = max(1, min(max_c, self.table.columnCount()))
        if max_c < min_c:
            min_c, max_c = max_c, min_c

        effective = self._fit_scale * self._user_zoom
        effective = max(self.MIN_EFFECTIVE_SCALE, min(effective, self.MAX_EFFECTIVE_SCALE))
        if effective <= 0:
            effective = 1.0

        self._suppress_resize_handlers = True
        try:
            with QSignalBlocker(self.table.horizontalHeader()):
                for c1 in range(min_c, max_c + 1):
                    c0 = c1 - 1
                    if self.table.isColumnHidden(c0):
                        continue

                    max_w = 0
                    for r0 in range(self.table.rowCount()):
                        it = self.table.item(r0, c0)
                        if it is None:
                            continue
                        try:
                            wrap_flag = bool(it.data(self.WRAP_ROLE))
                        except Exception:
                            wrap_flag = False
                        if not wrap_flag:
                            continue

                        text = (it.text() or "").strip()
                        if not text:
                            continue

                        try:
                            font = it.font() if it is not None else self.table.font()
                        except Exception:
                            font = self.table.font()

                        try:
                            fm = QFontMetrics(font)
                        except Exception:
                            fm = None

                        lines = text.splitlines() if "\n" in text else [text]
                        for line in lines:
                            line = str(line)
                            if not line:
                                continue
                            if fm is None:
                                w = len(line) * 7
                            else:
                                try:
                                    w = int(fm.horizontalAdvance(line))
                                except Exception:
                                    w = int(fm.horizontalAdvance(line)) if hasattr(fm, "horizontalAdvance") else len(line) * 7
                            if w > max_w:
                                max_w = w

                    if max_w > 0:
                        desired_px = max(int(max_w) + 12, 20)
                        try:
                            self.table.setColumnWidth(c0, int(desired_px))
                        except Exception:
                            pass

                        while len(self._base_col_widths) <= c0:
                            self._base_col_widths.append(max(1, int(desired_px / effective)))
                        self._base_col_widths[c0] = max(1, int(int(desired_px) / effective))
        finally:
            self._suppress_resize_handlers = False

        try:
            self._persist_base_sizes()
        except Exception:
            self._schedule_persist()

    def fit_rows_to_wrapped_text(self, min_col_1based: int = 1, max_col_1based: int = 20) -> None:
        """Auto-fit row heights using wrapped text in a column range.

        Only considers cells with WRAP_ROLE True (Excel Wrap Text).
        Columns are 1-based (Excel-style).
        """
        if not self._rendered:
            return
        if self.table.columnCount() <= 0 or self.table.rowCount() <= 0:
            return

        try:
            min_c = int(min_col_1based)
            max_c = int(max_col_1based)
        except Exception:
            return
        if min_c <= 0 or max_c <= 0:
            return

        min_c = max(1, min(min_c, self.table.columnCount()))
        max_c = max(1, min(max_c, self.table.columnCount()))
        if max_c < min_c:
            min_c, max_c = max_c, min_c

        cols0 = {c - 1 for c in range(min_c, max_c + 1)}

        effective = self._fit_scale * self._user_zoom
        effective = max(self.MIN_EFFECTIVE_SCALE, min(effective, self.MAX_EFFECTIVE_SCALE))
        if effective <= 0:
            effective = 1.0

        self._suppress_resize_handlers = True
        try:
            with QSignalBlocker(self.table.verticalHeader()):
                for r0 in range(self.table.rowCount()):
                    desired_px = self._compute_row_height_px(int(r0), cols0)
                    if desired_px is None:
                        continue
                    try:
                        self.table.setRowHeight(int(r0), int(desired_px))
                    except Exception:
                        pass

                    if 0 <= int(r0) < len(self._base_row_heights):
                        self._base_row_heights[int(r0)] = max(self._base_row_heights[int(r0)], int(int(desired_px) / effective))
                    else:
                        while len(self._base_row_heights) <= int(r0):
                            self._base_row_heights.append(max(1, int(int(desired_px) / effective)))
                        self._base_row_heights[int(r0)] = max(1, int(int(desired_px) / effective))
        finally:
            self._suppress_resize_handlers = False

        try:
            self._persist_base_sizes()
        except Exception:
            self._schedule_persist()

    def set_row_height_pixels(self, row_1based: int, height_px: int) -> None:
        """Set a row height in pixels, updating base sizes + persistence.

        height_px is the desired on-screen height at the current effective scale.
        """
        try:
            row_1based = int(row_1based)
            height_px = int(height_px)
        except Exception:
            return
        if row_1based <= 0 or height_px <= 0:
            return

        effective = self._fit_scale * self._user_zoom
        effective = max(self.MIN_EFFECTIVE_SCALE, min(effective, self.MAX_EFFECTIVE_SCALE))
        if effective <= 0:
            effective = 1.0

        base_px = max(int(height_px / effective), 1)
        idx = row_1based - 1

        # Ensure base size list is long enough.
        while len(self._base_row_heights) <= idx:
            try:
                cur_px = int(self.table.rowHeight(len(self._base_row_heights)))
            except Exception:
                cur_px = base_px
            self._base_row_heights.append(max(int(cur_px / effective), 1))

        self._base_row_heights[idx] = max(int(base_px), 1)

        try:
            self.table.setRowHeight(idx, max(int(base_px * effective), 5))
        except Exception:
            pass

        # Persist immediately so a subsequent render doesn't restore stale sizes.
        try:
            self._persist_base_sizes()
        except Exception:
            self._schedule_persist()

    def undo(self) -> None:
        if not self._undo_stack:
            return
        cmd = self._undo_stack.pop()
        self._apply_command(cmd, forward=False)
        self._redo_stack.append(cmd)

    def redo(self) -> None:
        if not self._redo_stack:
            return
        cmd = self._redo_stack.pop()
        self._apply_command(cmd, forward=True)
        self._undo_stack.append(cmd)

    def set_wrap_text_for_selection(self, wrap: bool) -> bool:
        """Set Excel 'Wrap Text' for the current selection.

        Returns True if the operation applied to at least one cell.
        """
        if self._ws is None:
            return False

        ranges = list(self.table.selectedRanges() or [])
        if not ranges:
            return False

        ws = self._ws

        try:
            from openpyxl.styles import Alignment
            from openpyxl.cell.cell import MergedCell
        except Exception:
            return False

        # Collect target cells (1-based). Use a set to avoid duplicates.
        targets: set[tuple[int, int]] = set()
        for rng in ranges:
            r0 = int(rng.topRow()) + 1
            r1 = int(rng.bottomRow()) + 1
            c0 = int(rng.leftColumn()) + 1
            c1 = int(rng.rightColumn()) + 1
            for rr in range(r0, r1 + 1):
                for cc in range(c0, c1 + 1):
                    targets.add((rr, cc))

        def _top_left_for_merged(row: int, col: int) -> tuple[int, int]:
            try:
                for merged in list(getattr(ws.merged_cells, "ranges", []) or []):
                    if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
                        return int(merged.min_row), int(merged.min_col)
            except Exception:
                pass
            return row, col

        applied = False
        for rr, cc in sorted(targets):
            try:
                cell = ws.cell(row=rr, column=cc)
            except Exception:
                continue

            if isinstance(cell, MergedCell) or cell.__class__.__name__ == "MergedCell":
                rr2, cc2 = _top_left_for_merged(rr, cc)
                try:
                    cell = ws.cell(row=rr2, column=cc2)
                except Exception:
                    continue

            try:
                cur = getattr(cell, "alignment", None)
                if cur is not None:
                    try:
                        cell.alignment = cur.copy(wrapText=bool(wrap))
                    except Exception:
                        cell.alignment = Alignment(wrapText=bool(wrap))
                else:
                    cell.alignment = Alignment(wrapText=bool(wrap))
                applied = True
            except Exception:
                continue

            # Update visible item state.
            try:
                it = self.table.item(int(rr) - 1, int(cc) - 1)
                if it is not None:
                    it.setData(self.WRAP_ROLE, bool(wrap))
            except Exception:
                pass

        if applied:
            # Excel-like behavior: wrap changes should expand/shrink row heights.
            try:
                self.table.resizeRowsToContents()
            except Exception:
                pass

        return applied

    def render(self) -> None:
        if self._ws is None:
            self.table.setRowCount(0)
            self.table.setColumnCount(0)
            self._rendered = False
            return

        ws = self._ws

        # Determine used area, including merges.
        cell_coords = list(getattr(ws, "_cells", {}).keys())  # (row, col)
        max_row = 1
        max_col = 1
        if cell_coords:
            max_row = max(r for r, _c in cell_coords)
            max_col = max(c for _r, c in cell_coords)

        for merged in list(getattr(ws.merged_cells, "ranges", [])):
            max_row = max(max_row, merged.max_row)
            max_col = max(max_col, merged.max_col)

        self.table.clear()
        self.table.setRowCount(max_row)
        self.table.setColumnCount(max_col)

        # Excel-style header labels
        try:
            from openpyxl.utils import get_column_letter

            self.table.setHorizontalHeaderLabels([get_column_letter(i) for i in range(1, max_col + 1)])
        except Exception:
            pass
        try:
            self.table.setVerticalHeaderLabels([str(i) for i in range(1, max_row + 1)])
        except Exception:
            pass

        # Capture base sizes so we can scale them later.
        self._base_row_heights = [self.table.rowHeight(r) for r in range(max_row)]
        self._base_col_widths = [self.table.columnWidth(c) for c in range(max_col)]
        self._last_effective_scale = 1.0
        self._fit_scale = 1.0

        # Apply row/col sizing (best-effort) without triggering our resize handlers.
        with QSignalBlocker(self.table.horizontalHeader()), QSignalBlocker(self.table.verticalHeader()):
            for r in range(1, max_row + 1):
                h = getattr(ws.row_dimensions.get(r), "height", None)
                if h:
                    # Excel row height is in points.
                    px = int(float(h) * 96 / 72)
                    self.table.setRowHeight(r - 1, max(px, 10))

            # Column width is in "character" units; approximate to pixels.
            for c in range(1, max_col + 1):
                letter = get_column_letter(c)
                w = getattr(ws.column_dimensions.get(letter), "width", None)
                if w:
                    px = int(float(w) * 7 + 10)
                    self.table.setColumnWidth(c - 1, max(px, 20))

        # Refresh base sizes after applying worksheet dimensions.
        self._base_row_heights = [self.table.rowHeight(r) for r in range(max_row)]
        self._base_col_widths = [self.table.columnWidth(c) for c in range(max_col)]

        # Apply any requested column hiding before fitting/scaling.
        self._apply_hidden_columns()

        # If persistence is enabled, restore user base sizes before fitting.
        # This ensures fit-to-width uses the last user-resized layout.
        self._restore_base_sizes()

        # Build merged spans.
        merged_map: Dict[Tuple[int, int], Tuple[int, int]] = {}  # top-left -> (rowSpan, colSpan)
        covered: set[BorderKey] = set()
        for merged in list(getattr(ws.merged_cells, "ranges", [])):
            tl = (merged.min_row, merged.min_col)
            merged_map[tl] = (merged.max_row - merged.min_row + 1, merged.max_col - merged.min_col + 1)
            for rr in range(merged.min_row, merged.max_row + 1):
                for cc in range(merged.min_col, merged.max_col + 1):
                    if (rr, cc) != tl:
                        covered.add((rr, cc))

        self._covered_cells = covered

        for (r, c), (rs, cs) in merged_map.items():
            # QTableWidget uses 0-based indices
            self.table.setSpan(r - 1, c - 1, rs, cs)

        def _qcolor_from_openpyxl(color_obj) -> Optional[QColor]:
            if color_obj is None:
                return None
            rgb = getattr(color_obj, "rgb", None)
            if not rgb:
                return None
            rgb = str(rgb)
            # openpyxl often uses ARGB
            if len(rgb) == 8:
                rgb = rgb[2:]
            if len(rgb) != 6:
                return None
            try:
                r = int(rgb[0:2], 16)
                g = int(rgb[2:4], 16)
                b = int(rgb[4:6], 16)
                return QColor(r, g, b)
            except Exception:
                return None

        def _ideal_text_color(bg: Optional[QColor]) -> Optional[QColor]:
            """Pick a readable text color when background is explicit."""
            if bg is None:
                return None
            # Use perceived lightness to select black/white.
            # QColor.lightness() is 0..255.
            return QColor(0, 0, 0) if bg.lightness() >= 140 else QColor(255, 255, 255)

        def _border_spec(side) -> Dict[str, Any]:
            style = getattr(side, "style", None)
            if not style or style == "none":
                return {"width": 0, "color": None}
            width = 1
            if style in {"medium"}:
                width = 2
            elif style in {"thick"}:
                width = 3
            color = _qcolor_from_openpyxl(getattr(side, "color", None)) or QColor(0, 0, 0)
            return {"width": width, "color": color}

        _fmt_decimal_re = re.compile(r"^[#0]+\.([0]+)$")

        def _format_display_value(raw: Any, number_format: Any) -> str:
            if raw is None:
                return ""
            # bool is a subclass of int; avoid rendering booleans as 1.0000.
            if isinstance(raw, bool):
                return str(raw)

            fmt = str(number_format or "").strip()
            if isinstance(raw, (int, float)) and fmt:
                try:
                    m = _fmt_decimal_re.match(fmt)
                    if m:
                        decimals = len(m.group(1))
                        return f"{float(raw):.{decimals}f}"
                except Exception:
                    pass

            return str(raw)

        # Populate items.
        self._in_programmatic_change = True
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                if (r, c) in covered:
                    continue

                cell = ws.cell(row=r, column=c)

                # Value
                raw_value = getattr(cell, "value", None)
                display_value = _format_display_value(raw_value, getattr(cell, "number_format", None))

                override = self._overrides.get((r, c))
                if override is not None and override.value is not None:
                    display_value = str(override.value)

                item = QTableWidgetItem(display_value)
                # Allow edits for normal cells. Covered merged cells are not created.
                item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
                # Store 1-based coordinate for later mapping.
                item.setData(Qt.ItemDataRole.UserRole + 1, (r, c))

                # Alignment
                align = getattr(cell, "alignment", None)
                if align:
                    h = getattr(align, "horizontal", None)
                    v = getattr(align, "vertical", None)

                    # Per-cell wrap flag (Excel Wrap Text).
                    try:
                        item.setData(self.WRAP_ROLE, bool(getattr(align, "wrapText", False)))
                    except Exception:
                        pass

                    qt_h = {
                        "left": Qt.AlignmentFlag.AlignLeft,
                        "center": Qt.AlignmentFlag.AlignHCenter,
                        "right": Qt.AlignmentFlag.AlignRight,
                        "justify": Qt.AlignmentFlag.AlignJustify,
                    }.get(h, Qt.AlignmentFlag.AlignLeft)

                    qt_v = {
                        "top": Qt.AlignmentFlag.AlignTop,
                        "center": Qt.AlignmentFlag.AlignVCenter,
                        "bottom": Qt.AlignmentFlag.AlignBottom,
                    }.get(v, Qt.AlignmentFlag.AlignVCenter)

                    item.setTextAlignment(int(qt_h | qt_v))
                else:
                    try:
                        item.setData(self.WRAP_ROLE, False)
                    except Exception:
                        pass

                # Font
                font = getattr(cell, "font", None)
                if font:
                    qfont = QFont()
                    if getattr(font, "name", None):
                        qfont.setFamily(str(font.name))
                    if getattr(font, "sz", None):
                        qfont.setPointSizeF(float(font.sz))
                    qfont.setBold(bool(getattr(font, "b", False)))
                    qfont.setItalic(bool(getattr(font, "i", False)))
                    item.setFont(qfont)

                # Record a base font size for scaling (use explicit cell font size when present).
                base_size = None
                if font and getattr(font, "sz", None):
                    try:
                        base_size = float(font.sz)
                    except Exception:
                        base_size = None
                if base_size is None:
                    # Some styles return 0/-1 point size; keep a sane default.
                    ps = float(self.table.font().pointSizeF())
                    base_size = ps if ps > 0 else 9.0
                item.setData(self.BASE_FONT_SIZE_ROLE, base_size)

                # Fill
                bg = None
                fill = getattr(cell, "fill", None)
                if fill and getattr(fill, "patternType", None) == "solid":
                    bg = _qcolor_from_openpyxl(getattr(fill, "fgColor", None))

                if override is not None and override.background is not None:
                    bg = override.background

                if bg is not None:
                    item.setBackground(bg)

                # Font color (important for light/dark mode correctness)
                # Use explicit Excel font color if present; otherwise, if the
                # cell has an explicit background, choose a contrasting color.
                fg = None
                if font and getattr(font, "color", None) is not None:
                    fg = _qcolor_from_openpyxl(getattr(font, "color", None))

                if fg is None:
                    fg = _ideal_text_color(bg)

                if fg is not None:
                    item.setForeground(fg)

                # Borders
                border = getattr(cell, "border", None)
                if border:
                    borders = {
                        "top": _border_spec(getattr(border, "top", None)),
                        "bottom": _border_spec(getattr(border, "bottom", None)),
                        "left": _border_spec(getattr(border, "left", None)),
                        "right": _border_spec(getattr(border, "right", None)),
                    }
                    item.setData(_ExcelBorderDelegate.BORDER_ROLE, borders)

                self.table.setItem(r - 1, c - 1, item)

            self._in_programmatic_change = False

        self._rendered = True
        # Defer fit until after layout/viewport sizing has settled.
        QTimer.singleShot(0, self._post_render_fit)

    def _apply_hidden_columns(self) -> None:
        if not self._hidden_columns_1based:
            return
        # QTableWidget uses 0-based columns.
        for c1 in self._hidden_columns_1based:
            idx = c1 - 1
            if 0 <= idx < self.table.columnCount():
                self.table.setColumnHidden(idx, True)

    def _post_render_fit(self) -> None:
        if not self._rendered:
            return
        self._recompute_fit_scale()
        if self._pending_effective_scale is not None:
            pending = self._pending_effective_scale
            self._pending_effective_scale = None
            self.set_effective_scale(pending)
        self._schedule_apply()

    def resizeEvent(self, event):
        super().resizeEvent(event)
        # If starting scale is locked, do not auto-recompute scale on resize.
        if self._lock_effective_scale:
            return
        if self._rendered:
            self._recompute_fit_scale()
            self._schedule_apply()

    def eventFilter(self, obj, event):
        # Zoom shortcuts should work regardless of focus child.
        if obj in (self.table.viewport(), self.table):
            if event.type() == QEvent.Type.MouseButtonPress:
                try:
                    if event.button() == Qt.MouseButton.LeftButton:
                        mods = event.modifiers()
                        if mods == Qt.KeyboardModifier.NoModifier:
                            # If the clicked cell has a list validation, treat
                            # it like a dropdown on click.
                            pos = event.position().toPoint() if hasattr(event, "position") else event.pos()
                            index = self.table.indexAt(pos) if obj == self.table.viewport() else self.table.currentIndex()
                            if index.isValid():
                                r0 = index.row()
                                c0 = index.column()
                                if (r0 + 1, c0 + 1) in self._validation_lists and (r0 + 1, c0 + 1) not in self._covered_cells:
                                    def _start_dropdown():
                                        item = self.table.item(r0, c0)
                                        if item is None:
                                            return
                                        self.table.setCurrentCell(r0, c0)
                                        self.table.editItem(item)

                                    QTimer.singleShot(0, _start_dropdown)
                except Exception:
                    pass
            if event.type() == QEvent.Type.Wheel:
                modifiers = event.modifiers()
                if modifiers & Qt.KeyboardModifier.ControlModifier:
                    # Ctrl+Wheel zoom
                    delta = event.angleDelta().y()
                    if delta > 0:
                        self.zoom_in()
                    elif delta < 0:
                        self.zoom_out()
                    event.accept()
                    return True

                # Optional wheel navigation override (used by Form 3 bubble follow).
                try:
                    handler = getattr(self, "_wheel_navigation_handler", None)
                    if handler is not None:
                        try:
                            cur_r0 = int(self.table.currentRow())
                            cur_c0 = int(self.table.currentColumn())
                        except Exception:
                            cur_r0 = -1
                            cur_c0 = -1
                        if cur_r0 >= 0 and cur_c0 >= 0:
                            dy = 0
                            try:
                                dy = int(event.angleDelta().y())
                            except Exception:
                                dy = 0
                            try:
                                consumed = bool(handler(int(cur_r0) + 1, int(cur_c0) + 1, int(dy)))
                            except Exception:
                                consumed = False
                            if consumed:
                                event.accept()
                                return True
                except Exception:
                    pass
            elif event.type() == QEvent.Type.KeyPress:
                modifiers = event.modifiers()
                key = event.key()

                try:
                    if event.matches(QKeySequence.Undo):
                        self._on_undo_shortcut()
                        event.accept()
                        return True
                except Exception:
                    pass

                # Excel-like Enter behavior: move down/up a row.
                # - When NOT editing: handle immediately.
                # - When editing: let the editor commit/close, then move.
                if key in (Qt.Key.Key_Return, Qt.Key.Key_Enter) and modifiers in (
                    Qt.KeyboardModifier.NoModifier,
                    Qt.KeyboardModifier.ShiftModifier,
                ):
                    direction = -1 if modifiers == Qt.KeyboardModifier.ShiftModifier else 1
                    try:
                        if self.table.state() == QAbstractItemView.State.EditingState:
                            self._pending_enter_nav_dir = int(direction)
                            return super().eventFilter(obj, event)
                    except Exception:
                        pass

                    self._move_current_cell_row(int(direction))
                    event.accept()
                    return True

                # While editing a cell, let the editor handle clipboard shortcuts.
                try:
                    if self.table.state() == QAbstractItemView.State.EditingState:
                        return super().eventFilter(obj, event)
                except Exception:
                    pass

                # Delete clears contents of selected cell(s) (Excel-like).
                if key == Qt.Key.Key_Delete and modifiers == Qt.KeyboardModifier.NoModifier:
                    self._clear_selection_contents()
                    event.accept()
                    return True

                # Optional Up/Down navigation override (used by Form 3 bubble follow).
                if modifiers == Qt.KeyboardModifier.NoModifier and key in (Qt.Key.Key_Up, Qt.Key.Key_Down):
                    try:
                        # Don't interfere while editing a cell.
                        try:
                            if self.table.state() == QAbstractItemView.State.EditingState:
                                return super().eventFilter(obj, event)
                        except Exception:
                            pass

                        handler = getattr(self, "_key_navigation_handler", None)
                        if handler is not None:
                            try:
                                cur_r0 = int(self.table.currentRow())
                                cur_c0 = int(self.table.currentColumn())
                            except Exception:
                                cur_r0 = -1
                                cur_c0 = -1
                            if cur_r0 >= 0 and cur_c0 >= 0:
                                direction = 1 if key == Qt.Key.Key_Down else -1
                                try:
                                    consumed = bool(handler(int(cur_r0) + 1, int(cur_c0) + 1, int(direction)))
                                except Exception:
                                    consumed = False
                                if consumed:
                                    event.accept()
                                    return True
                    except Exception:
                        pass

                # ESC clears click-to-paint mode.
                if key == Qt.Key.Key_Escape and modifiers == Qt.KeyboardModifier.NoModifier:
                    self.clear_click_paint_mode()
                    event.accept()
                    return True

                # Undo/redo
                if modifiers & Qt.KeyboardModifier.ControlModifier and key == Qt.Key.Key_Z:
                    self.undo()
                    event.accept()
                    return True
                if modifiers & Qt.KeyboardModifier.ControlModifier and key == Qt.Key.Key_Y:
                    self.redo()
                    event.accept()
                    return True

                # Copy/Cut/Paste
                if modifiers & Qt.KeyboardModifier.ControlModifier and key == Qt.Key.Key_C:
                    self._copy_selection()
                    event.accept()
                    return True
                if modifiers & Qt.KeyboardModifier.ControlModifier and key == Qt.Key.Key_X:
                    self._cut_selection()
                    event.accept()
                    return True
                if modifiers & Qt.KeyboardModifier.ControlModifier and key == Qt.Key.Key_V:
                    self._paste_clipboard()
                    event.accept()
                    return True

                # Fill down / fill right
                if modifiers & Qt.KeyboardModifier.ControlModifier and key == Qt.Key.Key_D:
                    self._fill_down()
                    event.accept()
                    return True
                if modifiers & Qt.KeyboardModifier.ControlModifier and key == Qt.Key.Key_R:
                    self._fill_right()
                    event.accept()
                    return True

                if modifiers & Qt.KeyboardModifier.ControlModifier:
                    if key in (Qt.Key.Key_Plus, Qt.Key.Key_Equal):
                        self.zoom_in()
                        event.accept()
                        return True
                    if key == Qt.Key.Key_Minus:
                        self.zoom_out()
                        event.accept()
                        return True
                    if key == Qt.Key.Key_0:
                        self.zoom_reset()
                        event.accept()
                        return True

        # Excel-like header selection (rows/cols) with Ctrl/Shift multi-select.
        if obj in (self.table.verticalHeader(), self.table.horizontalHeader()):
            try:
                if event.type() == QEvent.Type.MouseButtonPress and event.button() == Qt.MouseButton.LeftButton:
                    mods = event.modifiers()
                    pos = event.position().toPoint() if hasattr(event, "position") else event.pos()
                    idx = int(obj.logicalIndexAt(pos))
                    if idx >= 0:
                        if obj == self.table.verticalHeader():
                            self._header_select_rows(idx, mods)
                        else:
                            self._header_select_cols(idx, mods)
                        event.accept()
                        return True
                if event.type() == QEvent.Type.MouseButtonDblClick and event.button() == Qt.MouseButton.LeftButton:
                    # Double-click row header divider (or header) to auto-fit selected rows.
                    if obj == self.table.verticalHeader():
                        pos = event.position().toPoint() if hasattr(event, "position") else event.pos()
                        idx = int(obj.logicalIndexAt(pos))
                        rows0 = set()
                        try:
                            rows0 = {int(ix.row()) for ix in (self.table.selectionModel().selectedRows() or [])}
                        except Exception:
                            rows0 = set()
                        if not rows0:
                            try:
                                rows0 = {int(ix.row()) for ix in (self.table.selectedIndexes() or [])}
                            except Exception:
                                rows0 = set()
                        if not rows0 and idx >= 0:
                            rows0 = {int(idx)}
                        if rows0:
                            self._auto_fit_rows(set(rows0), force=True)
                            event.accept()
                            return True
            except Exception:
                pass
        return super().eventFilter(obj, event)

    def _header_select_rows(self, row0: int, modifiers) -> None:
        if row0 < 0 or row0 >= self.table.rowCount():
            return
        if self.table.columnCount() <= 0:
            return

        ctrl = bool(modifiers & Qt.KeyboardModifier.ControlModifier)
        shift = bool(modifiers & Qt.KeyboardModifier.ShiftModifier)

        if not shift or self._row_header_anchor0 is None:
            self._row_header_anchor0 = int(row0)

        start = int(self._row_header_anchor0 if shift and self._row_header_anchor0 is not None else row0)
        end = int(row0)
        r0 = min(start, end)
        r1 = max(start, end)

        model = self.table.model()
        if model is None:
            return

        sel = QItemSelection(model.index(r0, 0), model.index(r1, self.table.columnCount() - 1))
        flags = QItemSelectionModel.SelectionFlag.Select
        if not ctrl:
            flags |= QItemSelectionModel.SelectionFlag.Clear
        self.table.selectionModel().select(sel, flags)
        try:
            self.table.setCurrentCell(int(row0), int(self.table.currentColumn() if self.table.currentColumn() >= 0 else 0))
        except Exception:
            pass

    def _header_select_cols(self, col0: int, modifiers) -> None:
        if col0 < 0 or col0 >= self.table.columnCount():
            return
        if self.table.rowCount() <= 0:
            return

        ctrl = bool(modifiers & Qt.KeyboardModifier.ControlModifier)
        shift = bool(modifiers & Qt.KeyboardModifier.ShiftModifier)

        if not shift or self._col_header_anchor0 is None:
            self._col_header_anchor0 = int(col0)

        start = int(self._col_header_anchor0 if shift and self._col_header_anchor0 is not None else col0)
        end = int(col0)
        c0 = min(start, end)
        c1 = max(start, end)

        model = self.table.model()
        if model is None:
            return

        sel = QItemSelection(model.index(0, c0), model.index(self.table.rowCount() - 1, c1))
        flags = QItemSelectionModel.SelectionFlag.Select
        if not ctrl:
            flags |= QItemSelectionModel.SelectionFlag.Clear
        self.table.selectionModel().select(sel, flags)
        try:
            self.table.setCurrentCell(int(self.table.currentRow() if self.table.currentRow() >= 0 else 0), int(col0))
        except Exception:
            pass

    def _clear_selection_contents(self) -> None:
        """Clear the contents of all selected cells (without touching clipboard)."""
        selected = self._selected_cells0()
        if not selected:
            return

        updates: dict[BorderKey, Any] = {}
        changed_rows0: set[int] = set()
        for r0, c0 in sorted(selected):
            try:
                it = self.table.item(int(r0), int(c0))
                coord = it.data(Qt.ItemDataRole.UserRole + 1) if it is not None else None
            except Exception:
                coord = None
            if not coord:
                continue
            try:
                r1, c1 = int(coord[0]), int(coord[1])
            except Exception:
                continue
            if (r1, c1) in self._covered_cells:
                continue
            updates[(r1, c1)] = None
            if (c1 - 1) in self._auto_fit_row_height_cols:
                changed_rows0.add(r1 - 1)

        if updates:
            self._batch_set_cells(updates)
            self._auto_fit_rows(changed_rows0)

    def _on_close_editor(self, _editor=None, _hint=None) -> None:
        """Called when a cell editor closes; used for Excel-like Enter navigation."""
        try:
            direction = self._pending_enter_nav_dir
            self._pending_enter_nav_dir = None
        except Exception:
            direction = None
            self._pending_enter_nav_dir = None

        if direction not in (-1, 1):
            return
        # Defer so the model/selection state is fully updated.
        QTimer.singleShot(0, lambda d=int(direction): self._move_current_cell_row(d))

    def _move_current_cell_row(self, direction: int) -> None:
        """Move the current cell up/down one row and select it (Excel-like)."""
        try:
            direction = int(direction)
        except Exception:
            return
        if direction not in (-1, 1):
            return

        try:
            r0 = int(self.table.currentRow())
            c0 = int(self.table.currentColumn())
        except Exception:
            return
        if r0 < 0 or c0 < 0:
            return

        try:
            max_r = int(self.table.rowCount()) - 1
            max_c = int(self.table.columnCount()) - 1
        except Exception:
            return
        if max_r < 0 or max_c < 0:
            return

        nr = max(0, min(r0 + direction, max_r))
        nc = max(0, min(c0, max_c))

        try:
            self.table.clearSelection()
        except Exception:
            pass
        try:
            self.table.setCurrentCell(int(nr), int(nc))
        except Exception:
            return
        try:
            self.table.setRangeSelected(QTableWidgetSelectionRange(int(nr), int(nc), int(nr), int(nc)), True)
        except Exception:
            pass
        try:
            it = self.table.item(int(nr), int(nc))
            if it is not None:
                self.table.scrollToItem(it, QAbstractItemView.ScrollHint.EnsureVisible)
        except Exception:
            pass

    def _schedule_apply(self) -> None:
        # Coalesce rapid resize/wheel events.
        self._apply_timer.start()

    def _recompute_fit_scale(self) -> None:
        # Available area inside the table viewport.
        viewport = self.table.viewport()
        avail_w = max(viewport.width() - 2, 1)
        avail_h = max(viewport.height() - 2, 1)

        if not self._base_col_widths or not self._base_row_heights:
            return

        total_w = sum(
            w
            for i, w in enumerate(self._base_col_widths)
            if not self.table.isColumnHidden(i)
        )
        total_h = sum(self._base_row_heights)
        if total_w <= 0 or total_h <= 0:
            return

        scale_w = avail_w / total_w
        scale_h = avail_h / total_h
        if self._fit_mode == "both":
            self._fit_scale = min(scale_w, scale_h)
        elif self._fit_mode == "width":
            self._fit_scale = scale_w
        else:
            self._fit_scale = 1.0

        # If the user locked the effective scale, keep it fixed by adjusting zoom.
        if self._lock_effective_scale and self._locked_effective_scale is not None and self._fit_scale > 0:
            desired_zoom = self._locked_effective_scale / self._fit_scale
            self._user_zoom = max(self.MIN_USER_ZOOM, min(desired_zoom, self.MAX_USER_ZOOM))
            return

        # Keep current zoom within the effective [50%, 200%] range.
        # If a sheet is huge and fit_scale is tiny, this will increase user_zoom
        # so the effective scale doesn't get stuck at very small percentages.
        if self._fit_scale > 0:
            min_zoom = max(self.MIN_EFFECTIVE_SCALE / self._fit_scale, self.MIN_USER_ZOOM)
            max_zoom = min(self.MAX_EFFECTIVE_SCALE / self._fit_scale, self.MAX_USER_ZOOM)
            self._user_zoom = max(min_zoom, min(self._user_zoom, max_zoom))

    def _apply_scale(self) -> None:
        if not self._base_col_widths or not self._base_row_heights:
            return

        effective = self._fit_scale * self._user_zoom
        effective = max(self.MIN_EFFECTIVE_SCALE, min(effective, self.MAX_EFFECTIVE_SCALE))

        # Avoid thrashing on tiny scale changes.
        if abs(effective - self._last_effective_scale) < 0.01:
            # Still emit occasionally so UI stays in sync.
            self.scaleChanged.emit(float(effective))
            return

        # Avoid our own sectionResized handlers treating programmatic scaling as user resize.
        with QSignalBlocker(self.table.horizontalHeader()), QSignalBlocker(self.table.verticalHeader()):
            for i, base_w in enumerate(self._base_col_widths):
                if self.table.isColumnHidden(i):
                    continue
                self.table.setColumnWidth(i, max(int(base_w * effective), 5))
            for i, base_h in enumerate(self._base_row_heights):
                self.table.setRowHeight(i, max(int(base_h * effective), 5))

            # Post-correct for fit-to-width: use actual rendered width and adjust once.
            # If the user requested a locked starting scale, do not override it.
            if self._fit_mode == "width" and not self._lock_effective_scale:
                viewport = self.table.viewport()
                target_w = max(viewport.width() - 2, 1)
                actual_w = 0
                for i in range(self.table.columnCount()):
                    if self.table.isColumnHidden(i):
                        continue
                    actual_w += self.table.columnWidth(i)
                if actual_w > 0:
                    correction = target_w / float(actual_w)
                    if abs(1.0 - correction) >= 0.02:
                        desired = effective * correction
                        desired = max(self.MIN_EFFECTIVE_SCALE, min(desired, self.MAX_EFFECTIVE_SCALE))
                        if abs(desired - effective) >= 0.01:
                            effective = desired
                            for i, base_w in enumerate(self._base_col_widths):
                                if self.table.isColumnHidden(i):
                                    continue
                                self.table.setColumnWidth(i, max(int(base_w * effective), 5))
                            for i, base_h in enumerate(self._base_row_heights):
                                self.table.setRowHeight(i, max(int(base_h * effective), 5))

        # Update last applied effective scale after any correction.
        self._last_effective_scale = effective

        # Scale fonts for visible cells only (big speedup for large sheets).
        self._apply_font_scale_visible()

        # Auto-fit visible row heights (optional).
        self._apply_row_height_visible()

        # Notify UI of current scale.
        self.scaleChanged.emit(float(effective))

    def _apply_font_scale_visible(self) -> None:
        if not self._rendered:
            return
        if not self._base_col_widths or not self._base_row_heights:
            return

        effective = self._fit_scale * self._user_zoom
        effective = max(self.MIN_EFFECTIVE_SCALE, min(effective, self.MAX_EFFECTIVE_SCALE))

        try:
            effective_font = float(effective) * float(self._font_scale_multiplier)
        except Exception:
            effective_font = float(effective)

        vp = self.table.viewport()
        tl = self.table.indexAt(QPoint(0, 0))
        br = self.table.indexAt(QPoint(max(vp.width() - 1, 0), max(vp.height() - 1, 0)))
        if not tl.isValid() or not br.isValid():
            return

        r0, r1 = tl.row(), br.row()
        c0, c1 = tl.column(), br.column()
        if r0 < 0 or c0 < 0:
            return

        # Clamp to bounds
        r0 = max(0, min(r0, self.table.rowCount() - 1))
        r1 = max(0, min(r1, self.table.rowCount() - 1))
        c0 = max(0, min(c0, self.table.columnCount() - 1))
        c1 = max(0, min(c1, self.table.columnCount() - 1))

        for r in range(r0, r1 + 1):
            for c in range(c0, c1 + 1):
                item = self.table.item(r, c)
                if item is None:
                    continue
                base_size = item.data(self.BASE_FONT_SIZE_ROLE)
                if not base_size:
                    continue
                try:
                    base_size = float(base_size)
                except Exception:
                    continue
                f = item.font()
                f.setPointSizeF(max(base_size * effective_font, 6.0))
                item.setFont(f)

    def _apply_row_height_visible(self) -> None:
        """Resize visible rows to contents for configured wrap columns."""
        if not self._rendered:
            return
        if not self._auto_fit_row_height_cols:
            return

        vp = self.table.viewport()
        tl = self.table.indexAt(QPoint(0, 0))
        br = self.table.indexAt(QPoint(max(vp.width() - 1, 0), max(vp.height() - 1, 0)))
        if not tl.isValid() or not br.isValid():
            return

        r0, r1 = tl.row(), br.row()
        if r0 < 0:
            return
        r0 = max(0, min(r0, self.table.rowCount() - 1))
        r1 = max(0, min(r1, self.table.rowCount() - 1))

        # Only resize rows that actually have text in the configured columns.
        self._suppress_resize_handlers = True
        try:
            with QSignalBlocker(self.table.verticalHeader()):
                for r in range(r0, r1 + 1):
                    desired_px = self._compute_row_height_px(r, self._auto_fit_row_height_cols)
                    if desired_px is None:
                        continue
                    before_h = self.table.rowHeight(r)
                    if desired_px > before_h:
                        self.table.setRowHeight(r, int(desired_px))
                        # Persist growth into base sizes so scaling keeps it.
                        effective = self._fit_scale * self._user_zoom
                        effective = max(self.MIN_EFFECTIVE_SCALE, min(effective, self.MAX_EFFECTIVE_SCALE))
                        if effective <= 0:
                            effective = 1.0
                        if 0 <= r < len(self._base_row_heights):
                            self._base_row_heights[r] = max(self._base_row_heights[r], int(int(desired_px) / effective))
        finally:
            self._suppress_resize_handlers = False

    def _auto_fit_rows(self, rows0: set[int], *, force: bool = False) -> None:
        """Resize given 0-based rows to contents and persist as base row heights."""
        if not self._rendered:
            return
        if not self._auto_fit_row_height_cols and not force:
            return
        if not rows0:
            return

        effective = self._fit_scale * self._user_zoom
        effective = max(self.MIN_EFFECTIVE_SCALE, min(effective, self.MAX_EFFECTIVE_SCALE))
        if effective <= 0:
            effective = 1.0

        rows_sorted = sorted({r for r in rows0 if 0 <= r < self.table.rowCount()})
        if not rows_sorted:
            return

        # If no auto-fit columns are configured, consider all columns.
        if self._auto_fit_row_height_cols:
            cols0 = set(self._auto_fit_row_height_cols)
        else:
            cols0 = set(range(self.table.columnCount()))

        self._suppress_resize_handlers = True
        try:
            with QSignalBlocker(self.table.verticalHeader()):
                for r in rows_sorted:
                    desired_px = self._compute_row_height_px(r, cols0)
                    if desired_px is None:
                        continue
                    before_h = self.table.rowHeight(r)
                    if force:
                        self.table.setRowHeight(r, int(desired_px))
                    else:
                        if desired_px > before_h:
                            self.table.setRowHeight(r, int(desired_px))
                    if 0 <= r < len(self._base_row_heights):
                        self._base_row_heights[r] = max(self._base_row_heights[r], int(int(desired_px) / effective))
        finally:
            self._suppress_resize_handlers = False

    def _compute_row_height_px(self, row0: int, cols0: set[int]) -> int | None:
        """Compute desired row height based on cell text and wrap flags.

        Returns pixel height at current scale, or None if no text was found.
        """
        if row0 < 0 or row0 >= self.table.rowCount():
            return None
        if not cols0:
            return None

        max_h = 0
        saw_text = False

        for c0 in cols0:
            if c0 < 0 or c0 >= self.table.columnCount():
                continue
            if self.table.isColumnHidden(int(c0)):
                continue
            it = self.table.item(int(row0), int(c0))
            if it is None:
                continue
            text = it.text() or ""
            if not text.strip():
                continue
            saw_text = True

            try:
                font = it.font() if it is not None else self.table.font()
            except Exception:
                font = self.table.font()

            try:
                fm = QFontMetrics(font)
                base_h = int(fm.height()) + 6
            except Exception:
                fm = None
                base_h = 18

            try:
                wrap_flag = bool(it.data(self.WRAP_ROLE))
            except Exception:
                wrap_flag = False
            try:
                if not wrap_flag and int(c0) in self._auto_fit_row_height_cols:
                    wrap_flag = True
            except Exception:
                pass

            if wrap_flag:
                try:
                    col_w = int(self.table.columnWidth(int(c0)))
                except Exception:
                    col_w = 0
                padding = 8
                avail = max(int(col_w) - padding, 20)
                if fm is not None and avail > 0:
                    try:
                        rect = fm.boundingRect(QRect(0, 0, int(avail), 10000), Qt.TextWordWrap, str(text))
                        h = int(rect.height()) + 6
                    except Exception:
                        h = base_h
                else:
                    h = base_h
            else:
                h = base_h

            if h > max_h:
                max_h = int(h)

        if not saw_text:
            return None
        return max(int(max_h), 10)

    def _on_item_changed(self, item: QTableWidgetItem) -> None:
        if self._in_programmatic_change:
            return
        if self._ws is None:
            return

        coord = item.data(Qt.ItemDataRole.UserRole + 1)
        if not coord:
            return
        r, c = coord
        if (r, c) in self._covered_cells:
            return

        new_text = item.text()
        cell = self._ws.cell(row=r, column=c)
        old_value = cell.value
        new_value = self._coerce_text_value(new_text)
        if old_value == new_value:
            return

        cell.value = new_value
        self._undo_stack.append({(r, c): (old_value, new_value)})
        self._redo_stack.clear()
        self.modified.emit()

    def _coerce_text_value(self, text: str) -> Any:
        t = (text or "")
        if t == "":
            return None
        # Preserve formulas
        if t.startswith("="):
            return t
        # Try int/float
        try:
            if t.strip().isdigit() or (t.strip().startswith("-") and t.strip()[1:].isdigit()):
                return int(t.strip())
        except Exception:
            pass
        try:
            return float(t.strip())
        except Exception:
            return t

    def _apply_command(self, cmd: dict[BorderKey, tuple[Any, Any]], forward: bool) -> None:
        if self._ws is None:
            return
        self._in_programmatic_change = True
        try:
            for (r, c), (old, new) in cmd.items():
                value = new if forward else old

                # Support both value edits and fill edits in the undo stack.
                if isinstance(value, dict) and "fill_rgb" in value:
                    try:
                        self._apply_cell_fill(int(r), int(c), value.get("fill_rgb"), push_undo=False)
                    except Exception:
                        pass
                    continue

                self._ws.cell(row=r, column=c).value = value
                it = self.table.item(r - 1, c - 1)
                if it is not None:
                    it.setText("" if value is None else str(value))
            self.modified.emit()
        finally:
            self._in_programmatic_change = False

    def _selected_cells0(self) -> set[tuple[int, int]]:
        """Return selected cell coordinates as (row0, col0)."""
        out: set[tuple[int, int]] = set()
        try:
            for ix in (self.table.selectedIndexes() or []):
                try:
                    out.add((int(ix.row()), int(ix.column())))
                except Exception:
                    continue
        except Exception:
            pass
        return out

    def _selected_bounds0(self, cells0: set[tuple[int, int]]) -> Optional[tuple[int, int, int, int]]:
        if not cells0:
            return None
        rs = [r for r, _c in cells0]
        cs = [c for _r, c in cells0]
        return min(rs), min(cs), max(rs), max(cs)

    def _selected_rect(self) -> Optional[tuple[int, int, int, int]]:
        """Back-compat: return selection bounds for the current selection."""
        return self._selected_bounds0(self._selected_cells0())

    def _copy_selection(self) -> None:
        selected = self._selected_cells0()
        bounds = self._selected_bounds0(selected)
        if bounds is None:
            # Fallback: copy current cell.
            try:
                it = self.table.currentItem()
                QGuiApplication.clipboard().setText(it.text() if it is not None else "")
            except Exception:
                pass
            return

        r0, c0, r1, c1 = bounds
        lines: list[str] = []
        for r in range(int(r0), int(r1) + 1):
            row_vals: list[str] = []
            for c in range(int(c0), int(c1) + 1):
                if (int(r), int(c)) not in selected:
                    row_vals.append("")
                    continue
                it = self.table.item(int(r), int(c))
                row_vals.append(it.text() if it is not None else "")
            lines.append("\t".join(row_vals))
        QGuiApplication.clipboard().setText("\n".join(lines))

    def _cut_selection(self) -> None:
        self._copy_selection()
        selected = self._selected_cells0()
        if not selected:
            return

        updates: dict[BorderKey, Any] = {}
        changed_rows0: set[int] = set()
        for r0, c0 in sorted(selected):
            try:
                it = self.table.item(int(r0), int(c0))
                coord = it.data(Qt.ItemDataRole.UserRole + 1) if it is not None else None
            except Exception:
                coord = None
            if not coord:
                continue
            try:
                r1, c1 = int(coord[0]), int(coord[1])
            except Exception:
                continue
            if (r1, c1) in self._covered_cells:
                continue
            updates[(r1, c1)] = None
            if (c1 - 1) in self._auto_fit_row_height_cols:
                changed_rows0.add(r1 - 1)

        if updates:
            self._batch_set_cells(updates)
            self._auto_fit_rows(changed_rows0)

    def _paste_clipboard(self) -> None:
        if self._ws is None:
            return
        text = QGuiApplication.clipboard().text()
        if not text:
            return

        # Normalize newlines then parse into a matrix.
        norm = text.replace("\r\n", "\n").replace("\r", "\n")
        rows = [line.split("\t") for line in norm.split("\n")]
        # Trim trailing completely-empty row (common when copying from Excel).
        try:
            if rows and len(rows) > 1 and all(v == "" for v in rows[-1]):
                rows = rows[:-1]
        except Exception:
            pass
        if not rows:
            return

        clip_h = len(rows)
        clip_w = max((len(r) for r in rows), default=0)
        if clip_w <= 0:
            return

        # Selection-aware paste rules (Excel-like):
        # - If clipboard is 1 cell and selection exists, fill all selected cells.
        # - Else if selection is a full rectangle of same size, paste into selection.
        # - Else paste starting at current cell.
        selected = self._selected_cells0()
        bounds = self._selected_bounds0(selected)

        start_r0 = None
        start_c0 = None

        if selected and clip_h == 1 and clip_w == 1:
            # Fill selection
            updates: dict[BorderKey, Any] = {}
            changed_rows0: set[int] = set()
            v = self._coerce_text_value(rows[0][0] if rows[0] else "")
            for r0, c0 in sorted(selected):
                it = self.table.item(int(r0), int(c0))
                if it is None:
                    continue
                coord = it.data(Qt.ItemDataRole.UserRole + 1)
                if not coord:
                    continue
                try:
                    r1, c1 = int(coord[0]), int(coord[1])
                except Exception:
                    continue
                if (r1, c1) in self._covered_cells:
                    continue
                updates[(r1, c1)] = v
                if (c1 - 1) in self._auto_fit_row_height_cols:
                    changed_rows0.add(r1 - 1)
            if updates:
                self._batch_set_cells(updates)
                self._auto_fit_rows(changed_rows0)
            return

        if bounds is not None:
            r0, c0, r1, c1 = bounds
            sel_h = int(r1) - int(r0) + 1
            sel_w = int(c1) - int(c0) + 1
            # Is the selection a full rectangle?
            full_rect = len(selected) == (sel_h * sel_w)
            if full_rect and sel_h == clip_h and sel_w == clip_w:
                start_r0, start_c0 = int(r0), int(c0)

        if start_r0 is None or start_c0 is None:
            try:
                start_r0 = int(self.table.currentRow())
                start_c0 = int(self.table.currentColumn())
            except Exception:
                start_r0, start_c0 = -1, -1
            if start_r0 < 0 or start_c0 < 0:
                return

        updates: dict[BorderKey, Any] = {}
        changed_rows0: set[int] = set()
        for dr in range(clip_h):
            for dc in range(clip_w):
                tr0 = int(start_r0) + int(dr)
                tc0 = int(start_c0) + int(dc)
                if tr0 >= self.table.rowCount() or tc0 >= self.table.columnCount():
                    continue
                it = self.table.item(int(tr0), int(tc0))
                if it is None:
                    continue
                coord = it.data(Qt.ItemDataRole.UserRole + 1)
                if not coord:
                    continue
                try:
                    r1, c1 = int(coord[0]), int(coord[1])
                except Exception:
                    continue
                if (r1, c1) in self._covered_cells:
                    continue
                val = ""
                try:
                    val = rows[dr][dc] if dc < len(rows[dr]) else ""
                except Exception:
                    val = ""
                updates[(r1, c1)] = self._coerce_text_value(val)
                if (c1 - 1) in self._auto_fit_row_height_cols:
                    changed_rows0.add(r1 - 1)

        if updates:
            self._batch_set_cells(updates)
            self._auto_fit_rows(changed_rows0)

    def _fill_down(self) -> None:
        rect = self._selected_rect()
        if rect is None or self._ws is None:
            return
        r0, c0, r1, c1 = rect
        if r1 <= r0:
            return
        updates: dict[BorderKey, Any] = {}
        for c in range(c0, c1 + 1):
            src_item = self.table.item(r0, c)
            src_val = self._coerce_text_value(src_item.text() if src_item else "")
            for r in range(r0 + 1, r1 + 1):
                updates[(r + 1, c + 1)] = src_val
        self._batch_set_cells(updates)

    def _fill_right(self) -> None:
        rect = self._selected_rect()
        if rect is None or self._ws is None:
            return
        r0, c0, r1, c1 = rect
        if c1 <= c0:
            return
        updates: dict[BorderKey, Any] = {}
        for r in range(r0, r1 + 1):
            src_item = self.table.item(r, c0)
            src_val = self._coerce_text_value(src_item.text() if src_item else "")
            for c in range(c0 + 1, c1 + 1):
                updates[(r + 1, c + 1)] = src_val
        self._batch_set_cells(updates)

    def _batch_set_cells(self, updates: dict[BorderKey, Any]) -> None:
        if self._ws is None:
            return
        cmd: dict[BorderKey, tuple[Any, Any]] = {}
        self._in_programmatic_change = True
        try:
            for (r, c), new_value in updates.items():
                if (r, c) in self._covered_cells:
                    continue
                cell = self._ws.cell(row=r, column=c)
                old_value = cell.value
                if old_value == new_value:
                    continue
                cell.value = new_value
                it = self.table.item(r - 1, c - 1)
                if it is not None:
                    it.setText("" if new_value is None else str(new_value))
                cmd[(r, c)] = (old_value, new_value)
        finally:
            self._in_programmatic_change = False
        if cmd:
            self._undo_stack.append(cmd)
            self._redo_stack.clear()
            try:
                self.modified.emit()
            except Exception:
                pass

    def _on_column_resized(self, logicalIndex: int, _oldSize: int, newSize: int) -> None:
        if self._suppress_resize_handlers:
            return
        if not self._base_col_widths:
            return
        effective = self._fit_scale * self._user_zoom
        effective = max(self.MIN_EFFECTIVE_SCALE, min(effective, self.MAX_EFFECTIVE_SCALE))
        if effective <= 0:
            effective = 1.0
        if 0 <= logicalIndex < len(self._base_col_widths):
            self._base_col_widths[logicalIndex] = max(int(newSize / effective), 1)
            if not self._lock_effective_scale:
                self._recompute_fit_scale()
            self._schedule_apply()
            self._schedule_persist()

    def _on_row_resized(self, logicalIndex: int, _oldSize: int, newSize: int) -> None:
        if self._suppress_resize_handlers:
            return
        if not self._base_row_heights:
            return
        effective = self._fit_scale * self._user_zoom
        effective = max(self.MIN_EFFECTIVE_SCALE, min(effective, self.MAX_EFFECTIVE_SCALE))
        if effective <= 0:
            effective = 1.0
        if 0 <= logicalIndex < len(self._base_row_heights):
            new_base = max(int(newSize / effective), 1)
            self._base_row_heights[logicalIndex] = new_base
            
            # Persist manual resize to the openpyxl worksheet so it survives re-renders.
            try:
                if self._ws is not None:
                    # Convert pixels back to points (approx 72/96)
                    points = float(new_base * 72 / 96)
                    # row_dimensions is 1-based
                    row_idx = logicalIndex + 1
                    self._ws.row_dimensions[row_idx].height = points
            except Exception:
                pass

            if not self._lock_effective_scale:
                self._recompute_fit_scale()
            self._schedule_apply()
            self._schedule_persist()

    def _build_validation_list_map(
        self,
        ws,
    ) -> tuple[Dict[BorderKey, list[str]], Dict[BorderKey, dict[str, str]], Dict[BorderKey, str]]:
        """Parse list validations into per-cell mappings.

        Returns:
          - mapping: (row, col) -> list of display items
          - display_to_value: (row, col) -> {display_text: stored_value}
        """
        mapping: Dict[BorderKey, list[str]] = {}
        display_to_value: Dict[BorderKey, dict[str, str]] = {}
        kind_map: Dict[BorderKey, str] = {}
        dvs = getattr(ws, "data_validations", None)
        if not dvs:
            return mapping, display_to_value, kind_map

        def _parse_list(formula: str) -> Optional[list[str]]:
            if not formula:
                return None
            f = str(formula).strip()
            # Inline list like "A,B,C"
            if len(f) >= 2 and f[0] == '"' and f[-1] == '"':
                inner = f[1:-1]
                return [x.strip() for x in inner.split(',') if x.strip()]

            # Range reference, e.g. =Sheet!$A$1:$A$10
            # We'll support simple cases and pull values from the workbook.
            # Accept optional leading '=' and optional quoted sheet names.
            try:
                from openpyxl.utils.cell import range_boundaries
            except Exception:
                return None

            f2 = f[1:].strip() if f.startswith("=") else f
            m = re.match(
                r"^(?:(?P<sheet>'[^']+'|[^!]+)!)?(?P<rng>\$?[A-Z]{1,3}\$?\d+:\$?[A-Z]{1,3}\$?\d+)$",
                f2,
            )
            if not m:
                return None

            sheet_name = m.group("sheet")
            rng = m.group("rng")
            if sheet_name:
                sheet_name = str(sheet_name)
                if sheet_name.startswith("'") and sheet_name.endswith("'"):
                    sheet_name = sheet_name[1:-1]

            try:
                min_col, min_row, max_col, max_row = range_boundaries(rng)
            except Exception:
                return None

            wb = getattr(ws, "parent", None)
            ws_ref = ws
            if sheet_name and wb is not None:
                try:
                    ws_ref = wb[sheet_name]
                except Exception:
                    ws_ref = ws

            values: list[str] = []
            for rr in range(min_row, max_row + 1):
                for cc in range(min_col, max_col + 1):
                    v = ws_ref.cell(row=rr, column=cc).value
                    if v is None:
                        continue
                    s = str(v).strip()
                    if not s:
                        continue
                    values.append(s)

            return values or None

        def _parse_supplier_display_map(formula: str) -> Optional[tuple[list[str], dict[str, str]]]:
            """Special case: __as9102_suppliers has A=code, B=customer.

            Validation references codes (col A) so stored values remain valid,
            but the UI should show customers.
            """
            if not formula:
                return None
            f = str(formula).strip()
            f2 = f[1:].strip() if f.startswith("=") else f
            m = re.match(
                r"^(?:(?P<sheet>'[^']+'|[^!]+)!)?(?P<rng>\$?[A-Z]{1,3}\$?\d+:\$?[A-Z]{1,3}\$?\d+)$",
                f2,
            )
            if not m:
                return None
            sheet_name = m.group("sheet")
            rng = m.group("rng")
            if sheet_name:
                sheet_name = str(sheet_name)
                if sheet_name.startswith("'") and sheet_name.endswith("'"):
                    sheet_name = sheet_name[1:-1]
            if (sheet_name or "") != "__as9102_suppliers":
                return None

            try:
                from openpyxl.utils.cell import range_boundaries
            except Exception:
                return None

            try:
                min_col, min_row, max_col, max_row = range_boundaries(rng)
            except Exception:
                return None

            # Expect codes in column A only
            if min_col != 1 or max_col != 1:
                return None

            wb = getattr(ws, "parent", None)
            if wb is None:
                return None
            try:
                ws_ref = wb["__as9102_suppliers"]
            except Exception:
                return None

            items: list[str] = []
            d2v: dict[str, str] = {}
            for rr in range(min_row, max_row + 1):
                code = ws_ref.cell(row=rr, column=1).value
                customer = ws_ref.cell(row=rr, column=2).value
                if code is None:
                    continue
                code_s = str(code).strip()
                if not code_s:
                    continue
                customer_s = str(customer).strip() if customer is not None else ""
                display = customer_s if customer_s else code_s
                if display not in d2v:
                    items.append(display)
                    d2v[display] = code_s

            if not items:
                return None
            return items, d2v

        def _parse_supplier_master_code_display_map(formula: str) -> Optional[tuple[list[str], dict[str, str]]]:
            """Special case: __as9102_supplier_master has A=company, C=code.

            Validation references codes (col C) so stored values remain valid,
            but the UI should show companies.
            """
            if not formula:
                return None
            f = str(formula).strip()
            f2 = f[1:].strip() if f.startswith("=") else f
            m = re.match(
                r"^(?:(?P<sheet>'[^']+'|[^!]+)!)?(?P<rng>\$?[A-Z]{1,3}\$?\d+:\$?[A-Z]{1,3}\$?\d+)$",
                f2,
            )
            if not m:
                return None
            sheet_name = m.group("sheet")
            rng = m.group("rng")
            if sheet_name:
                sheet_name = str(sheet_name)
                if sheet_name.startswith("'") and sheet_name.endswith("'"):
                    sheet_name = sheet_name[1:-1]
            if (sheet_name or "") != "__as9102_supplier_master":
                return None

            try:
                from openpyxl.utils.cell import range_boundaries
            except Exception:
                return None

            try:
                min_col, min_row, max_col, max_row = range_boundaries(rng)
            except Exception:
                return None

            # Expect codes in column C only
            if min_col != 3 or max_col != 3:
                return None

            wb = getattr(ws, "parent", None)
            if wb is None:
                return None
            try:
                ws_ref = wb["__as9102_supplier_master"]
            except Exception:
                return None

            items: list[str] = []
            d2v: dict[str, str] = {}
            for rr in range(min_row, max_row + 1):
                comp = ws_ref.cell(row=rr, column=1).value
                code = ws_ref.cell(row=rr, column=3).value
                if code is None:
                    continue
                code_s = str(code).strip()
                if not code_s:
                    continue
                comp_s = str(comp).strip() if comp is not None else ""
                display = comp_s if comp_s else code_s
                if display not in d2v:
                    items.append(display)
                    d2v[display] = code_s

            if not items:
                return None
            return items, d2v

        def _parse_supplier_master_address_display_map(formula: str) -> Optional[tuple[list[str], dict[str, str]]]:
            """Special case: __as9102_supplier_master has A=company, B=address.

            Validation references addresses (col B) so stored values are addresses,
            but the UI should show companies.
            """
            if not formula:
                return None
            f = str(formula).strip()
            f2 = f[1:].strip() if f.startswith("=") else f
            m = re.match(
                r"^(?:(?P<sheet>'[^']+'|[^!]+)!)?(?P<rng>\$?[A-Z]{1,3}\$?\d+:\$?[A-Z]{1,3}\$?\d+)$",
                f2,
            )
            if not m:
                return None
            sheet_name = m.group("sheet")
            rng = m.group("rng")
            if sheet_name:
                sheet_name = str(sheet_name)
                if sheet_name.startswith("'") and sheet_name.endswith("'"):
                    sheet_name = sheet_name[1:-1]
            if (sheet_name or "") != "__as9102_supplier_master":
                return None

            try:
                from openpyxl.utils.cell import range_boundaries
            except Exception:
                return None

            try:
                min_col, min_row, max_col, max_row = range_boundaries(rng)
            except Exception:
                return None
            # Expect addresses in column B only
            if min_col != 2 or max_col != 2:
                return None

            wb = getattr(ws, "parent", None)
            if wb is None:
                return None
            try:
                ws_ref = wb["__as9102_supplier_master"]
            except Exception:
                return None

            items: list[str] = []
            d2v: dict[str, str] = {}
            for rr in range(min_row, max_row + 1):
                comp = ws_ref.cell(row=rr, column=1).value
                address = ws_ref.cell(row=rr, column=2).value
                if address is None:
                    continue
                addr_s = str(address).strip()
                if not addr_s:
                    continue
                comp_s = str(comp).strip() if comp is not None else ""
                display = comp_s if comp_s else addr_s
                if display not in d2v:
                    items.append(display)
                    d2v[display] = addr_s

            if not items:
                return None
            return items, d2v

        def _parse_supplier_directory_display_map(formula: str) -> Optional[tuple[list[str], dict[str, str]]]:
            """Special case: __as9102_supplier_directory.

            Validation references stored address values (legacy col B; current col G)
            but UI should show companies.
            """
            if not formula:
                return None
            f = str(formula).strip()
            f2 = f[1:].strip() if f.startswith("=") else f
            m = re.match(
                r"^(?:(?P<sheet>'[^']+'|[^!]+)!)?(?P<rng>\$?[A-Z]{1,3}\$?\d+:\$?[A-Z]{1,3}\$?\d+)$",
                f2,
            )
            if not m:
                return None
            sheet_name = m.group("sheet")
            rng = m.group("rng")
            if sheet_name:
                sheet_name = str(sheet_name)
                if sheet_name.startswith("'") and sheet_name.endswith("'"):
                    sheet_name = sheet_name[1:-1]
            if (sheet_name or "") != "__as9102_supplier_directory":
                return None

            try:
                from openpyxl.utils.cell import range_boundaries
            except Exception:
                return None

            try:
                min_col, min_row, max_col, max_row = range_boundaries(rng)
            except Exception:
                return None
            # Expect stored addresses in column B (legacy) or column G (current)
            if (min_col, max_col) == (2, 2):
                address_col = 2
            elif (min_col, max_col) == (7, 7):
                address_col = 7
            elif (min_col, max_col) == (8, 8):
                # Backward compat for older builds that used column H.
                address_col = 8
            else:
                return None

            wb = getattr(ws, "parent", None)
            if wb is None:
                return None
            try:
                ws_ref = wb["__as9102_supplier_directory"]
            except Exception:
                return None

            items: list[str] = []
            d2v: dict[str, str] = {}
            for rr in range(min_row, max_row + 1):
                company = ws_ref.cell(row=rr, column=1).value
                address = ws_ref.cell(row=rr, column=address_col).value
                if address is None:
                    continue
                addr_s = str(address).strip()
                if not addr_s:
                    continue
                comp_s = str(company).strip() if company is not None else ""
                display = comp_s if comp_s else addr_s
                if display not in d2v:
                    items.append(display)
                    d2v[display] = addr_s
            if not items:
                return None
            return items, d2v

        for dv in list(getattr(dvs, "dataValidation", [])):
            if getattr(dv, "type", None) != "list":
                continue
            formula1 = getattr(dv, "formula1", None)
            master_code = _parse_supplier_master_code_display_map(formula1)
            master_addr = _parse_supplier_master_address_display_map(formula1)
            supplier = _parse_supplier_display_map(formula1)
            directory = _parse_supplier_directory_display_map(formula1)
            values = _parse_list(formula1)
            if not values and master_code is None and master_addr is None and supplier is None and directory is None:
                continue
            sqref = getattr(dv, "sqref", None)
            ranges = getattr(sqref, "ranges", None) if sqref is not None else None
            if not ranges:
                continue
            for cr in list(ranges):
                min_row = getattr(cr, "min_row", None)
                min_col = getattr(cr, "min_col", None)
                max_row = getattr(cr, "max_row", None)
                max_col = getattr(cr, "max_col", None)
                if not (min_row and min_col and max_row and max_col):
                    continue
                for r in range(min_row, max_row + 1):
                    for c in range(min_col, max_col + 1):
                        if master_code is not None:
                            items, d2v = master_code
                            mapping[(r, c)] = items
                            display_to_value[(r, c)] = d2v
                            kind_map[(r, c)] = "supplier_code"
                        elif master_addr is not None:
                            items, d2v = master_addr
                            mapping[(r, c)] = items
                            display_to_value[(r, c)] = d2v
                            kind_map[(r, c)] = "supplier_directory"
                        elif supplier is not None:
                            items, d2v = supplier
                            mapping[(r, c)] = items
                            display_to_value[(r, c)] = d2v
                            kind_map[(r, c)] = "supplier_code"
                        elif directory is not None:
                            items, d2v = directory
                            mapping[(r, c)] = items
                            display_to_value[(r, c)] = d2v
                            kind_map[(r, c)] = "supplier_directory"
                        else:
                            mapping[(r, c)] = values
        return mapping, display_to_value, kind_map


class _ExcelEditDelegate(QStyledItemDelegate):
    """Provides combo-box editors for list validations and keeps border painting."""

    @staticmethod
    def _apply_editor_font(editor, index) -> None:
        try:
            f = index.data(Qt.ItemDataRole.FontRole)
        except Exception:
            f = None
        try:
            if f is not None and hasattr(editor, "setFont"):
                editor.setFont(f)
        except Exception:
            pass

    def __init__(self, viewer: ExcelSheetViewer):
        super().__init__(viewer.table)
        self.viewer = viewer
        self.border_delegate: Optional[_ExcelBorderDelegate] = None

    def paint(self, painter, option, index):
        # Copy option so we can safely adjust selection/focus rendering.
        try:
            opt = QStyleOptionViewItem(option)
        except Exception:
            opt = option

        # Apply per-cell wrap based on WRAP_ROLE.
        try:
            wrap_flag = bool(index.data(ExcelSheetViewer.WRAP_ROLE))
        except Exception:
            wrap_flag = False

        try:
            wrap_feature = QStyleOptionViewItem.ViewItemFeature.WrapText
            if wrap_flag:
                opt.features |= wrap_feature
            else:
                opt.features &= ~wrap_feature
        except Exception:
            pass

        # Excel-like selection: no filled selection background, only an outline.
        outline_only = False
        try:
            outline_only = bool(getattr(self.viewer, "_selection_outline_only", False))
        except Exception:
            outline_only = False

        is_selected = False
        try:
            is_selected = bool(opt.state & QStyle.StateFlag.State_Selected)
        except Exception:
            try:
                is_selected = bool(opt.state & QStyle.State_Selected)
            except Exception:
                is_selected = False

        if outline_only and is_selected:
            try:
                opt.state &= ~QStyle.StateFlag.State_Selected
                opt.state &= ~QStyle.StateFlag.State_HasFocus
            except Exception:
                try:
                    opt.state &= ~QStyle.State_Selected
                    opt.state &= ~QStyle.State_HasFocus
                except Exception:
                    pass

        # Paint base cell (including borders).
        if self.border_delegate is not None:
            self.border_delegate.paint(painter, opt, index)
        else:
            super().paint(painter, opt, index)

        # Draw an Excel-like outline for selected cells.
        if outline_only and is_selected:
            try:
                rect = opt.rect.adjusted(1, 1, -1, -1)
            except Exception:
                rect = opt.rect
            try:
                painter.save()
                # Excel selection green (approx.)
                pen = QPen(QColor(0x21, 0xA3, 0x66))
                pen.setWidth(2)
                painter.setPen(pen)
                painter.setBrush(Qt.BrushStyle.NoBrush)
                painter.drawRect(rect)
            except Exception:
                pass
            finally:
                try:
                    painter.restore()
                except Exception:
                    pass

        return

    def createEditor(self, parent, option, index):
        coord = index.data(Qt.ItemDataRole.UserRole + 1)
        if isinstance(coord, list):
            coord = tuple(coord)
        if coord:
            r, c = coord
            values = self.viewer._validation_lists.get((r, c))
            if values:
                from PySide6.QtWidgets import QComboBox

                combo = QComboBox(parent)
                combo.addItems(values)
                combo.setEditable(True)

                # Capture Enter/Shift+Enter while editing to behave like Excel.
                try:
                    combo.installEventFilter(self)
                except Exception:
                    pass

                # Match the cell's visual font while editing (important for
                # symbol fonts like GD&T).
                self._apply_editor_font(combo, index)

                # Make it feel like an Excel-style dropdown: when editing begins,
                # immediately show the list.
                QTimer.singleShot(0, combo.showPopup)
                return combo

        # If the cell is marked as wrapped, use a QTextEdit editor so
        # it stays wrapped during in-cell editing (Excel-like behavior).
        try:
            wrap_flag = bool(index.data(ExcelSheetViewer.WRAP_ROLE))
        except Exception:
            wrap_flag = False
        if wrap_flag:
            try:
                editor = QTextEdit(parent)
                editor.setAcceptRichText(False)
                editor.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
                editor.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
                try:
                    editor.setFrameStyle(0)
                except Exception:
                    pass
                try:
                    editor.setLineWrapMode(QTextEdit.LineWrapMode.WidgetWidth)
                    opt = editor.document().defaultTextOption()
                    opt.setWrapMode(QTextOption.WrapMode.WrapAtWordBoundaryOrAnywhere)
                    editor.document().setDefaultTextOption(opt)
                except Exception:
                    pass

                # Match the cell's visual font while editing.
                self._apply_editor_font(editor, index)

                # Capture Enter/Shift+Enter while editing to behave like Excel.
                try:
                    editor.installEventFilter(self)
                except Exception:
                    pass
                return editor
            except Exception:
                pass

        editor = super().createEditor(parent, option, index)
        try:
            if editor is not None:
                self._apply_editor_font(editor, index)
                # Capture Enter/Shift+Enter while editing to behave like Excel.
                try:
                    editor.installEventFilter(self)
                except Exception:
                    pass
        except Exception:
            pass
        return editor

    def eventFilter(self, obj, event):
        # While an editor has focus, key events won't reach the table's eventFilter.
        # Intercept Enter/Shift+Enter here, commit/close, then let the viewer move.
        try:
            if event.type() == QEvent.Type.KeyPress:
                key = event.key()
                mods = event.modifiers()
                if key in (Qt.Key.Key_Return, Qt.Key.Key_Enter) and mods in (
                    Qt.KeyboardModifier.NoModifier,
                    Qt.KeyboardModifier.ShiftModifier,
                ):
                    direction = -1 if mods == Qt.KeyboardModifier.ShiftModifier else 1
                    try:
                        self.viewer._pending_enter_nav_dir = int(direction)
                    except Exception:
                        pass
                    try:
                        self.commitData.emit(obj)
                    except Exception:
                        pass
                    try:
                        self.closeEditor.emit(obj, QAbstractItemDelegate.EndEditHint.NoHint)
                    except Exception:
                        try:
                            self.closeEditor.emit(obj)
                        except Exception:
                            pass
                    try:
                        event.accept()
                    except Exception:
                        pass
                    return True
        except Exception:
            pass
        return super().eventFilter(obj, event)

    def setEditorData(self, editor, index):
        from PySide6.QtWidgets import QComboBox

        # Ensure editor keeps the same font as the cell.
        try:
            self._apply_editor_font(editor, index)
        except Exception:
            pass

        if isinstance(editor, QComboBox):
            txt = index.data(Qt.ItemDataRole.DisplayRole) or ""
            coord = index.data(Qt.ItemDataRole.UserRole + 1)
            if isinstance(coord, list):
                coord = tuple(coord)
            if coord and coord in self.viewer._validation_display_to_value:
                # Stored value is supplier code; map it back to display (customer).
                d2v = self.viewer._validation_display_to_value.get(coord) or {}
                target_val = str(txt)
                display = None
                for disp, val in d2v.items():
                    if str(val) == target_val:
                        display = disp
                        break
                if display is not None:
                    i = editor.findText(display)
                    if i >= 0:
                        editor.setCurrentIndex(i)
                        return
                editor.setEditText(target_val)
                return

            i = editor.findText(str(txt))
            if i >= 0:
                editor.setCurrentIndex(i)
            else:
                editor.setEditText(str(txt))
            return
        if isinstance(editor, QTextEdit):
            try:
                txt = index.data(Qt.ItemDataRole.DisplayRole) or ""
                editor.setPlainText(str(txt))
                # Place cursor at end (consistent with default editing behavior).
                try:
                    cursor = editor.textCursor()
                    cursor.movePosition(cursor.MoveOperation.End)
                    editor.setTextCursor(cursor)
                except Exception:
                    pass
            except Exception:
                pass
            return
        return super().setEditorData(editor, index)

    def setModelData(self, editor, model, index):
        from PySide6.QtWidgets import QComboBox

        if isinstance(editor, QComboBox):
            coord = index.data(Qt.ItemDataRole.UserRole + 1)
            if isinstance(coord, list):
                coord = tuple(coord)
            txt = editor.currentText()
            if coord and coord in self.viewer._validation_display_to_value:
                d2v = self.viewer._validation_display_to_value.get(coord) or {}
                model.setData(index, d2v.get(txt, txt), Qt.ItemDataRole.EditRole)
                return
            model.setData(index, txt, Qt.ItemDataRole.EditRole)
            return
        if isinstance(editor, QTextEdit):
            try:
                txt = editor.toPlainText()
            except Exception:
                txt = ""
            model.setData(index, txt, Qt.ItemDataRole.EditRole)
            return
        return super().setModelData(editor, model, index)

    def updateEditorGeometry(self, editor, option, index):
        try:
            editor.setGeometry(option.rect)
        except Exception:
            pass
        return super().updateEditorGeometry(editor, option, index)
