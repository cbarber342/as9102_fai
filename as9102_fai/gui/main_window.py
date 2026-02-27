"""Main window - merged from external `cbarber342/as9102_fai` implementation.

This file replaces the local GUI with the external project's `MainWindow`.
Backups of the original local files were saved as `.bak` files.
"""

import sys
import os
import re
import json
import copy
import datetime
import logging
import io
from PySide6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                               QHBoxLayout, QPushButton, QLabel, QFileDialog, 
                               QGroupBox, QFormLayout, QLineEdit, QMessageBox, 
                               QTableWidget, QTableWidgetItem, QHeaderView, QAbstractItemView, QTableWidgetSelectionRange,
                               QTabWidget, QGridLayout, QCheckBox, QScrollArea, QSpinBox, QToolBar, QComboBox, QButtonGroup, QColorDialog, QMenu,
                               QDialog, QProgressBar, QSizePolicy)
from PySide6.QtCore import Qt, QTimer, Signal, QRect, QEvent, QObject
from PySide6.QtCore import QSettings
from PySide6.QtGui import QColor, QFontDatabase, QPalette, QFontMetrics, QKeySequence, QShortcut
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Alignment, Font

logger = logging.getLogger(__name__)


class _DeleteClearsTableCellsFilter(QObject):
    """Event filter: pressing Delete clears selected QTableWidget cell contents."""

    def eventFilter(self, obj, event):
        try:
            if event.type() == QEvent.Type.KeyPress and event.key() == Qt.Key.Key_Delete:
                try:
                    if event.modifiers() != Qt.KeyboardModifier.NoModifier:
                        return super().eventFilter(obj, event)
                except Exception:
                    pass

                if isinstance(obj, QTableWidget):
                    table = obj
                else:
                    return super().eventFilter(obj, event)

                # Don't interfere with in-cell editing; editor should handle Delete.
                try:
                    if table.state() == QAbstractItemView.State.EditingState:
                        return super().eventFilter(obj, event)
                except Exception:
                    pass

                indexes = []
                try:
                    indexes = list(table.selectedIndexes() or [])
                except Exception:
                    indexes = []
                if not indexes:
                    return super().eventFilter(obj, event)

                try:
                    table.blockSignals(True)
                    for ix in indexes:
                        r = int(ix.row())
                        c = int(ix.column())
                        item = table.item(r, c)
                        if item is None:
                            item = QTableWidgetItem("")
                            table.setItem(r, c, item)
                        item.setText("")
                finally:
                    try:
                        table.blockSignals(False)
                    except Exception:
                        pass

                event.accept()
                return True
        except Exception:
            pass
        return super().eventFilter(obj, event)


class _ColorSwatchCheckBox(QCheckBox):
    """A checkbox rendered as a color swatch that supports right-click color picking."""

    colorChanged = Signal()

    def __init__(self, label: str = "", rgb: str | None = None, parent: QWidget | None = None):
        super().__init__(label, parent)
        self._swatch_rgb: str | None = None
        self.setFixedSize(22, 22)
        self.set_swatch_rgb(rgb)

    def swatch_rgb(self) -> str | None:
        return self._swatch_rgb

    def set_swatch_rgb(self, rgb: str | None) -> None:
        s = str(rgb or "").strip().lstrip("#")
        self._swatch_rgb = s.upper() if s else None
        bg = f"#{self._swatch_rgb}" if self._swatch_rgb else "palette(base)"
        self.setStyleSheet(
            "QCheckBox::indicator { width: 18px; height: 18px; }"
            f"QCheckBox::indicator:unchecked {{ background: {bg}; border: 1px solid #666; }}"
            f"QCheckBox::indicator:checked {{ background: {bg}; border: 2px solid #000; }}"
        )

    def mousePressEvent(self, event):
        try:
            if event.button() == Qt.MouseButton.RightButton:
                menu = QMenu(self)
                act_pick = menu.addAction("Choose color…")
                act_none = menu.addAction("No color")
                chosen = menu.exec(event.globalPosition().toPoint() if hasattr(event, "globalPosition") else event.globalPos())

                if chosen == act_none:
                    self.set_swatch_rgb(None)
                    try:
                        self.colorChanged.emit()
                    except Exception:
                        pass
                elif chosen == act_pick:
                    cur = QColor("#" + self._swatch_rgb) if self._swatch_rgb else QColor()
                    color = QColorDialog.getColor(cur, self, "Select Color")
                    if color.isValid():
                        self.set_swatch_rgb(color.name()[1:])
                        try:
                            self.colorChanged.emit()
                        except Exception:
                            pass
                event.accept()
                return
        except Exception:
            pass
        return super().mousePressEvent(event)

US_STATE_CODES: list[str] = [
    "AL","AK","AZ","AR","CA","CO","CT","DE","DC","FL","GA","HI","ID","IL","IN","IA","KS","KY","LA","ME",
    "MD","MA","MI","MN","MS","MO","MT","NE","NV","NH","NJ","NM","NY","NC","ND","OH","OK","OR","PA","RI",
    "SC","SD","TN","TX","UT","VT","VA","WA","WV","WI","WY",
]

US_STATE_NAME_TO_CODE: dict[str, str] = {
    "alabama": "AL",
    "alaska": "AK",
    "arizona": "AZ",
    "arkansas": "AR",
    "california": "CA",
    "colorado": "CO",
    "connecticut": "CT",
    "delaware": "DE",
    "district of columbia": "DC",
    "florida": "FL",
    "georgia": "GA",
    "hawaii": "HI",
    "idaho": "ID",
    "illinois": "IL",
    "indiana": "IN",
    "iowa": "IA",
    "kansas": "KS",
    "kentucky": "KY",
    "louisiana": "LA",
    "maine": "ME",
    "maryland": "MD",
    "massachusetts": "MA",
    "michigan": "MI",
    "minnesota": "MN",
    "mississippi": "MS",
    "missouri": "MO",
    "montana": "MT",
    "nebraska": "NE",
    "nevada": "NV",
    "new hampshire": "NH",
    "new jersey": "NJ",
    "new mexico": "NM",
    "new york": "NY",
    "north carolina": "NC",
    "north dakota": "ND",
    "ohio": "OH",
    "oklahoma": "OK",
    "oregon": "OR",
    "pennsylvania": "PA",
    "rhode island": "RI",
    "south carolina": "SC",
    "south dakota": "SD",
    "tennessee": "TN",
    "texas": "TX",
    "utah": "UT",
    "vermont": "VT",
    "virginia": "VA",
    "washington": "WA",
    "west virginia": "WV",
    "wisconsin": "WI",
    "wyoming": "WY",
}

_US_STATE_NAMES_BY_LENGTH: list[str] = sorted(US_STATE_NAME_TO_CODE.keys(), key=len, reverse=True)


def _clean_company_prefix(company: str, address: str) -> str:
    comp = str(company or "").strip()
    addr = str(address or "").strip()
    if not comp or not addr:
        return addr

    a = addr.lstrip(" ,")
    c = comp.strip().rstrip(",")
    if a.lower().startswith(c.lower()):
        rest = a[len(c):].lstrip(" ,")
        return rest
    return addr


def _parse_us_city_state_zip(s: str) -> tuple[str, str, str]:
    """Parse 'City, ST 12345' (or 'City ST 12345') into (city, state, zip)."""
    s = str(s or "").strip()
    if not s:
        return "", "", ""

    # Split on comma first
    parts = [p.strip() for p in s.split(",") if p.strip()]
    if len(parts) >= 2:
        city = parts[-2]
        tail = parts[-1]
    else:
        city = ""
        tail = parts[-1] if parts else s

    tail_up = tail.upper()
    m = re.search(r"\b([A-Z]{2})\b\s*(\d{5}(?:-\d{4})?)?\b", tail_up)
    if m:
        st = m.group(1)
        z = m.group(2) or ""
        # If we didn't get city from comma form, try to infer from leading tokens
        if not city:
            city_guess = re.sub(r"\b([A-Z]{2})\b\s*(\d{5}(?:-\d{4})?)?\b.*$", "", s, flags=re.IGNORECASE).strip(" ,")
            city = city_guess
        return city.strip(), st.strip(), z.strip()

    # Try state names + ZIP
    tail_low = tail.lower()
    zip_m = re.search(r"\b(\d{5}(?:-\d{4})?)\b", tail)
    zipc = zip_m.group(1) if zip_m else ""
    for nm in _US_STATE_NAMES_BY_LENGTH:
        if nm in tail_low:
            st = US_STATE_NAME_TO_CODE.get(nm, "")
            if not city:
                # City is typically the prior comma chunk in the caller.
                city_guess = re.sub(re.escape(nm), "", tail_low, flags=re.IGNORECASE).strip(" ,")
                if city_guess and city_guess != tail_low:
                    city = city_guess
            return city.strip(), st, zipc.strip()

    return city.strip(), "", ""


def _split_address_lines(full_address: str) -> tuple[str, str, str, str, str, str]:
    """Split a one-line address into (addr1, addr2, addr3, city, state, zip)."""
    s = str(full_address or "").strip()
    if not s:
        return "", "", "", "", "", ""

    # Normalize whitespace
    s = re.sub(r"\s+", " ", s)

    # Try comma-based split
    parts = [p.strip() for p in s.split(",") if p.strip()]
    city = state = zipc = ""

    if len(parts) >= 2:
        # Attempt to parse last two chunks as city/state/zip
        city2, st2, z2 = _parse_us_city_state_zip(", ".join(parts[-2:]))
        if st2 or z2:
            city, state, zipc = city2, st2, z2
            parts = parts[:-2]
        else:
            # Sometimes last chunk is 'ST ZIP' and previous is city
            city2, st2, z2 = _parse_us_city_state_zip(parts[-1])
            if st2 or z2:
                city = parts[-2]
                state, zipc = st2, z2
                parts = parts[:-2]

    # Whatever remains are street lines
    addr_lines = parts if parts else [s]
    addr1 = addr_lines[0] if len(addr_lines) > 0 else ""
    addr2 = addr_lines[1] if len(addr_lines) > 1 else ""
    addr3 = addr_lines[2] if len(addr_lines) > 2 else ""
    return addr1, addr2, addr3, city, state, zipc


def _build_full_address(addr1: str, addr2: str, addr3: str, city: str, state: str, zipc: str) -> str:
    parts: list[str] = []
    for a in (addr1, addr2, addr3):
        a = str(a or "").strip().strip(",")
        if a:
            parts.append(a)
    cs = str(city or "").strip().strip(",")
    st = str(state or "").strip().upper()
    z = str(zipc or "").strip()
    tail = ""
    if cs and st and z:
        tail = f"{cs}, {st} {z}"
    elif cs and st:
        tail = f"{cs}, {st}"
    elif cs:
        tail = cs
    elif st and z:
        tail = f"{st} {z}"
    elif st:
        tail = st
    elif z:
        tail = z
    if tail:
        parts.append(tail)
    return ", ".join(parts)


def _build_full_address_v2(addr1: str, addr2: str, city: str, state: str, zipc: str) -> str:
    return _build_full_address(addr1, addr2, "", city, state, zipc)


def _build_full_address_with_company(company: str, addr1: str, addr2: str, city: str, state: str, zipc: str) -> str:
    """Build the stored dropdown value shown in Form 1/2 cells.

    The dropdown displays Company, and the stored value includes Company + address.
    """
    comp_s = str(company or "").strip().strip(",")
    base = _build_full_address_v2(addr1, addr2, city, state, zipc)
    if comp_s and base:
        return f"{comp_s}, {base}"
    return comp_s or base


# Supplier directory seed rows EXACTLY as provided in the screenshot.
# Columns: Company, Address 1, Address 2, City, State, Zip Code
DEFAULT_SUPPLIER_DIRECTORY_SEED: list[tuple[str, str, str, str, str, str]] = [
    ("A. M. Castle & Co", "2602 Pinewood Drive", "", "Grand Prairie", "TX", "75050"),
    ("Adept Fasteners", "27949 Hancock Parkway", "", "Valencia", "CA", "91355"),
    ("Aeromil Industrial Coatings LLC", "49 West Pima Street", "", "Phoenix", "AZ", "85003"),
    ("Altemp Alloys LLC", "330 West Taft Avenue", "", "Orange", "CA", "92865"),
    ("Alumiplate Incorporated", "8960 Springbrook Drive NW", "Suite 105", "Minneapolis", "MN", "55433"),
    ("Applied Thermal Technologies", "2169 N 100 E", "", "Warsaw", "IN", "46582"),
    ("Atlantic Casting & Engineering Corp.", "810 Bloomfield Ave", "", "Clifton", "NJ", "7012"),
    ("Aviva Metals Inc.", "2929 West 12th Street", "", "Houston", "TX", "77008"),
    ("B&M Finishers Inc. / Kenelworth Anodizing", "201 South 31st Street", "", "Kenilworth", "NJ", "7033"),
    ("Bodycote", "4008 Clay Avenue", "Suite 200", "Haltom City", "TX", "76117"),
    ("Boedeker Plastics", "904 West Sixth Street", "", "Shiner", "TX", "77984"),
    ("Bralco Metals #08", "410 Mars Drive", "", "Garland", "TX", "75040"),
    ("Camtron Incorporated", "3101 Summit Ave", "Suite 300", "Plano", "TX", "75074"),
    ("Carr Lane Manufacturing Co.", "4200 Carr Lane Ct", "", "Saint Louis", "MO", "63119"),
    ("Courter-Hall", "1910 North First St.", "", "Garland", "TX", "75040"),
    ("D.B. Roberts Company", "3100 Summit Ave.", "Suite 100", "Plano", "TX", "75074"),
    ("Ed Fagan Inc.", "10537 Humbolt Street", "", "Los Alamitos", "CA", "90720"),
    ("EFINEA", "1847 W Business Center Drive", "", "Orange", "CA", "92867"),
    ("Ellsworth Adhesives", "W129 N10825 Washington Dr.", "", "Germantown", "WI", "53022"),
    ("Embee Processing LLC", "2158 South Hathaway Street", "", "Santa Ana", "CA", "92705"),
    ("F.M. Callahan and SON Inc.", "22 Sharon Street", "", "Malden", "MA", "2148"),
    ("Farmers Copper LTD.", "9900 Emmett F Lowry Expy", "", "Texas City", "TX", "77591"),
    ("General Metal Finishing", "42 Frank Mossberg Drive", "", "Attleboro", "MA", "2703"),
    ("Hadco Metal Trading Co. LLC", "24403 Amah Parkway", "", "Claremore", "OK", "74019"),
    ("Har-Conn Aerospace", "5000 Augusta Drive", "", "Fort Worth", "TX", "76106"),
    ("Hardware Specialty Co. Inc.", "48-75 36th Street", "", "Long Island City", "NY", "11101"),
    ("Hydraflow", "1881 W. Malvern", "", "Fullerton", "CA", "92833"),
    ("Indian Industries", "432 West Fork Dr", "", "Arlington", "TX", "76012"),
    ("Industrial Precision Coating", "2 Trim Way", "", "Randolph", "MA", "2368"),
    ("J. L. Anthony & Company", "115 Baker Street", "", "Providence", "RI", "2905"),
    ("Krayden", "1491 West 124th Avenue", "", "Westminster", "CO", "80234"),
    ("Mark Finishing LLC", "2509 Silver Maple Drive", "", "St. Paul", "TX", "75098"),
    ("McCarty & Sons Inc.", "81 Westbrook Industrial Park Road", "PO Box 543", "Westbrook", "CT", "6498"),
    ("McMaster-Carr", "1901 Riverside Pkwy", "", "Douglasville", "GA", "30135"),
    ("Mead Metals Inc.", "555 Cardigan Road", "", "St. Paul", "MN", "55126"),
    ("Metallurgical Engineering Services Inc.", "845 E. Arapaho Road", "", "Richardson", "TX", "75081"),
    ("MetalMart International", "5828 Smithway Street", "", "Commerce", "CA", "90040"),
    ("National Aluminum and Alloy", "PO Box 66", "", "Muenster", "TX", "76252"),
    ("National Electronic Alloys", "1335 East Warner Ave.", "", "Santa Ana", "CA", "92705"),
    ("OnlineMetals", "1500 Cherokee Parkway", "", "Acworth", "GA", "30102"),
    ("Precision Sensors", "340 Woodmont Road", "", "Milford", "CT", "6460"),
    ("PTI Industries", "2 Peerless Way", "", "Enfield", "CT", "6082"),
    ("Rolled Alloys Inc.", "3173 Crenshaw Parkway", "", "Richburg", "SC", "29729"),
    ("Sapa Industrial Extrusions", "1550 Kirby Lane", "", "Spanish Fork", "UT", "84660"),
    ("Sciaky Inc.", "4915 W. 67th Street", "", "Chicago", "IL", "60638"),
    ("Simco Coatings Inc.", "211 Gunther Ln.", "", "Belle Chasse", "LA", "70037"),
    ("Skygeek Logistics Inc.", "30 Airway Drive", "Suite 2", "Lagrangeville", "NY", "12540"),
    ("Spectrum Coatings Inc.", "217 Chapman Street", "", "Providence", "RI", "2905"),
    ("Spira Manufacturing Corporation", "650 Jessie Street", "", "San Fernando", "CA", "91340"),
    ("The Hitt Companies Inc.", "3231 W. MacArthur Blvd.", "", "Santa Ana", "CA", "92704"),
    ("The Indium Corporation of America", "5836 Success Drive", "", "Rome", "NY", "13440"),
    ("Titanium Industries Inc.", "1450 N. Hwy 77", "", "Hillsboro", "TX", "76645"),
    ("Triad Product Finishing", "1440 South Highway 121", "Suite 13", "Lewisville", "TX", "75067"),
    ("Trident Company", "405 N Plano Rd.", "", "Richardson", "TX", "75081"),
    ("Mi Tech Tungsten Metals LLC", "4701 Massachusetts Avenue", "", "Indianapolis", "IN", "46218"),
]


DEFAULT_CALIBRATED_EQUIPMENT_SEED: list[tuple[str, str, str, str]] = [
    ("Spectrum", "480537", "Zeiss CMM", "4-23-26"),
    ("Contura", "501477", "Zeiss CMM", "6-18-26"),
    ("Micura", "510009", "Zeiss CMM", "6-17-26"),
    ("Contura", "560711", "Zeiss CMM", "2-6-26"),
]

from as9102_fai.parsers.chr_parser import ChrParser
from as9102_fai.gui.pdf_viewer import PdfViewer
from as9102_fai.gui.drawing_viewer_window import DrawingViewerWindow
from as9102_fai.reports.fai_generator import FaiGenerator
from as9102_fai.parsers.pdf_extractor import PdfTextExtractor
from as9102_fai.gui.excel_sheet_viewer import ExcelSheetViewer


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("AS9102 FAI Generator")
        self.resize(1400, 900)
        
        self.parser = ChrParser()
        self.characteristics = []
        self.template_path = ""
        self.pdf_path = ""
        self.drawing_pdf_path = ""
        # Template workbook is loaded lazily when a template is selected.
        # Initialize to avoid AttributeError during early load_defaults/load_chr.
        self._template_wb = None
        # Forms are keyed to match the common template structure
        # (Form 1, Form 2, Form 2 Cont., Form 3)
        self._form_sheet_names = {"1": None, "2": None, "2c": None, "3": None}
        self._form_viewers = {"1": None, "2": None, "2c": None, "3": None}
        self._supplier_directory_table = None
        self._supplier_directory_suppress_changes = False

        # Calibrated Equipment tab state
        self._calibrated_equipment_table = None
        self._calibrated_equipment_suppress_changes = False

        # Per-user/per-machine settings (Windows registry on Windows).
        self._settings = QSettings("as9102_fai", "as9102_fai_gui")

        # Form 3 option: include derived thread rows (Go/No-Go + Minor Dia).
        self._form3_include_thread_extras = self._settings.value(
            "forms/3/include_thread_extras",
            False,
            type=bool,
        )

        # Form 3 option: GD&T callout rendering mode.
        # Form 3 option: GD&T callout rendering.
        # Per request: always use installed-font mode and default to the "GDT" font.
        self._form3_gdt_callout_mode = "font"
        self._form3_gdt_font_family = "GDT"

        # Debounce timers for persisting table column/row sizes.
        self._table_persist_timers: dict[str, QTimer] = {}

        # Form 3 undo stack (for row delete operations).
        self._form3_undo_stack: list[bytes] = []
        self._form3_undo_max = 20
        
        # Default file paths
        # Priority order:
        # 1) Explicit env var per file (AS9102_FAI_DEFAULT_CHR/TEMPLATE/DRAWING)
        # 2) AS9102_FAI_SAMPLE_DIR (sample_chr.txt/template.xlsx/drawing.pdf)
        # 3) Fallback paths (matching the screenshot)
        sample_dir = os.environ.get("AS9102_FAI_SAMPLE_DIR", "").strip()

        def _norm(p: str) -> str:
            p = (p or "").strip().replace("/", os.sep)
            return os.path.normpath(os.path.expanduser(p)) if p else ""

        screenshot_chr = ""
        screenshot_template = ""
        screenshot_drawing = ""

        env_chr = _norm(os.environ.get("AS9102_FAI_DEFAULT_CHR", ""))
        env_template = _norm(os.environ.get("AS9102_FAI_DEFAULT_TEMPLATE", ""))
        env_drawing = _norm(os.environ.get("AS9102_FAI_DEFAULT_DRAWING", ""))

        sample_chr = _norm(os.path.join(sample_dir, "sample_chr.txt")) if sample_dir else ""
        sample_template = _norm(os.path.join(sample_dir, "template.xlsx")) if sample_dir else ""
        sample_drawing = _norm(os.path.join(sample_dir, "drawing.pdf")) if sample_dir else ""

        self.default_chr_path = env_chr or (sample_chr if sample_dir else "") or _norm(screenshot_chr)
        self.default_template_path = env_template or (sample_template if sample_dir else "") or _norm(screenshot_template)
        self.default_drawing_path = env_drawing or (sample_drawing if sample_dir else "") or _norm(screenshot_drawing)
        
        self.setup_ui()
        # Defer default loading until the Qt event loop starts; this matches the
        # timing of user-driven Browse selection and ensures the preview table
        # updates reliably on startup.
        QTimer.singleShot(0, self.load_defaults)

        # Ctrl+Z (undo) for Form 3 row deletions.
        # Use a single QAction to avoid ambiguous shortcut overloads.
        self._form3_undo_shortcut = None

        self._form3_undo_action = None

        # Global event filter for reliable Ctrl+Z when focus is inside Form 3.
        try:
            app = QApplication.instance()
            if app is not None:
                app.installEventFilter(self)
        except Exception:
            pass

    def _sync_bubbles_to_form3(self, bubbled_numbers: set[int] | None = None) -> None:
        """Ensure Form 3 bubble fill colors match the drawing's current bubbles.

        This is intentionally tolerant of startup ordering (drawing/template/chr
        can load in different orders).
        """

        if bubbled_numbers is None:
            try:
                dv = getattr(self, "drawing_viewer_tab", None)
                pv = getattr(dv, "_pdf_viewer", None) if dv is not None else None
                if pv is not None and hasattr(pv, "get_bubbled_numbers"):
                    bubbled_numbers = set(pv.get_bubbled_numbers() or set())
            except Exception:
                bubbled_numbers = None

        try:
            s = set(int(x) for x in (bubbled_numbers or set()))
        except Exception:
            s = set()

        try:
            self._last_bubbled_numbers = set(s)
        except Exception:
            pass

        try:
            self._update_form3_bubble_fills(s)
        except Exception:
            pass

        try:
            v3 = self._form_viewers.get("3")
            if v3 is not None:
                tbl = getattr(v3, "table", None)
                if tbl is not None:
                    tbl.viewport().update()
        except Exception:
            pass

    def eventFilter(self, obj, event):
        try:
            if event.type() == QEvent.Type.KeyPress:
                is_undo = False
                try:
                    if event.matches(QKeySequence.Undo):
                        is_undo = True
                except Exception:
                    is_undo = False
                if not is_undo:
                    try:
                        if event.key() == Qt.Key.Key_Z and (event.modifiers() & Qt.KeyboardModifier.ControlModifier):
                            is_undo = True
                    except Exception:
                        is_undo = False

                if not is_undo:
                    return super().eventFilter(obj, event)

                try:
                    print("Form3 Ctrl+Z keypress detected")
                except Exception:
                    pass

                viewer = self._get_focused_excel_viewer()
                if viewer is not None:
                    if str(getattr(viewer, "form_key", "")) == "3":
                        try:
                            if viewer.table.state() == QAbstractItemView.State.EditingState:
                                return super().eventFilter(obj, event)
                        except Exception:
                            pass
                        handled = bool(self._on_form3_undo_requested())
                        if handled:
                            event.accept()
                            return True
                else:
                    # If no focused viewer, still allow undo when Form 3 tab is active.
                    try:
                        if self._is_form3_tab_active():
                            handled = bool(self._on_form3_undo_requested())
                            if handled:
                                event.accept()
                                return True
                    except Exception:
                        pass
        except Exception:
            pass
        return super().eventFilter(obj, event)

    def _get_focused_excel_viewer(self):
        try:
            fw = QApplication.focusWidget()
        except Exception:
            fw = None
        while fw is not None:
            try:
                if isinstance(fw, ExcelSheetViewer):
                    return fw
            except Exception:
                pass
            try:
                fw = fw.parentWidget()
            except Exception:
                break
        return None

    def _is_form3_tab_active(self) -> bool:
        try:
            tabs = getattr(self, "forms_tabs", None)
            key_map = getattr(self, "_form_tab_to_key", {}) or {}
            if tabs is None:
                return False
            w = tabs.currentWidget()
            return str(key_map.get(w, "")) == "3"
        except Exception:
            return False

    def _refresh_form3_view(self) -> None:
        if self._template_wb is None:
            return
        ws3_name = self._form_sheet_names.get("3")
        viewer3 = self._form_viewers.get("3")
        if not ws3_name or not viewer3 or ws3_name not in self._template_wb.sheetnames:
            return

        try:
            self._write_form3_to_worksheet(self._template_wb[ws3_name])
        except Exception:
            pass
        try:
            viewer3.set_overrides({})
            viewer3.render()
        except Exception:
            pass

    def load_defaults(self):
        """Populate the UI with default file paths; auto-load when files exist."""
        print("DEBUG: load_defaults started")
        last_chr = str(self._settings.value("paths/chr", ""))
        last_template = str(self._settings.value("paths/template", ""))
        last_drawing = str(self._settings.value("paths/drawing_pdf", ""))
        last_machine = str(self._settings.value("inputs/calibrated_equipment_machine", ""))

        chr_path = last_chr.strip() or self.default_chr_path
        template_path = last_template.strip() or self.default_template_path
        drawing_path = last_drawing.strip() or self.default_drawing_path

        if chr_path:
            # Use the same code path as the Browse button when possible.
            if os.path.exists(chr_path):
                print(f"DEBUG: Loading CHR: {chr_path}")
                self.load_chr(chr_path)
            else:
                self.chr_path_edit.setText(chr_path)

        if template_path:
            self.template_path_edit.setText(template_path)
            if os.path.exists(template_path):
                self.template_path = template_path
                print(f"DEBUG: Loading Template: {template_path}")
                self.load_template()

        if drawing_path:
            self.drawing_pdf_edit.setText(drawing_path)
            if os.path.exists(drawing_path):
                self.drawing_pdf_path = drawing_path
                try:
                    dv = getattr(self, "drawing_viewer_tab", None)
                    if dv is not None:
                        print(f"DEBUG: Loading PDF: {drawing_path}")
                        dv.load_pdf(self.drawing_pdf_path)
                        # Bubble state may already exist on the drawing. Sync
                        # Form 3 shading after the viewer finishes loading.
                        QTimer.singleShot(75, self._sync_bubbles_to_form3)
                except Exception:
                    pass

        # Populate machine dropdown from persisted equipment rows and restore selection.
        try:
            print("DEBUG: Refreshing equipment combo")
            self._refresh_calibrated_equipment_combo(preserve_selection=False)
            if last_machine and hasattr(self, "calibrated_equipment_combo") and self.calibrated_equipment_combo is not None:
                i = self.calibrated_equipment_combo.findText(last_machine)
                if i >= 0:
                    self.calibrated_equipment_combo.setCurrentIndex(i)
        except Exception:
            pass
        print("DEBUG: load_defaults finished")
        
    def setup_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(2, 2, 2, 2)
        main_layout.setSpacing(6)

        # Primary actions toolbar (best-practice: keep main actions always visible).
        toolbar = QToolBar("Main")
        toolbar.setMovable(False)
        self.addToolBar(toolbar)

        toolbar.addAction("Generate Report", self.generate_report)

        # Spacer to push theme toggle to the right (or just separate it)
        empty = QWidget()
        empty.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        toolbar.addWidget(empty)
        
        toolbar.addWidget(QLabel("Theme: "))
        self.theme_combo = QComboBox()
        self.theme_combo.addItems(["Light", "Dark"])
        # Restore last saved theme
        last_theme = str(self._settings.value("theme", "Dark"))
        self.theme_combo.setCurrentText(last_theme)
        # Apply initially
        QTimer.singleShot(0, lambda: self._set_theme(last_theme))
        
        self.theme_combo.currentTextChanged.connect(self._set_theme)
        toolbar.addWidget(self.theme_combo)

        # Build Inputs UI (hosted in a tab, not a dock)
        inputs_tab = QWidget()
        inputs_layout = QVBoxLayout(inputs_tab)
        inputs_layout.setContentsMargins(8, 8, 8, 8)
        inputs_layout.setSpacing(8)

        file_group = QGroupBox("Input Files")
        file_layout = QFormLayout(file_group)
        file_layout.setContentsMargins(8, 8, 8, 8)
        file_layout.setSpacing(6)
        
        self.chr_path_edit = QLineEdit()
        self.chr_path_edit.setPlaceholderText("Drop chr.txt here or click Browse")
        self.chr_browse_btn = QPushButton("Browse")
        self.chr_browse_btn.clicked.connect(self.browse_chr)

        # Calibrated equipment selection (machine names) - populated from the Calibrated Equipment tab.
        self.calibrated_equipment_combo = QComboBox()
        self.calibrated_equipment_combo.setEditable(False)
        # Keep it compact (~20 characters wide) per request.
        try:
            self.calibrated_equipment_combo.setMinimumContentsLength(20)
            self.calibrated_equipment_combo.setSizeAdjustPolicy(
                QComboBox.SizeAdjustPolicy.AdjustToMinimumContentsLengthWithIcon
            )
        except Exception:
            pass
        self.calibrated_equipment_combo.currentIndexChanged.connect(self._on_calibrated_equipment_combo_changed)
        
        self.drawing_pdf_edit = QLineEdit()
        self.drawing_pdf_edit.setPlaceholderText("Drop Drawing PDF here")
        self.drawing_pdf_browse_btn = QPushButton("Browse")
        self.drawing_pdf_browse_btn.clicked.connect(self.browse_drawing_pdf)

        try:
            self.drawing_pdf_edit.editingFinished.connect(self._on_drawing_pdf_edit_committed)
        except Exception:
            pass
        
        # Machine dropdown BEFORE the Calypso file path/browse, on the same row.
        machine_chr_widget = QWidget()
        machine_chr_layout = QHBoxLayout(machine_chr_widget)
        machine_chr_layout.setContentsMargins(0, 0, 0, 0)
        machine_chr_layout.setSpacing(6)
        machine_chr_layout.addWidget(self.calibrated_equipment_combo)
        machine_chr_layout.addWidget(self.chr_path_edit)
        machine_chr_layout.addWidget(self.chr_browse_btn)
        file_layout.addRow("Machine / Calypso File (*.txt):", machine_chr_widget)
        file_layout.addRow("Drawing PDF:", self.create_file_row(self.drawing_pdf_edit, self.drawing_pdf_browse_btn))

        template_group = QGroupBox("FAI Template")
        template_layout = QFormLayout(template_group)
        template_layout.setContentsMargins(8, 8, 8, 8)
        template_layout.setSpacing(6)

        self.template_path_edit = QLineEdit()
        self.template_path_edit.setPlaceholderText("Drop Excel Template here")
        self.template_browse_btn = QPushButton("Browse")
        self.template_browse_btn.clicked.connect(self.browse_template)
        template_layout.addRow("Template (*.xlsx):", self.create_file_row(self.template_path_edit, self.template_browse_btn))

        inputs_layout.addWidget(file_group)
        inputs_layout.addWidget(template_group)
        inputs_layout.addStretch(1)
        
        # Forms (tabs)
        self.forms_tabs = QTabWidget()
        self.forms_tabs.setDocumentMode(True)

        # Drawing Viewer tab (embedded) before Form 1
        drawing_tab = QWidget()
        self._drawing_tab_widget = drawing_tab
        drawing_tab_l = QVBoxLayout(drawing_tab)
        drawing_tab_l.setContentsMargins(6, 6, 6, 6)
        drawing_tab_l.setSpacing(6)
        self._drawing_tab_layout = drawing_tab_l

        # Embed the same layout as the standalone Drawing Viewer window
        # (Properties/Notes docks + drawing toolbar).
        self.drawing_viewer_tab = DrawingViewerWindow(
            pdf_path="",
            default_save_basename="",
            pop_out_callback=self._undock_drawing_viewer,
            dock_back_callback=self._dock_drawing_viewer,
        )
        try:
            self.drawing_viewer_tab.set_docked_state(True)
        except Exception:
            pass
        try:
            v = getattr(self.drawing_viewer_tab, "_pdf_viewer", None)
            if v is not None and hasattr(v, "bubbles_changed"):
                v.bubbles_changed.connect(self._on_drawing_bubbles_changed)
            if v is not None and hasattr(v, "drawing_saved"):
                v.drawing_saved.connect(self._on_drawing_saved)
            if v is not None and hasattr(v, "insert_notes_to_form3_requested"):
                v.insert_notes_to_form3_requested.connect(self._on_insert_notes_to_form3)
        except Exception:
            pass

        # Drawing Viewer bubble scroller: optionally highlight Form 3 when viewer is popped out.
        try:
            if hasattr(self.drawing_viewer_tab, "bubbleScrollSelected"):
                self.drawing_viewer_tab.bubbleScrollSelected.connect(self._on_drawing_bubble_scroller_selected)
        except Exception:
            pass
        drawing_tab_l.addWidget(self.drawing_viewer_tab, 1)

        tab1, viewer1 = self._create_form_tab("1")
        tab2, viewer2 = self._create_form_tab("2")
        tab2c, viewer2c = self._create_form_tab("2c")
        tab3, viewer3 = self._create_form_tab("3")

        # Map the tab widgets to their form keys so toolbar actions can target
        # the correct viewer based on the active tab.
        self._form_tab_to_key = {
            tab1: "1",
            tab2: "2",
            tab2c: "2c",
            tab3: "3",
        }

        self._form_viewers["1"] = viewer1
        self._form_viewers["2"] = viewer2
        self._form_viewers["2c"] = viewer2c
        self._form_viewers["3"] = viewer3

        self.forms_tabs.addTab(drawing_tab, "Drawing Viewer")
        self.forms_tabs.addTab(tab1, "Form 1")
        self.forms_tabs.addTab(tab2, "Form 2")
        self.forms_tabs.addTab(tab2c, "Form 2 Cont.")
        self.forms_tabs.addTab(tab3, "Form 3")

        customer_tab = self._create_suppliers_tab()
        self.forms_tabs.addTab(customer_tab, "Customer")

        supplier_tab = self._create_supplier_directory_tab()
        self.forms_tabs.addTab(supplier_tab, "Supplier")

        calibrated_equipment_tab = self._create_calibrated_equipment_tab()
        self.forms_tabs.addTab(calibrated_equipment_tab, "Calibrated Equipment")

        self.forms_tabs.addTab(inputs_tab, "Inputs")



        main_layout.addWidget(self.forms_tabs)

        # Initial population of the machine dropdown.
        try:
            self._refresh_calibrated_equipment_combo(preserve_selection=False)
        except Exception:
            pass



    def _undock_drawing_viewer(self) -> None:
        """Remove the Drawing Viewer from tabs and show it as a window."""
        try:
            if not hasattr(self, "forms_tabs") or self.forms_tabs is None:
                return
            tab = getattr(self, "_drawing_tab_widget", None)
            if tab is None:
                return

            idx = self.forms_tabs.indexOf(tab)
            if idx != -1:
                self.forms_tabs.removeTab(idx)

            dv = getattr(self, "drawing_viewer_tab", None)
            if dv is None:
                return

            # Detach from layout and show as a top-level window.
            dv.setParent(None)
            dv.setWindowFlags(Qt.Window)
            try:
                dv.set_docked_state(False)
            except Exception:
                pass
            dv.showMaximized()
        except Exception:
            pass

    def _dock_drawing_viewer(self) -> None:
        """Dock the Drawing Viewer back into the first tab."""
        try:
            dv = getattr(self, "drawing_viewer_tab", None)
            tab = getattr(self, "_drawing_tab_widget", None)
            lay = getattr(self, "_drawing_tab_layout", None)
            if dv is None or tab is None or lay is None:
                return

            # If tab is not currently present, insert it back at the front.
            idx = self.forms_tabs.indexOf(tab)
            if idx == -1:
                self.forms_tabs.insertTab(0, tab, "Drawing Viewer")
                self.forms_tabs.setCurrentIndex(0)
            else:
                self.forms_tabs.setCurrentIndex(idx)

            dv.setParent(tab)
            dv.setWindowFlags(Qt.Widget)
            try:
                dv.set_docked_state(True)
            except Exception:
                pass

            # Ensure it's in the tab layout.
            try:
                lay.addWidget(dv, 1)
            except Exception:
                pass
            dv.show()
        except Exception:
            pass

    def _calibrated_equipment_rows_for_dropdown(self) -> list[tuple[str, str, str, str]]:
        """Return calibrated equipment rows for the Inputs dropdown.

        Prefer the currently loaded template workbook's hidden sheet, but fall back
        to QSettings persistence so the dropdown works even before a template is loaded.
        """

        try:
            if self._template_wb is not None:
                rows = self._calibrated_equipment_rows()
                if rows:
                    return rows
        except Exception:
            pass

        try:
            return self._load_persistent_calibrated_equipment_rows()
        except Exception:
            return []

    def _refresh_calibrated_equipment_combo(self, preserve_selection: bool = True) -> None:
        if not hasattr(self, "calibrated_equipment_combo") or self.calibrated_equipment_combo is None:
            return

        combo = self.calibrated_equipment_combo
        prev = str(combo.currentText() or "").strip() if preserve_selection else ""

        rows = self._calibrated_equipment_rows_for_dropdown()
        items: list[tuple[str, tuple[str, str, str]]] = []
        seen: set[str] = set()
        for name, mid, mtype, due in rows:
            n = str(name or "").strip()
            if not n:
                continue
            key = n.lower()
            if key in seen:
                continue
            seen.add(key)
            items.append((n, (str(mid or "").strip(), str(mtype or "").strip(), str(due or "").strip())))

        try:
            combo.blockSignals(True)
            combo.clear()
            combo.addItem("", None)
            for name, data in items:
                combo.addItem(name, data)
            if prev:
                i = combo.findText(prev)
                if i >= 0:
                    combo.setCurrentIndex(i)
        finally:
            combo.blockSignals(False)

    def _selected_calibrated_equipment_details(self) -> tuple[str, str, str] | None:
        """Return (machine_id, machine_type, calibration_due_date) for current selection."""

        if not hasattr(self, "calibrated_equipment_combo") or self.calibrated_equipment_combo is None:
            return None

        try:
            data = self.calibrated_equipment_combo.currentData()
            if isinstance(data, tuple) and len(data) == 3:
                mid, mtype, due = data
                return (str(mid or "").strip(), str(mtype or "").strip(), str(due or "").strip())
        except Exception:
            pass

        selected_name = str(self.calibrated_equipment_combo.currentText() or "").strip()
        if not selected_name:
            return None

        for name, mid, mtype, due in self._calibrated_equipment_rows_for_dropdown():
            if str(name or "").strip().lower() == selected_name.lower():
                return (str(mid or "").strip(), str(mtype or "").strip(), str(due or "").strip())

        return None

    def _on_calibrated_equipment_combo_changed(self) -> None:
        if not hasattr(self, "calibrated_equipment_combo") or self.calibrated_equipment_combo is None:
            return

        name = str(self.calibrated_equipment_combo.currentText() or "").strip()
        try:
            self._settings.setValue("inputs/calibrated_equipment_machine", name)
        except Exception:
            pass

        # Only apply when a Calypso file is selected (per request).
        if not str(getattr(self, "chr_path_edit", QLineEdit()).text() or "").strip():
            return
        if self._template_wb is None:
            return

        try:
            self._apply_selected_calibrated_equipment_to_workbook()
        except Exception:
            return

        # Re-render relevant forms so the user sees the change immediately.
        try:
            for key in ("2", "2c", "3"):
                name = self._form_sheet_names.get(key)
                viewer = self._form_viewers.get(key)
                if name and viewer and name in self._template_wb.sheetnames:
                    if key == "3" and self.characteristics:
                        try:
                            self._write_form3_to_worksheet(self._template_wb[name])
                        except Exception:
                            pass
                    viewer.set_overrides({})
                    viewer.render()
        except Exception:
            pass

    @staticmethod
    def _norm_label(text: object) -> str:
        if text is None:
            return ""
        s = str(text).strip().lower()
        # Normalize common punctuation and spacing.
        s = re.sub(r"[\s\t\r\n]+", " ", s)
        s = re.sub(r"[^a-z0-9 ]+", "", s)
        s = s.replace(" ", "")
        return s

    def _apply_selected_calibrated_equipment_to_workbook(self) -> None:
        """Write selected equipment details into the template workbook.

        - Machine ID + Calibration Due Date -> "Designed/Qualified tooling" field
        - Machine Type -> first column under "Additional Data/Comments"
        """

        if self._template_wb is None:
            return
        details = self._selected_calibrated_equipment_details()
        if not details:
            return
        machine_id, machine_type, due_date = details

        # Form 3 (and some Form 2 templates) display tooling like:
        #   Gage ID: <id>  Cal Due: <date>
        # Convert older persisted prefixes ("ID:" / "Due:") if present.
        mid_clean = re.sub(r"^\s*(?:gage\s*id|id)\s*:\s*", "", str(machine_id or ""), flags=re.IGNORECASE).strip()
        due_clean = re.sub(r"^\s*(?:cal\s*due|due)\s*:\s*", "", str(due_date or ""), flags=re.IGNORECASE).strip()
        tooling_parts: list[str] = []
        if mid_clean:
            tooling_parts.append(f"Gage ID: {mid_clean}")
        if due_clean:
            tooling_parts.append(f"Cal Due: {due_clean}")
        tooling_text = "  ".join(tooling_parts).strip()

        def is_tooling_label(v: object) -> bool:
            n = self._norm_label(v)
            # Handle variants like "Designed/Qualified tooling" or typos like "Designed?Qualified".
            return ("tooling" in n) and ("designed" in n) and ("qualified" in n)

        def is_additional_comments_header(v: object) -> bool:
            n = self._norm_label(v)
            # Handle typo "Addtion" too.
            return ("additionaldata" in n or "addtiondata" in n) and ("comment" in n)

        def cell_in_merged_range(ws, row: int, col: int):
            try:
                coord = ws.cell(row=row, column=col).coordinate
                for mr in getattr(ws, "merged_cells", []).ranges:
                    if coord in mr:
                        return mr
            except Exception:
                return None
            return None

        def merged_top_left(ws, row: int, col: int) -> tuple[int, int]:
            mr = cell_in_merged_range(ws, row, col)
            if mr is None:
                return (row, col)
            try:
                return (mr.min_row, mr.min_col)
            except Exception:
                return (row, col)

        def looks_like_field_number(v: object, n: int) -> bool:
            if v is None:
                return False
            s = str(v).strip().lower()
            # Common patterns: "10", "10.", "10)"
            return s == str(n) or s.startswith(f"{n}.") or s.startswith(f"{n})")

        def row_has_keywords(ws, row: int, keywords: list[str], max_col: int) -> int:
            score = 0
            for cc in range(1, max_col + 1):
                v = ws.cell(row=row, column=cc).value
                n = self._norm_label(v)
                if not n:
                    continue
                hit = 0
                for kw in keywords:
                    if kw in n:
                        hit += 1
                score = max(score, hit)
            return score

        def find_numbered_field_write_cell(ws, field_no: int, keywords: list[str]) -> tuple[int, int] | None:
            """Find a best-effort input cell for a numbered field.

            This handles templates that split the number ("10.") and the label text across cells.
            """

            max_row = min(getattr(ws, "max_row", 0) or 0, 250)
            max_col = min(getattr(ws, "max_column", 0) or 0, 60)
            best: tuple[int, int, int] | None = None  # (score, row, number_col)

            for rr in range(1, max_row + 1):
                for cc in range(1, min(max_col, 10) + 1):
                    if not looks_like_field_number(ws.cell(row=rr, column=cc).value, field_no):
                        continue
                    score = row_has_keywords(ws, rr, keywords, max_col)
                    if score <= 0:
                        continue
                    if best is None or score > best[0]:
                        best = (score, rr, cc)

            if best is None:
                return None

            _score, rr, num_col = best
            # Find the label cell on this row that contains most keywords.
            best_label_col = num_col
            best_label_hits = 0
            for cc in range(1, max_col + 1):
                v = ws.cell(row=rr, column=cc).value
                n = self._norm_label(v)
                if not n:
                    continue
                hits = sum(1 for kw in keywords if kw in n)
                if hits > best_label_hits:
                    best_label_hits = hits
                    best_label_col = cc

            # Start searching for an input box to the right of the label area.
            start_col = best_label_col
            mr = cell_in_merged_range(ws, rr, best_label_col)
            if mr is not None:
                try:
                    start_col = mr.max_col
                except Exception:
                    start_col = best_label_col

            for off in range(1, 20):
                tc = start_col + off
                if tc > max_col:
                    break
                tcell = ws.cell(row=rr, column=tc)
                # Prefer a blank cell or a merged input box.
                if tcell.value is None or str(tcell.value).strip() == "" or cell_in_merged_range(ws, rr, tc) is not None:
                    tr, tc2 = merged_top_left(ws, rr, tc)
                    return (tr, tc2)

            # Fallback: immediate right of the number cell.
            tr, tc2 = merged_top_left(ws, rr, min(num_col + 1, max_col))
            return (tr, tc2)

        def write_next_to_label(ws, label_pred, value: str) -> bool:
            if not value:
                return False
            max_row = min(getattr(ws, "max_row", 0) or 0, 250)
            max_col = min(getattr(ws, "max_column", 0) or 0, 40)
            for rr in range(1, max_row + 1):
                for cc in range(1, max_col + 1):
                    cell = ws.cell(row=rr, column=cc)
                    if not label_pred(cell.value):
                        continue
                    # Try a few cells to the right for a value slot.
                    for off in range(1, 7):
                        tc = cc + off
                        if tc > max_col:
                            break
                        target = ws.cell(row=rr, column=tc)
                        tv = target.value
                        if tv is None or str(tv).strip() == "":
                            target.value = value
                            return True
                    # Fallback: immediate right.
                    ws.cell(row=rr, column=min(cc + 1, max_col)).value = value
                    return True
            return False

        def write_under_header_first_col(ws, header_pred, value: str) -> bool:
            if not value:
                return False
            max_row = min(getattr(ws, "max_row", 0) or 0, 250)
            max_col = min(getattr(ws, "max_column", 0) or 0, 60)
            for rr in range(1, max_row + 1):
                for cc in range(1, max_col + 1):
                    cell = ws.cell(row=rr, column=cc)
                    if not header_pred(cell.value):
                        continue

                    base_col = cc
                    # If the header is merged across columns, use the left-most column.
                    try:
                        for mr in getattr(ws, "merged_cells", []).ranges:
                            if cell.coordinate in mr:
                                base_col = mr.min_col
                                break
                    except Exception:
                        base_col = cc

                    for r2 in range(rr + 1, min(rr + 30, max_row) + 1):
                        t = ws.cell(row=r2, column=base_col)
                        if t.value is None or str(t.value).strip() == "":
                            t.value = value
                            return True
                    # Fallback: write directly below header.
                    ws.cell(row=min(rr + 1, max_row), column=base_col).value = value
                    return True
            return False

        # Apply to likely form sheets first; then fall back to scanning all sheets.
        sheet_candidates: list[str] = []
        # NOTE: Do not apply this to Form 3. Form 3 uses per-row columns for tooling/comments,
        # and writing into the header area can land in unintended merged cells (e.g. O4/P5).
        for key in ("2", "2c", "1"):
            nm = self._form_sheet_names.get(key)
            if nm and nm in self._template_wb.sheetnames:
                sheet_candidates.append(nm)
        # De-dup while preserving order.
        sheet_candidates = list(dict.fromkeys(sheet_candidates))
        if not sheet_candidates:
            sheet_candidates = list(self._template_wb.sheetnames)

        wrote_any = False
        for sname in sheet_candidates:
            try:
                ws = self._template_wb[sname]
            except Exception:
                continue
            # First try label-based matching.
            if tooling_text:
                wrote_any = write_next_to_label(ws, is_tooling_label, tooling_text) or wrote_any
            if machine_type:
                wrote_any = write_under_header_first_col(ws, is_additional_comments_header, machine_type) or wrote_any

            # Then try numbered-field matching (AS9102 Form 2 style):
            # 10. Designed/Qualified Tooling
            if tooling_text and not wrote_any:
                loc = find_numbered_field_write_cell(ws, 10, ["designed", "qualified", "tooling"])
                if loc is not None:
                    r0, c0 = loc
                    ws.cell(row=r0, column=c0).value = tooling_text
                    wrote_any = True

            # 12. Additional Data/Comments
            if machine_type and not wrote_any:
                loc = find_numbered_field_write_cell(ws, 12, ["additional", "data", "comment"])
                if loc is not None:
                    r0, c0 = loc
                    ws.cell(row=r0, column=c0).value = machine_type
                    wrote_any = True

        return
        
        # Enable drops for the window
        self.setAcceptDrops(True)

    def _create_suppliers_tab(self) -> QWidget:
        container = QWidget()
        layout = QVBoxLayout(container)
        layout.setContentsMargins(6, 6, 6, 6)
        layout.setSpacing(6)

        header = QWidget()
        header_layout = QHBoxLayout(header)
        header_layout.setContentsMargins(0, 0, 0, 0)

        add_btn = QPushButton("Add")
        del_btn = QPushButton("Delete")
        header_layout.addWidget(add_btn)
        header_layout.addWidget(del_btn)
        header_layout.addStretch()
        layout.addWidget(header)

        table = QTableWidget()
        table.setColumnCount(2)
        table.setHorizontalHeaderLabels(["Customer", "Supplier Code"])
        table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        # Allow user to drag-resize columns and row heights.
        table.horizontalHeader().setStretchLastSection(False)
        table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        table.verticalHeader().setDefaultSectionSize(24)
        table.resizeColumnsToContents()
        table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        table.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        table.setEditTriggers(
            QAbstractItemView.EditTrigger.DoubleClicked
            | QAbstractItemView.EditTrigger.EditKeyPressed
            | QAbstractItemView.EditTrigger.AnyKeyPressed
        )

        self._install_delete_clears_cells(table)
        layout.addWidget(table)

        self._register_persistent_table(table, "tables/customer")

        self._suppliers_table = table

        def _add_row() -> None:
            if self._suppliers_table is None:
                return
            table = self._suppliers_table
            try:
                table.blockSignals(True)
                r = table.rowCount()
                table.insertRow(r)
                for c in range(2):
                    table.setItem(r, c, QTableWidgetItem(""))
            finally:
                table.blockSignals(False)

            table.setCurrentCell(r, 0)
            table.editItem(table.item(r, 0))

        def _delete_rows() -> None:
            if self._suppliers_table is None:
                return
            rows = sorted({idx.row() for idx in self._suppliers_table.selectionModel().selectedRows()}, reverse=True)
            if not rows:
                return
            for r in rows:
                self._suppliers_table.removeRow(r)
            self._save_suppliers_from_tab()

        add_btn.clicked.connect(_add_row)
        del_btn.clicked.connect(_delete_rows)
        table.itemChanged.connect(lambda _it: self._save_suppliers_from_tab())

        return container

    def _create_supplier_directory_tab(self) -> QWidget:
        container = QWidget()
        layout = QVBoxLayout(container)
        layout.setContentsMargins(6, 6, 6, 6)
        layout.setSpacing(6)

        header = QWidget()
        header_layout = QHBoxLayout(header)
        header_layout.setContentsMargins(0, 0, 0, 0)

        add_btn = QPushButton("Add")
        del_btn = QPushButton("Delete")
        header_layout.addWidget(add_btn)
        header_layout.addWidget(del_btn)
        header_layout.addStretch()
        layout.addWidget(header)

        table = QTableWidget()
        table.setColumnCount(6)
        table.setHorizontalHeaderLabels(["Company", "Address 1", "Address 2", "City", "State", "Zip Code"])
        table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)
        # Allow user to drag-resize columns and row heights.
        table.horizontalHeader().setStretchLastSection(False)
        table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        table.verticalHeader().setDefaultSectionSize(24)
        table.resizeColumnsToContents()
        table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        table.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        table.setEditTriggers(
            QAbstractItemView.EditTrigger.DoubleClicked
            | QAbstractItemView.EditTrigger.EditKeyPressed
            | QAbstractItemView.EditTrigger.AnyKeyPressed
        )

        self._install_delete_clears_cells(table)
        layout.addWidget(table)

        self._register_persistent_table(table, "tables/supplier")

        self._supplier_directory_table = table

        class _StateDelegate(QHeaderView):
            pass

        from PySide6.QtWidgets import QStyledItemDelegate, QComboBox

        class _StateComboDelegate(QStyledItemDelegate):
            def createEditor(self, parent, option, index):
                combo = QComboBox(parent)
                combo.setEditable(True)
                combo.addItem("")
                combo.addItems(US_STATE_CODES)
                QTimer.singleShot(0, combo.showPopup)
                return combo

            def setEditorData(self, editor, index):
                try:
                    txt = str(index.data(Qt.ItemDataRole.DisplayRole) or "").strip().upper()
                    i = editor.findText(txt)
                    if i >= 0:
                        editor.setCurrentIndex(i)
                    else:
                        editor.setCurrentText(txt)
                except Exception:
                    return

            def setModelData(self, editor, model, index):
                try:
                    txt = str(editor.currentText() or "").strip().upper()
                    model.setData(index, txt)
                except Exception:
                    return

        # State column (index 4)
        table.setItemDelegateForColumn(4, _StateComboDelegate(table))

        def _add_row() -> None:
            if self._supplier_directory_table is None:
                return
            table = self._supplier_directory_table
            try:
                table.blockSignals(True)
                r = table.rowCount()
                table.insertRow(r)
                for c in range(6):
                    table.setItem(r, c, QTableWidgetItem(""))
            finally:
                table.blockSignals(False)

            table.setCurrentCell(r, 0)
            table.editItem(table.item(r, 0))

        def _delete_rows() -> None:
            if self._supplier_directory_table is None:
                return
            rows = sorted({idx.row() for idx in self._supplier_directory_table.selectionModel().selectedRows()}, reverse=True)
            if not rows:
                return
            for r in rows:
                self._supplier_directory_table.removeRow(r)
            self._save_supplier_directory_from_tab()

        add_btn.clicked.connect(_add_row)
        del_btn.clicked.connect(_delete_rows)
        table.itemChanged.connect(lambda _it: self._save_supplier_directory_from_tab())

        return container

    def _create_calibrated_equipment_tab(self) -> QWidget:
        container = QWidget()
        layout = QVBoxLayout(container)
        layout.setContentsMargins(6, 6, 6, 6)
        layout.setSpacing(6)

        header = QWidget()
        header_layout = QHBoxLayout(header)
        header_layout.setContentsMargins(0, 0, 0, 0)

        add_btn = QPushButton("Add")
        del_btn = QPushButton("Delete")
        header_layout.addWidget(add_btn)
        header_layout.addWidget(del_btn)
        header_layout.addStretch()
        layout.addWidget(header)

        table = QTableWidget()
        table.setColumnCount(4)
        table.setHorizontalHeaderLabels(["Machine Name", "ID", "Machine Type", "Calibration Due Date"])
        table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        # Allow user to drag-resize columns and row heights.
        table.horizontalHeader().setStretchLastSection(False)
        table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        table.verticalHeader().setDefaultSectionSize(24)
        table.resizeColumnsToContents()
        table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        table.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        table.setEditTriggers(
            QAbstractItemView.EditTrigger.DoubleClicked
            | QAbstractItemView.EditTrigger.EditKeyPressed
            | QAbstractItemView.EditTrigger.AnyKeyPressed
        )

        self._install_delete_clears_cells(table)
        layout.addWidget(table)

        self._register_persistent_table(table, "tables/calibrated_equipment")

        self._calibrated_equipment_table = table

        def _add_row() -> None:
            if self._calibrated_equipment_table is None:
                return
            table = self._calibrated_equipment_table
            try:
                table.blockSignals(True)
                r = table.rowCount()
                table.insertRow(r)
                for c in range(4):
                    table.setItem(r, c, QTableWidgetItem(""))
            finally:
                table.blockSignals(False)

            table.setCurrentCell(r, 0)
            table.editItem(table.item(r, 0))

        def _delete_rows() -> None:
            if self._calibrated_equipment_table is None:
                return
            rows = sorted({idx.row() for idx in self._calibrated_equipment_table.selectionModel().selectedRows()}, reverse=True)
            if not rows:
                return
            for r in rows:
                self._calibrated_equipment_table.removeRow(r)
            self._save_calibrated_equipment_from_tab()

        add_btn.clicked.connect(_add_row)
        del_btn.clicked.connect(_delete_rows)
        table.itemChanged.connect(lambda _it: self._save_calibrated_equipment_from_tab())

        return container

    def _register_persistent_table(self, table: QTableWidget, key: str) -> None:
        """Persist column widths + row heights for a QTableWidget.

        Stored in QSettings under `{key}/sizes`.
        """
        if table is None:
            return
        key = str(key or "").strip()
        if not key:
            return

        # Avoid double-connecting if called twice.
        try:
            if bool(table.property("_persist_sizes_registered")):
                return
        except Exception:
            pass
        try:
            table.setProperty("_persist_sizes_registered", True)
            table.setProperty("_persist_sizes_key", key)
        except Exception:
            return

        self._restore_persistent_table_sizes(table, key)

        def _schedule() -> None:
            self._schedule_persist_table_sizes(table, key)

        try:
            table.horizontalHeader().sectionResized.connect(lambda *_a: _schedule())
        except Exception:
            pass
        try:
            table.verticalHeader().sectionResized.connect(lambda *_a: _schedule())
        except Exception:
            pass

    def _install_delete_clears_cells(self, table: QTableWidget) -> None:
        """Ensure Delete clears contents for the provided table."""
        if table is None:
            return
        try:
            if getattr(self, "_delete_clears_cells_filter", None) is None:
                self._delete_clears_cells_filter = _DeleteClearsTableCellsFilter(self)
            # Avoid double install.
            if bool(table.property("_delete_clears_cells_installed")):
                return
            table.setProperty("_delete_clears_cells_installed", True)
            table.installEventFilter(self._delete_clears_cells_filter)
        except Exception:
            return

    def _schedule_persist_table_sizes(self, table: QTableWidget, key: str) -> None:
        key = str(key or "").strip()
        if not key:
            return
        timer = self._table_persist_timers.get(key)
        if timer is None:
            timer = QTimer(self)
            timer.setSingleShot(True)
            timer.setInterval(250)
            timer.timeout.connect(lambda t=table, k=key: self._save_persistent_table_sizes(t, k))
            self._table_persist_timers[key] = timer
        timer.start()

    def _save_persistent_table_sizes(self, table: QTableWidget, key: str) -> None:
        if table is None:
            return
        try:
            payload = {
                "cols": [int(table.columnWidth(i)) for i in range(int(table.columnCount()))],
                "rows": [int(table.rowHeight(i)) for i in range(int(table.rowCount()))],
            }
            self._settings.setValue(f"{key}/sizes", json.dumps(payload))
        except Exception:
            return

    def _restore_persistent_table_sizes(self, table: QTableWidget, key: str) -> None:
        if table is None:
            return
        try:
            raw = self._settings.value(f"{key}/sizes", "", type=str)
        except Exception:
            raw = ""
        if not raw:
            return
        try:
            payload = json.loads(str(raw))
        except Exception:
            return
        if not isinstance(payload, dict):
            return

        cols = payload.get("cols")
        rows = payload.get("rows")
        if isinstance(cols, list):
            for i in range(min(len(cols), int(table.columnCount()))):
                try:
                    w = int(cols[i])
                except Exception:
                    continue
                if w > 0:
                    table.setColumnWidth(i, w)
        if isinstance(rows, list):
            for i in range(min(len(rows), int(table.rowCount()))):
                try:
                    h = int(rows[i])
                except Exception:
                    continue
                if h > 0:
                    table.setRowHeight(i, h)

    def _refresh_all_viewer_validations(self) -> None:
        for v in self._form_viewers.values():
            if v is not None:
                try:
                    v.refresh_validations()
                except Exception:
                    pass
        for win in list(self._popout_windows):
            try:
                for v in win.findChildren(ExcelSheetViewer):
                    v.refresh_validations()
            except Exception:
                continue

    def _supplier_master_sheet(self, wb):
        """Hidden sheet used for both supplier-code and supplier-directory dropdowns.

        Columns:
          A: Company
          B: Address
          C: Supplier Code
        """
        name = "__as9102_supplier_master"
        created = False
        if name in wb.sheetnames:
            ws = wb[name]
        else:
            ws = wb.create_sheet(name)
            created = True
        try:
            ws.sheet_state = "hidden"
        except Exception:
            pass

        # Seed only if empty/new.
        a1 = ws.cell(row=1, column=1).value
        b1 = ws.cell(row=1, column=2).value
        c1 = ws.cell(row=1, column=3).value
        is_empty = all(v is None or str(v).strip() == "" for v in (a1, b1, c1))
        if created or is_empty:
            # Merge any existing legacy hidden sheets if present.
            directory_rows: dict[str, str] = {}
            code_rows: dict[str, str] = {}

            if "__as9102_supplier_directory" in wb.sheetnames:
                sd = wb["__as9102_supplier_directory"]
                for rr in range(1, min((getattr(sd, "max_row", 0) or 0), 5000) + 1):
                    comp = sd.cell(row=rr, column=1).value
                    addr = sd.cell(row=rr, column=2).value
                    comp_s = str(comp).strip() if comp is not None else ""
                    addr_s = str(addr).strip() if addr is not None else ""
                    if comp_s and addr_s:
                        directory_rows[comp_s] = addr_s

            if "__as9102_suppliers" in wb.sheetnames:
                sc = wb["__as9102_suppliers"]
                for rr in range(1, min((getattr(sc, "max_row", 0) or 0), 5000) + 1):
                    code = sc.cell(row=rr, column=1).value
                    comp = sc.cell(row=rr, column=2).value
                    comp_s = str(comp).strip() if comp is not None else ""
                    code_s = str(code).strip() if code is not None else ""
                    if comp_s and code_s:
                        code_rows[comp_s] = code_s

            # If nothing existed, seed with the original D9 defaults.
            if not directory_rows and not code_rows:
                code_rows = {
                    "Raytheon": "10033672",
                    "DRS": "10V001518",
                    "SAES Getters": "Camtron Incorporated",
                    "Vallen": "497735",
                }

            companies = sorted(set(directory_rows.keys()) | set(code_rows.keys()), key=lambda s: s.lower())
            for i, comp in enumerate(companies, start=1):
                ws.cell(row=i, column=1).value = comp
                ws.cell(row=i, column=2).value = directory_rows.get(comp, "")
                ws.cell(row=i, column=3).value = code_rows.get(comp, "")

        return ws

    def _load_persistent_customer_rows(self) -> list[tuple[str, str]]:
        """Load persisted Customer rows as [(customer, code)]."""
        try:
            raw = self._settings.value("lists/customer_rows", "", type=str)
        except Exception:
            raw = ""
        if not raw:
            return []
        try:
            data = json.loads(raw)
        except Exception:
            return []
        rows: list[tuple[str, str]] = []
        if isinstance(data, list):
            for it in data:
                if not isinstance(it, dict):
                    continue
                customer = str(it.get("customer", "") or "").strip()
                code = str(it.get("code", "") or "").strip()
                if not (customer or code):
                    continue
                rows.append((customer, code))
        # De-dupe by customer name (case-insensitive), keep first occurrence.
        out: list[tuple[str, str]] = []
        seen: set[str] = set()
        for customer, code in rows:
            k = (customer or "").strip().lower()
            if k and k in seen:
                continue
            if k:
                seen.add(k)
            out.append((customer, code))
        return out

    def _save_persistent_customer_rows(self, rows: list[tuple[str, str]]) -> None:
        """Persist Customer rows as JSON in QSettings."""
        payload = []
        for customer, code in rows:
            customer = str(customer or "").strip()
            code = str(code or "").strip()
            if not (customer or code):
                continue
            payload.append({"customer": customer, "code": code})
        try:
            self._settings.setValue("lists/customer_rows", json.dumps(payload))
        except Exception:
            pass

    def _load_persistent_supplier_directory_rows(self) -> list[tuple[str, str, str, str, str, str]]:
        """Load persisted Supplier rows as [(company, addr1, addr2, city, state, zip)]."""
        try:
            raw = self._settings.value("lists/supplier_directory_rows", "", type=str)
        except Exception:
            raw = ""
        if not raw:
            return []
        try:
            data = json.loads(raw)
        except Exception:
            return []
        rows: list[tuple[str, str, str, str, str, str]] = []
        if isinstance(data, list):
            for it in data:
                if not isinstance(it, dict):
                    continue
                company = str(it.get("company", "") or "").strip()
                # Backward-compat: old key "address" stored full address in one field.
                full = str(it.get("address", "") or "").strip()
                addr1 = str(it.get("addr1", "") or "").strip() or full
                addr2 = str(it.get("addr2", "") or "").strip()
                addr3 = str(it.get("addr3", "") or "").strip()
                city = str(it.get("city", "") or "").strip()
                state = str(it.get("state", "") or "").strip().upper()
                zipc = str(it.get("zip", "") or "").strip()
                if full and (not city and not state and not zipc and not addr2 and not addr3):
                    a1, a2, a3, ct, st, z = _split_address_lines(full)
                    addr1, addr2, addr3, city, state, zipc = a1, a2, a3, ct, st, z

                # Fold legacy Address3 into Address2 to match screenshot.
                if addr3:
                    addr2 = f"{addr2}, {addr3}" if addr2 else addr3

                if not (company or addr1 or addr2 or city or state or zipc):
                    continue
                rows.append((company, addr1, addr2, city, state, zipc))
        # De-dupe by company name (case-insensitive), keep first occurrence.
        out: list[tuple[str, str, str, str, str, str]] = []
        seen: set[str] = set()
        for company, addr1, addr2, city, state, zipc in rows:
            k = (company or "").strip().lower()
            if k and k in seen:
                continue
            if k:
                seen.add(k)
            out.append((company, addr1, addr2, city, state, zipc))
        return out

    def _save_persistent_supplier_directory_rows(self, rows: list[tuple[str, str, str, str, str, str]]) -> None:
        """Persist Supplier rows as JSON in QSettings."""
        payload = []
        for company, addr1, addr2, city, state, zipc in rows:
            company = str(company or "").strip()
            addr1 = str(addr1 or "").strip()
            addr2 = str(addr2 or "").strip()
            city = str(city or "").strip()
            state = str(state or "").strip().upper()
            zipc = str(zipc or "").strip()
            if not (company or addr1 or addr2 or city or state or zipc):
                continue
            payload.append(
                {
                    "company": company,
                    "addr1": addr1,
                    "addr2": addr2,
                    "city": city,
                    "state": state,
                    "zip": zipc,
                }
            )
        try:
            self._settings.setValue("lists/supplier_directory_rows", json.dumps(payload))
        except Exception:
            pass

    def _load_persistent_calibrated_equipment_rows(self) -> list[tuple[str, str, str, str]]:
        """Load persisted calibrated equipment rows as [(name, id, type, due_date)]."""
        try:
            raw = self._settings.value("lists/calibrated_equipment_rows", "", type=str)
        except Exception:
            raw = ""
        if not raw:
            return []
        try:
            data = json.loads(raw)
        except Exception:
            return []
        rows: list[tuple[str, str, str, str]] = []
        if isinstance(data, list):
            for it in data:
                if not isinstance(it, dict):
                    continue
                name = str(it.get("machine_name", "") or "").strip()
                mid = str(it.get("id", "") or "").strip()
                mtype = str(it.get("machine_type", "") or "").strip()
                due = str(it.get("cal_due_date", "") or "").strip()
                if not (name or mid or mtype or due):
                    continue
                rows.append((name, mid, mtype, due))

        # De-dupe by ID if present, else by name (case-insensitive). Keep first occurrence.
        out: list[tuple[str, str, str, str]] = []
        seen: set[str] = set()
        for name, mid, mtype, due in rows:
            k = (mid or name or "").strip().lower()
            if k and k in seen:
                continue
            if k:
                seen.add(k)
            out.append((name, mid, mtype, due))
        return out

    def _save_persistent_calibrated_equipment_rows(self, rows: list[tuple[str, str, str, str]]) -> None:
        """Persist calibrated equipment rows as JSON in QSettings."""
        payload = []
        for name, mid, mtype, due in rows:
            name = str(name or "").strip()
            mid = str(mid or "").strip()
            mtype = str(mtype or "").strip()
            due = str(due or "").strip()
            if not (name or mid or mtype or due):
                continue
            payload.append(
                {
                    "machine_name": name,
                    "id": mid,
                    "machine_type": mtype,
                    "cal_due_date": due,
                }
            )
        try:
            self._settings.setValue("lists/calibrated_equipment_rows", json.dumps(payload))
        except Exception:
            pass

    def _apply_persistent_lists_to_workbook(self) -> None:
        """Apply persisted Customer/Supplier lists into the currently loaded workbook."""
        if self._template_wb is None:
            return

        # Customer list -> master sheet (A=Customer, C=Code). Preserve existing Address column values.
        persisted_customers = self._load_persistent_customer_rows()
        if persisted_customers:
            ws = self._supplier_master_sheet(self._template_wb)
            # Build an address map from existing master rows.
            addr_by_key: dict[str, str] = {}
            max_scan = min(max(getattr(ws, "max_row", 0) or 0, 0), 5000)
            for rr in range(1, max_scan + 1):
                comp = ws.cell(row=rr, column=1).value
                addr = ws.cell(row=rr, column=2).value
                comp_s = str(comp).strip() if comp is not None else ""
                addr_s = str(addr).strip() if addr is not None else ""
                if comp_s:
                    addr_by_key[comp_s.lower()] = addr_s

            # Clear then write.
            rows = sorted(persisted_customers, key=lambda x: (x[0] or "").lower())
            clear_to = max(len(rows) + 20, 50)
            for rr in range(1, clear_to + 1):
                ws.cell(row=rr, column=1).value = None
                ws.cell(row=rr, column=2).value = None
                ws.cell(row=rr, column=3).value = None
            for i, (customer, code) in enumerate(rows, start=1):
                ws.cell(row=i, column=1).value = customer
                ws.cell(row=i, column=2).value = addr_by_key.get(customer.lower(), "")
                ws.cell(row=i, column=3).value = code

            self._apply_supplier_master_validations()

        # Supplier directory list -> supplier directory sheet
        persisted_suppliers = self._load_persistent_supplier_directory_rows()
        if persisted_suppliers:
            sheet_name, _last = self._ensure_supplier_directory_sheet(self._template_wb)
            ws = self._template_wb[sheet_name]
            rows = list(persisted_suppliers)
            clear_to = max(len(rows) + 20, 50)
            for rr in range(1, clear_to + 1):
                for cc in range(1, 9):
                    ws.cell(row=rr, column=cc).value = None
            for i, (company, addr1, addr2, city, state, zipc) in enumerate(rows, start=1):
                addr1 = _clean_company_prefix(company, addr1)
                ws.cell(row=i, column=1).value = company
                ws.cell(row=i, column=2).value = addr1
                ws.cell(row=i, column=3).value = addr2
                ws.cell(row=i, column=4).value = city
                ws.cell(row=i, column=5).value = state
                ws.cell(row=i, column=6).value = zipc
                ws.cell(row=i, column=7).value = _build_full_address_with_company(company, addr1, addr2, city, state, zipc)
            # Mark as migrated/seeded so we don't re-import defaults on every load.
            ws.cell(row=1, column=8).value = "seeded_v4_reset"

            # Re-apply dropdowns.
            form1 = self._form_sheet_names.get("1")
            form2 = self._form_sheet_names.get("2")
            if form1 and form1 in self._template_wb.sheetnames:
                self._ensure_supplier_directory_dropdown(self._template_wb[form1], cell_range="E15:E500")
            if form2 and form2 in self._template_wb.sheetnames:
                self._ensure_supplier_directory_dropdown(self._template_wb[form2], cell_range="F5:F500")

        # Calibrated equipment list -> calibrated equipment sheet
        persisted_equipment = self._load_persistent_calibrated_equipment_rows()
        if persisted_equipment:
            ws = self._calibrated_equipment_sheet(self._template_wb)
            if ws is not None:
                rows = list(persisted_equipment)
                clear_to = max(len(rows) + 20, 50)
                for rr in range(1, clear_to + 1):
                    for cc in range(1, 6):
                        ws.cell(row=rr, column=cc).value = None
                for i, (name, mid, mtype, due) in enumerate(rows, start=1):
                    ws.cell(row=i, column=1).value = name
                    ws.cell(row=i, column=2).value = mid
                    ws.cell(row=i, column=3).value = mtype
                    ws.cell(row=i, column=4).value = due
                ws.cell(row=1, column=5).value = "seeded_v1"

    def _calibrated_equipment_sheet(self, wb):
        """Hidden sheet used for calibrated equipment list."""
        name = "__as9102_calibrated_equipment"
        created = False
        if name in wb.sheetnames:
            ws = wb[name]
        else:
            ws = wb.create_sheet(name)
            created = True
        try:
            ws.sheet_state = "hidden"
        except Exception:
            pass

        # Seed if empty/new and nothing has been persisted yet.
        a1 = ws.cell(row=1, column=1).value
        b1 = ws.cell(row=1, column=2).value
        c1 = ws.cell(row=1, column=3).value
        d1 = ws.cell(row=1, column=4).value
        is_empty = all(v is None or str(v).strip() == "" for v in (a1, b1, c1, d1))
        if created or is_empty:
            persisted = self._load_persistent_calibrated_equipment_rows()
            rows = persisted or list(DEFAULT_CALIBRATED_EQUIPMENT_SEED)
            for rr in range(1, max(len(rows) + 20, 50) + 1):
                for cc in range(1, 6):
                    ws.cell(row=rr, column=cc).value = None
            for i, (mn, mid, mt, due) in enumerate(rows, start=1):
                ws.cell(row=i, column=1).value = mn
                ws.cell(row=i, column=2).value = mid
                ws.cell(row=i, column=3).value = mt
                ws.cell(row=i, column=4).value = due
            ws.cell(row=1, column=5).value = "seeded_v1"
            # Persist seed so it survives even if the workbook changes.
            try:
                if not persisted:
                    self._save_persistent_calibrated_equipment_rows(rows)
            except Exception:
                pass

        return ws

    def _supplier_directory_sheet(self, wb):
        """Hidden sheet used for Supplier dropdowns.

        Layout:
          A: Company
          B: Address
        """
        sheet_name, _last_row = self._ensure_supplier_directory_sheet(wb)
        try:
            return wb[sheet_name]
        except Exception:
            return None

    def _supplier_master_rows(self) -> list[tuple[str, str, str]]:
        if self._template_wb is None:
            return []
        ws = self._supplier_master_sheet(self._template_wb)
        rows: list[tuple[str, str, str]] = []
        for rr in range(1, min((getattr(ws, "max_row", 0) or 0), 5000) + 1):
            comp = ws.cell(row=rr, column=1).value
            addr = ws.cell(row=rr, column=2).value
            code = ws.cell(row=rr, column=3).value
            comp_s = str(comp).strip() if comp is not None else ""
            addr_s = str(addr).strip() if addr is not None else ""
            code_s = str(code).strip() if code is not None else ""
            if not (comp_s or addr_s or code_s):
                continue
            rows.append((comp_s, addr_s, code_s))
        return rows

    def _write_supplier_master_rows(self, rows: list[tuple[str, str, str]]) -> None:
        if self._template_wb is None:
            return
        ws = self._supplier_master_sheet(self._template_wb)

        # Clear a bit beyond current size.
        clear_to = max(len(rows) + 20, 50)
        for rr in range(1, clear_to + 1):
            ws.cell(row=rr, column=1).value = None
            ws.cell(row=rr, column=2).value = None
            ws.cell(row=rr, column=3).value = None

        for i, (comp, addr, code) in enumerate(rows, start=1):
            ws.cell(row=i, column=1).value = comp
            ws.cell(row=i, column=2).value = addr
            ws.cell(row=i, column=3).value = code

        self._apply_supplier_master_validations()
        self._refresh_all_viewer_validations()

        # Persist customer list automatically (customer + code).
        try:
            cust_rows = [(c, cd) for (c, _a, cd) in rows if (c or cd)]
            self._save_persistent_customer_rows(cust_rows)
        except Exception:
            pass

    def _apply_list_validation(self, ws, sqref: str, formula1: str) -> None:
        dvs = getattr(ws, "data_validations", None)
        if dvs is None:
            ws.data_validations = openpyxl.worksheet.datavalidation.DataValidationList()
            dvs = ws.data_validations

        # Update an existing list DV that targets this sqref, else create.
        for dv in list(getattr(dvs, "dataValidation", []) or []):
            if getattr(dv, "type", None) != "list":
                continue
            try:
                if sqref.split(":", 1)[0] in str(getattr(dv, "sqref", "")):
                    dv.formula1 = formula1
                    return
            except Exception:
                continue

        dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(sqref)

    def _apply_supplier_master_validations(self) -> None:
        if self._template_wb is None:
            return
        master = self._supplier_master_sheet(self._template_wb)
        last_code = 1
        max_scan = min(max(getattr(master, "max_row", 0) or 0, 50), 5000)
        for rr in range(1, max_scan + 1):
            c = master.cell(row=rr, column=3).value
            if c is not None and str(c).strip() != "":
                last_code = rr
        code_rng = f"=__as9102_supplier_master!$C$1:$C${last_code}"

        # Apply to Form 1 worksheet if present (customer->code dropdown).
        form1 = self._form_sheet_names.get("1")
        if form1 and form1 in self._template_wb.sheetnames:
            ws1 = self._template_wb[form1]
            self._apply_list_validation(ws1, "D9", code_rng)

    def _load_suppliers_tab_from_workbook(self) -> None:
        if self._suppliers_table is None:
            return
        self._suppliers_suppress_changes = True
        try:
            self._suppliers_table.setRowCount(0)
            if self._template_wb is None:
                self._suppliers_table.setEnabled(False)
                return
            self._suppliers_table.setEnabled(True)

            rows = self._supplier_master_rows()
            self._suppliers_table.setRowCount(len(rows))
            for r, (comp, addr, code) in enumerate(rows):
                self._suppliers_table.setItem(r, 0, QTableWidgetItem(comp))
                self._suppliers_table.setItem(r, 1, QTableWidgetItem(code))
        finally:
            self._suppliers_suppress_changes = False

    def _supplier_directory_rows(self) -> list[tuple[str, str, str, str, str, str]]:
        if self._template_wb is None:
            return []
        ws = self._supplier_directory_sheet(self._template_wb)
        if ws is None:
            return []

        # Detect legacy schema (old marker was stored in column I).
        legacy_schema = False
        try:
            for rr in range(1, min((getattr(ws, "max_row", 0) or 0), 10) + 1):
                v = ws.cell(row=rr, column=9).value
                if v is not None and str(v).strip() != "":
                    legacy_schema = True
                    break
        except Exception:
            legacy_schema = False

        rows: list[tuple[str, str, str, str, str, str]] = []
        for rr in range(1, min((getattr(ws, "max_row", 0) or 0), 5000) + 1):
            comp = ws.cell(row=rr, column=1).value
            addr1 = ws.cell(row=rr, column=2).value
            addr2 = ws.cell(row=rr, column=3).value

            if legacy_schema:
                addr3 = ws.cell(row=rr, column=4).value
                city = ws.cell(row=rr, column=5).value
                state = ws.cell(row=rr, column=6).value
                zipc = ws.cell(row=rr, column=7).value
            else:
                addr3 = None
                city = ws.cell(row=rr, column=4).value
                state = ws.cell(row=rr, column=5).value
                zipc = ws.cell(row=rr, column=6).value

            comp_s = str(comp).strip() if comp is not None else ""
            a1_s = str(addr1).strip() if addr1 is not None else ""
            a2_s = str(addr2).strip() if addr2 is not None else ""
            a3_s = str(addr3).strip() if addr3 is not None else ""
            city_s = str(city).strip() if city is not None else ""
            state_s = str(state).strip().upper() if state is not None else ""
            zip_s = str(zipc).strip() if zipc is not None else ""

            # Backward compat: if only column B is populated and others are empty,
            # treat it as a single-line address and split it.
            if a1_s and not (a2_s or a3_s or city_s or state_s or zip_s):
                a1_clean = _clean_company_prefix(comp_s, a1_s)
                a1, a2, a3, ct, st, z = _split_address_lines(a1_clean)
                # Fold legacy address line 3 into Address 2 for screenshot schema.
                if a3:
                    a2 = f"{a2}, {a3}" if a2 else a3
                a1_s, a2_s, a3_s, city_s, state_s, zip_s = a1, a2, "", ct, st, z

            # Fold legacy Address 3 into Address 2.
            if a3_s:
                a2_s = f"{a2_s}, {a3_s}" if a2_s else a3_s
                a3_s = ""

            if not (comp_s or a1_s or a2_s or city_s or state_s or zip_s):
                continue
            rows.append((comp_s, a1_s, a2_s, city_s, state_s, zip_s))
        return rows

    def _write_supplier_directory_rows(self, rows: list[tuple[str, str, str, str, str, str]]) -> None:
        if self._template_wb is None:
            return
        ws = self._supplier_directory_sheet(self._template_wb)
        if ws is None:
            return

        clear_to = max(len(rows) + 20, 50)
        for rr in range(1, clear_to + 1):
            for cc in range(1, 9):
                ws.cell(row=rr, column=cc).value = None

        for i, (comp, addr1, addr2, city, state, zipc) in enumerate(rows, start=1):
            addr1 = _clean_company_prefix(comp, addr1)
            state = str(state or "").strip().upper()
            if state and state not in US_STATE_CODES:
                # Keep user value, but normalized.
                state = state

            # If user pasted a multi-line address into Address 1, split it.
            if addr1 and not (addr2 or city or state or zipc):
                a1, a2, a3, ct, st, z = _split_address_lines(addr1)
                if a3:
                    a2 = f"{a2}, {a3}" if a2 else a3
                addr1, addr2, city, state, zipc = a1, a2, ct, st, z

            ws.cell(row=i, column=1).value = comp
            ws.cell(row=i, column=2).value = addr1
            ws.cell(row=i, column=3).value = addr2
            ws.cell(row=i, column=4).value = city
            ws.cell(row=i, column=5).value = state
            ws.cell(row=i, column=6).value = zipc
            ws.cell(row=i, column=7).value = _build_full_address_with_company(comp, addr1, addr2, city, state, zipc)

        # Keep the one-time reset marker so we don't wipe user edits on next load.
        ws.cell(row=1, column=8).value = "seeded_v4_reset"

        # Re-apply dropdowns (Form 1 E15:E500, Form 2 F5:F500)
        form1 = self._form_sheet_names.get("1")
        form2 = self._form_sheet_names.get("2")
        if form1 and form1 in self._template_wb.sheetnames:
            self._ensure_supplier_directory_dropdown(self._template_wb[form1], cell_range="E15:E500")
        if form2 and form2 in self._template_wb.sheetnames:
            self._ensure_supplier_directory_dropdown(self._template_wb[form2], cell_range="F5:F500")

        self._refresh_all_viewer_validations()

        # Persist supplier directory automatically.
        try:
            self._save_persistent_supplier_directory_rows(rows)
        except Exception:
            pass

    def _load_supplier_directory_tab_from_workbook(self) -> None:
        if self._supplier_directory_table is None:
            return
        self._supplier_directory_suppress_changes = True
        try:
            self._supplier_directory_table.setRowCount(0)
            if self._template_wb is None:
                self._supplier_directory_table.setEnabled(False)
                return
            self._supplier_directory_table.setEnabled(True)

            rows = self._supplier_directory_rows()
            self._supplier_directory_table.setRowCount(len(rows))
            for r, (comp, addr1, addr2, city, state, zipc) in enumerate(rows):
                self._supplier_directory_table.setItem(r, 0, QTableWidgetItem(comp))
                self._supplier_directory_table.setItem(r, 1, QTableWidgetItem(addr1))
                self._supplier_directory_table.setItem(r, 2, QTableWidgetItem(addr2))
                self._supplier_directory_table.setItem(r, 3, QTableWidgetItem(city))
                self._supplier_directory_table.setItem(r, 4, QTableWidgetItem(state))
                self._supplier_directory_table.setItem(r, 5, QTableWidgetItem(zipc))
        finally:
            self._supplier_directory_suppress_changes = False

    def _save_supplier_directory_from_tab(self) -> None:
        if self._supplier_directory_suppress_changes:
            return
        if self._template_wb is None or self._supplier_directory_table is None:
            return

        rows: list[tuple[str, str, str, str, str, str]] = []
        for r in range(self._supplier_directory_table.rowCount()):
            comp = self._supplier_directory_table.item(r, 0).text().strip() if self._supplier_directory_table.item(r, 0) else ""
            addr1 = self._supplier_directory_table.item(r, 1).text().strip() if self._supplier_directory_table.item(r, 1) else ""
            addr2 = self._supplier_directory_table.item(r, 2).text().strip() if self._supplier_directory_table.item(r, 2) else ""
            city = self._supplier_directory_table.item(r, 3).text().strip() if self._supplier_directory_table.item(r, 3) else ""
            state = self._supplier_directory_table.item(r, 4).text().strip().upper() if self._supplier_directory_table.item(r, 4) else ""
            zipc = self._supplier_directory_table.item(r, 5).text().strip() if self._supplier_directory_table.item(r, 5) else ""
            if not (comp or addr1 or addr2 or city or state or zipc):
                continue
            rows.append((comp, addr1, addr2, city, state, zipc))

        self._write_supplier_directory_rows(rows)
        self._load_supplier_directory_tab_from_workbook()

    def _calibrated_equipment_rows(self) -> list[tuple[str, str, str, str]]:
        if self._template_wb is None:
            return []
        ws = self._calibrated_equipment_sheet(self._template_wb)
        if ws is None:
            return []

        rows: list[tuple[str, str, str, str]] = []
        for rr in range(1, min((getattr(ws, "max_row", 0) or 0), 5000) + 1):
            name = ws.cell(row=rr, column=1).value
            mid = ws.cell(row=rr, column=2).value
            mtype = ws.cell(row=rr, column=3).value
            due = ws.cell(row=rr, column=4).value
            name_s = str(name).strip() if name is not None else ""
            mid_s = str(mid).strip() if mid is not None else ""
            mtype_s = str(mtype).strip() if mtype is not None else ""
            due_s = str(due).strip() if due is not None else ""
            if not (name_s or mid_s or mtype_s or due_s):
                continue
            rows.append((name_s, mid_s, mtype_s, due_s))
        return rows

    def _write_calibrated_equipment_rows(self, rows: list[tuple[str, str, str, str]]) -> None:
        if self._template_wb is None:
            return
        ws = self._calibrated_equipment_sheet(self._template_wb)
        if ws is None:
            return

        clear_to = max(len(rows) + 20, 50)
        for rr in range(1, clear_to + 1):
            for cc in range(1, 6):
                ws.cell(row=rr, column=cc).value = None

        for i, (name, mid, mtype, due) in enumerate(rows, start=1):
            ws.cell(row=i, column=1).value = str(name or "").strip()
            ws.cell(row=i, column=2).value = str(mid or "").strip()
            ws.cell(row=i, column=3).value = str(mtype or "").strip()
            ws.cell(row=i, column=4).value = str(due or "").strip()

        ws.cell(row=1, column=5).value = "seeded_v1"

        try:
            self._save_persistent_calibrated_equipment_rows(rows)
        except Exception:
            pass

    def _load_calibrated_equipment_tab_from_workbook(self) -> None:
        if self._calibrated_equipment_table is None:
            return
        self._calibrated_equipment_suppress_changes = True
        try:
            self._calibrated_equipment_table.setRowCount(0)
            if self._template_wb is None:
                self._calibrated_equipment_table.setEnabled(False)
                return
            self._calibrated_equipment_table.setEnabled(True)

            rows = self._calibrated_equipment_rows()
            self._calibrated_equipment_table.setRowCount(len(rows))
            for r, (name, mid, mtype, due) in enumerate(rows):
                self._calibrated_equipment_table.setItem(r, 0, QTableWidgetItem(name))
                self._calibrated_equipment_table.setItem(r, 1, QTableWidgetItem(mid))
                self._calibrated_equipment_table.setItem(r, 2, QTableWidgetItem(mtype))
                self._calibrated_equipment_table.setItem(r, 3, QTableWidgetItem(due))
        finally:
            self._calibrated_equipment_suppress_changes = False

    def _save_calibrated_equipment_from_tab(self) -> None:
        if self._calibrated_equipment_suppress_changes:
            return
        if self._template_wb is None or self._calibrated_equipment_table is None:
            return

        rows: list[tuple[str, str, str, str]] = []
        for r in range(self._calibrated_equipment_table.rowCount()):
            name = self._calibrated_equipment_table.item(r, 0).text().strip() if self._calibrated_equipment_table.item(r, 0) else ""
            mid = self._calibrated_equipment_table.item(r, 1).text().strip() if self._calibrated_equipment_table.item(r, 1) else ""
            mtype = self._calibrated_equipment_table.item(r, 2).text().strip() if self._calibrated_equipment_table.item(r, 2) else ""
            due = self._calibrated_equipment_table.item(r, 3).text().strip() if self._calibrated_equipment_table.item(r, 3) else ""
            if not (name or mid or mtype or due):
                continue
            rows.append((name, mid, mtype, due))

        self._write_calibrated_equipment_rows(rows)
        self._load_calibrated_equipment_tab_from_workbook()

        # Keep Inputs dropdown in sync with table changes.
        try:
            self._refresh_calibrated_equipment_combo(preserve_selection=True)
        except Exception:
            pass

    def _save_suppliers_from_tab(self) -> None:
        if self._suppliers_suppress_changes:
            return
        if self._template_wb is None or self._suppliers_table is None:
            return

        # Preserve any existing address values in the master sheet.
        existing_addr_by_customer: dict[str, str] = {}
        try:
            for comp, addr, _code in self._supplier_master_rows():
                if comp:
                    existing_addr_by_customer[comp.strip().lower()] = addr
        except Exception:
            existing_addr_by_customer = {}

        rows: list[tuple[str, str, str]] = []
        missing_code = 0
        for r in range(self._suppliers_table.rowCount()):
            comp = self._suppliers_table.item(r, 0).text().strip() if self._suppliers_table.item(r, 0) else ""
            code = self._suppliers_table.item(r, 1).text().strip() if self._suppliers_table.item(r, 1) else ""
            if not (comp or code):
                continue
            if comp and not code:
                missing_code += 1
            addr = existing_addr_by_customer.get(comp.lower(), "")
            rows.append((comp, addr, code))

        if missing_code and not self._warned_missing_customer_code:
            self._warned_missing_customer_code = True
            try:
                QMessageBox.information(
                    self,
                    "Customer List",
                    "One or more Customer rows are missing a Supplier Code.\n"
                    "Those rows will not appear in the Form 1 D9 dropdown until a code is entered.",
                )
            except Exception:
                pass

        rows = sorted(rows, key=lambda x: (x[0] or "").lower())
        self._write_supplier_master_rows(rows)
        self._load_suppliers_tab_from_workbook()

    def _create_form_tab(self, form_key: str):
        container = QWidget()
        layout = QVBoxLayout(container)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(4)

        header = QWidget()
        header_layout = QHBoxLayout(header)
        header_layout.setContentsMargins(6, 4, 6, 0)

        # Per-form starting scales (requested defaults) with persisted overrides.
        start_defaults = {"1": 170, "2": 123, "2c": 191, "3": 88}
        start_default = int(start_defaults.get(form_key, 100))
        saved_use = self._settings.value(f"forms/{form_key}/use_starting_scale", True, type=bool)
        saved_scale = self._settings.value(f"forms/{form_key}/starting_scale_pct", start_default, type=int)

        use_start_cb = QCheckBox("Use starting scale")
        use_start_cb.setChecked(bool(saved_use))
        header_layout.addWidget(use_start_cb)

        header_layout.addWidget(QLabel("Starting scale:"))
        start_spin = QSpinBox()
        start_spin.setRange(50, 200)
        start_spin.setValue(max(50, min(200, int(saved_scale))))
        start_spin.setSuffix("%")
        start_spin.setMaximumWidth(90)
        header_layout.addWidget(start_spin)

        scale_label = QLabel("Scale: 100%")
        header_layout.addWidget(scale_label)

        include_thread_cb = None

        font_scale_spin = None

        wrap_cb = None

        # Fit rows (all forms): auto-fit row heights using wrapped text in columns A–T.
        try:
            fit_rows_btn = QPushButton("Fit Rows")
            fit_rows_btn.setToolTip("Auto-fit rows using wrapped text in columns A–T")
            fit_rows_btn.clicked.connect(lambda _=False, fk=str(form_key): self._on_fit_rows_requested(fk))
            header_layout.addWidget(fit_rows_btn)
        except Exception:
            pass

        # Form 3 specific controls.
        if form_key == "3":
            include_thread_cb = QCheckBox("Auto add Go/No Go + Minor Dia")
            include_thread_cb.setChecked(bool(getattr(self, "_form3_include_thread_extras", False)))
            header_layout.addWidget(include_thread_cb)

            # Fit Rows button is added for all forms above.

            try:
                renum_btn = QPushButton("Renumber Char/Bubble")
                renum_btn.setToolTip("Fill Char No. and Bubble No. for rows with Description/Note text")
                renum_btn.clicked.connect(self._on_form3_renumber_char_bubble_requested)
                header_layout.addWidget(renum_btn)
            except Exception:
                pass

            try:
                undo_btn = QPushButton("Undo Delete")
                undo_btn.setToolTip("Undo last Form 3 row delete")
                undo_btn.clicked.connect(lambda: self._on_form3_undo_requested())
                header_layout.addWidget(undo_btn)
            except Exception:
                pass

            # Form 3 GD&T controls removed per request:
            # - always use installed-font mode
            # - always use the "GDT" font family

            # Form 3: follow bubble selection into the Drawing Viewer.
            try:
                self._form3_follow_find_cb = QCheckBox("Click on Bubble to find on drawing")
                saved_follow = False
                try:
                    saved_follow = bool(self._settings.value("forms/3/follow_find", False, type=bool))
                except Exception:
                    saved_follow = False
                self._form3_follow_find_cb.setChecked(bool(saved_follow))
                try:
                    self._form3_follow_find_enabled = bool(saved_follow)
                except Exception:
                    self._form3_follow_find_enabled = bool(saved_follow)

                def _persist_follow(on: bool) -> None:
                    try:
                        self._form3_follow_find_enabled = bool(on)
                    except Exception:
                        pass
                    try:
                        self._settings.setValue("forms/3/follow_find", bool(on))
                    except Exception:
                        pass
                    # Keep pop-out checkbox (if present) in sync.
                    try:
                        cb2 = getattr(self, "_form3_follow_find_cb_pop", None)
                        if cb2 is not None and bool(cb2.isChecked()) != bool(on):
                            cb2.blockSignals(True)
                            cb2.setChecked(bool(on))
                            cb2.blockSignals(False)
                    except Exception:
                        pass

                self._form3_follow_find_cb.toggled.connect(lambda on: _persist_follow(bool(on)))
                header_layout.addWidget(self._form3_follow_find_cb)
            except Exception:
                pass

            # Form 3: font size control (replaces bubble backfill swatches in header).
            try:
                header_layout.addWidget(QLabel("Font size:"))
                saved_pt = self._settings.value("forms/3/font_point_size", None)
                if saved_pt is None:
                    # Back-compat with older % setting.
                    pct = self._settings.value("forms/3/font_scale_pct", 100, type=int)
                    try:
                        saved_pt = int(round(10.0 * (float(pct) / 100.0)))
                    except Exception:
                        saved_pt = 10
                font_scale_spin = QSpinBox()
                font_scale_spin.setRange(6, 24)
                font_scale_spin.setValue(max(6, min(24, int(saved_pt) if saved_pt is not None else 10)))
                font_scale_spin.setSuffix(" pt")
                font_scale_spin.setMaximumWidth(80)
                header_layout.addWidget(font_scale_spin)
                self._form3_font_scale_spin = font_scale_spin
            except Exception:
                font_scale_spin = None

        # Wrap Text toggle (requested) for all form tabs.
        if form_key in ("1", "2", "2c", "3"):
            wrap_cb = QCheckBox("Wrap Text")
            wrap_cb.setToolTip("Apply/clear Excel Wrap Text for the current selection")
            try:
                wrap_cb.setTristate(True)
            except Exception:
                pass
            header_layout.addWidget(wrap_cb)

        # Cell highlight color selectors for Forms 1/2/2c/3 (exclusive): color1, color2, clear/3rd.
        color_group = None
        red_btn = orange_btn = none_btn = None
        _clear_color_selection = None
        _persist_swatch_colors = None

        if form_key in ("1", "2", "2c", "3"):
            header_layout.addWidget(QLabel("Cell color:"))
            sw1 = str(self._settings.value("forms/3/paint_swatch1_rgb", "FFC7CE", type=str) or "").strip()
            sw2 = str(self._settings.value("forms/3/paint_swatch2_rgb", "FFEB9C", type=str) or "").strip()
            sw3 = str(self._settings.value("forms/3/paint_swatch3_rgb", "", type=str) or "").strip()
            red_btn = _ColorSwatchCheckBox("", sw1 or None, container)
            orange_btn = _ColorSwatchCheckBox("", sw2 or None, container)
            none_btn = _ColorSwatchCheckBox("", sw3 or None, container)

            red_btn.setToolTip("Color 1 (right-click to change)")
            orange_btn.setToolTip("Color 2 (right-click to change)")
            none_btn.setToolTip("Clear fill (right-click to change into a 3rd color)")

            color_group = QButtonGroup(container)
            color_group.setExclusive(True)
            color_group.addButton(red_btn, 1)
            color_group.addButton(orange_btn, 2)
            color_group.addButton(none_btn, 3)

            def _clear_color_selection() -> None:
                try:
                    color_group.setExclusive(False)
                    red_btn.setChecked(False)
                    orange_btn.setChecked(False)
                    none_btn.setChecked(False)
                finally:
                    color_group.setExclusive(True)

            # Default selection: none selected (paint mode off).
            _clear_color_selection()

            def _persist_swatch_colors() -> None:
                try:
                    self._settings.setValue("forms/3/paint_swatch1_rgb", red_btn.swatch_rgb() or "")
                    self._settings.setValue("forms/3/paint_swatch2_rgb", orange_btn.swatch_rgb() or "")
                    self._settings.setValue("forms/3/paint_swatch3_rgb", none_btn.swatch_rgb() or "")
                except Exception:
                    pass

            header_layout.addWidget(red_btn)
            header_layout.addWidget(orange_btn)
            header_layout.addWidget(none_btn)

            # Form 3 checkbox is independent, but it lives next to the paint controls.
            if include_thread_cb is not None:
                def _persist_and_refresh_form3() -> None:
                    try:
                        self._form3_include_thread_extras = include_thread_cb.isChecked()
                        self._settings.setValue(
                            "forms/3/include_thread_extras",
                            self._form3_include_thread_extras,
                        )
                    except Exception:
                        pass
                    self._refresh_form3_view()

                include_thread_cb.toggled.connect(lambda _checked: _persist_and_refresh_form3())


        header_layout.addStretch()

        layout.addWidget(header)

        viewer = ExcelSheetViewer()
        try:
            viewer.form_key = str(form_key)
        except Exception:
            pass

        # Visual polish: light-gray borders everywhere, but make Form 3 slightly thinner.
        try:
            if str(form_key) == "3" and hasattr(viewer, "set_border_width_scale"):
                viewer.set_border_width_scale(0.6)
        except Exception:
            pass
        viewer.set_persistence(self._settings, f"tables/forms/{form_key}")
        # Track modifications to set dirty flag
        try:
            viewer.modified.connect(self._set_wb_dirty)
            if form_key == "1":
                viewer.modified.connect(self._refresh_drawing_viewer_default_save_basename)
        except Exception:
            pass
            
        # Fit to width so the form uses the full window width.
        viewer.set_fit_mode("width")
        # Excel-like selection (outline only, no filled selection background).
        try:
            viewer.set_selection_outline_only(True)
        except Exception:
            pass
        if form_key == "3":
            viewer.set_hidden_columns([23, 24, 25])
            # Form 3: auto-fit row heights for wrapped Description/Note text (column G).
            viewer.set_auto_fit_row_height_columns([7])
            try:
                viewer.set_custom_undo_handler(self._on_form3_undo_requested)
            except Exception:
                pass
            try:
                viewer.rowDeleteRequested.disconnect()
            except Exception:
                pass
            try:
                viewer.rowDeleteManyRequested.disconnect()
            except Exception:
                pass
            try:
                viewer.rowDeleteRequested.connect(self._on_form3_row_delete_requested)
            except Exception:
                pass
            try:
                viewer.rowDeleteManyRequested.connect(self._on_form3_rows_delete_requested)
            except Exception:
                pass
            # Form 3: allow row insertion via right-click, but only below row 5.
            try:
                viewer.enable_row_insert_context_menu(True, min_row_1based=6)
                viewer.rowInsertRequested.connect(self._on_form3_row_insert_requested)
                try:
                    pass
                except Exception:
                    pass
            except Exception:
                pass

        if form_key == "2":
            # Form 2: auto-fit row heights for wrapped content in Column D.
            viewer.set_auto_fit_row_height_columns([4])

        if form_key == "3":
            # Form 3: click/scroll bubble follow (driven by Form 3 checkbox).
            try:
                def _follow_enabled() -> bool:
                    try:
                        return bool(getattr(self, "_form3_follow_find_enabled", False))
                    except Exception:
                        return False

                def _drawing_is_popped_out() -> bool:
                    try:
                        dv = getattr(self, "drawing_viewer_tab", None)
                        return bool(dv is not None and (dv.windowFlags() & Qt.Window))
                    except Exception:
                        return False

                def _select_bubble_on_drawing(n: int) -> None:
                    try:
                        nn = int(n)
                    except Exception:
                        return
                    if nn <= 0:
                        return

                    # If the Drawing Viewer is popped out, do NOT activate it;
                    # keep keyboard focus on Form 3 so Up/Down works.
                    try:
                        dv = getattr(self, "drawing_viewer_tab", None)
                        pv = getattr(dv, "_pdf_viewer", None) if dv is not None else None
                        popped_out = _drawing_is_popped_out()
                    except Exception:
                        pv = None
                        popped_out = False

                    if popped_out and pv is not None and hasattr(pv, "select_bubble_number"):
                        try:
                            pv.select_bubble_number(int(nn), center=True)
                        except Exception:
                            pass
                        return

                    try:
                        self._focus_drawing_and_select_bubble(int(nn))
                    except Exception:
                        pass

                def _bubble_number_at_row(ws, row_1based: int) -> int | None:
                    try:
                        v = ws.cell(row=int(row_1based), column=5).value
                        n = int(v)
                        return int(n) if n > 0 else None
                    except Exception:
                        return None

                def _select_row_bubble(rr_1based: int) -> int | None:
                    ws = getattr(viewer, "_ws", None)
                    tbl = getattr(viewer, "table", None)
                    if ws is None or tbl is None:
                        return None
                    n = _bubble_number_at_row(ws, int(rr_1based))
                    if n is None:
                        return None
                    r0 = int(rr_1based) - 1
                    c0 = 4
                    try:
                        tbl.clearSelection()
                    except Exception:
                        pass
                    try:
                        tbl.setCurrentCell(int(r0), int(c0))
                    except Exception:
                        pass
                    try:
                        tbl.setRangeSelected(QTableWidgetSelectionRange(int(r0), int(c0), int(r0), int(c0)), True)
                    except Exception:
                        pass
                    try:
                        it = tbl.item(int(r0), int(c0))
                        if it is not None:
                            off = max(0, int(r0) - 5)
                            it2 = tbl.item(int(off), int(c0))
                            if it2 is not None:
                                tbl.scrollToItem(it2, QAbstractItemView.ScrollHint.PositionAtTop)
                    except Exception:
                        pass
                    return int(n)

                def _on_form3_bubble_click(r0: int, c0: int) -> None:
                    if not _follow_enabled():
                        return
                    if int(c0) + 1 != 5:
                        return
                    ws = getattr(viewer, "_ws", None)
                    if ws is None:
                        return
                    n = _bubble_number_at_row(ws, int(r0) + 1)
                    if n is None:
                        return
                    _select_bubble_on_drawing(int(n))
                    if _drawing_is_popped_out():
                        try:
                            viewer.table.setFocus(Qt.FocusReason.OtherFocusReason)
                        except Exception:
                            pass

                def _wheel_nav(row_1based: int, col_1based: int, delta_y: int) -> bool:
                    if not _follow_enabled():
                        return False
                    if int(col_1based) != 5:
                        return False
                    if int(delta_y) == 0:
                        return False

                    ws = getattr(viewer, "_ws", None)
                    if ws is None:
                        return False

                    step = 1 if int(delta_y) < 0 else -1
                    rr = int(row_1based) + int(step)
                    if rr < 6:
                        rr = 6

                    try:
                        max_row = int(getattr(ws, "max_row", 0) or 0)
                    except Exception:
                        max_row = 0
                    if max_row <= 0:
                        max_row = int(rr) + 500

                    tries = 0
                    while 1 <= rr <= max_row and tries < 2000:
                        tries += 1
                        n = _select_row_bubble(int(rr))
                        if n is not None:
                            _select_bubble_on_drawing(int(n))
                            if _drawing_is_popped_out():
                                try:
                                    viewer.table.setFocus(Qt.FocusReason.OtherFocusReason)
                                except Exception:
                                    pass
                            return True
                        rr += int(step)
                    return False

                def _key_nav(row_1based: int, col_1based: int, direction: int) -> bool:
                    if not _follow_enabled():
                        return False
                    if int(col_1based) != 5:
                        return False
                    try:
                        direction = int(direction)
                    except Exception:
                        return False
                    if direction not in (-1, 1):
                        return False

                    ws = getattr(viewer, "_ws", None)
                    if ws is None:
                        return False

                    rr = int(row_1based) + int(direction)
                    if rr < 6:
                        rr = 6

                    try:
                        max_row = int(getattr(ws, "max_row", 0) or 0)
                    except Exception:
                        max_row = 0
                    if max_row <= 0:
                        max_row = int(rr) + 500

                    tries = 0
                    while 1 <= rr <= max_row and tries < 2000:
                        tries += 1
                        n = _select_row_bubble(int(rr))
                        if n is not None:
                            _select_bubble_on_drawing(int(n))
                            if _drawing_is_popped_out():
                                try:
                                    viewer.table.setFocus(Qt.FocusReason.OtherFocusReason)
                                except Exception:
                                    pass
                            return True
                        rr += int(direction)
                    return False

                try:
                    viewer.table.cellClicked.connect(_on_form3_bubble_click)
                except Exception:
                    pass
                try:
                    viewer.set_wheel_navigation_handler(_wheel_nav)
                except Exception:
                    pass
                try:
                    viewer.set_key_navigation_handler(_key_nav)
                except Exception:
                    pass
            except Exception:
                pass

        if wrap_cb is not None:
            _wrap_updating = {"busy": False}

            def _compute_wrap_state_for_selection() -> Qt.CheckState:
                try:
                    ws = getattr(viewer, "_ws", None)
                    tbl = getattr(viewer, "table", None)
                    if ws is None or tbl is None:
                        return Qt.CheckState.Unchecked
                    ranges = list(tbl.selectedRanges() or [])
                    if not ranges:
                        return Qt.CheckState.Unchecked
                except Exception:
                    return Qt.CheckState.Unchecked

                any_true = False
                any_false = False

                for rng in ranges:
                    r0 = int(rng.topRow()) + 1
                    r1 = int(rng.bottomRow()) + 1
                    c0 = int(rng.leftColumn()) + 1
                    c1 = int(rng.rightColumn()) + 1
                    for rr in range(r0, r1 + 1):
                        for cc in range(c0, c1 + 1):
                            try:
                                cell = ws.cell(row=int(rr), column=int(cc))
                                align = getattr(cell, "alignment", None)
                                w = bool(getattr(align, "wrapText", False)) if align is not None else False
                            except Exception:
                                w = False
                            if w:
                                any_true = True
                            else:
                                any_false = True
                            if any_true and any_false:
                                return Qt.CheckState.PartiallyChecked

                if any_true and not any_false:
                    return Qt.CheckState.Checked
                if any_false and not any_true:
                    return Qt.CheckState.Unchecked
                return Qt.CheckState.Unchecked

            def _update_wrap_checkbox_from_selection() -> None:
                if _wrap_updating["busy"]:
                    return
                _wrap_updating["busy"] = True
                try:
                    st = _compute_wrap_state_for_selection()
                    try:
                        wrap_cb.setCheckState(st)
                    except Exception:
                        wrap_cb.setChecked(st == Qt.CheckState.Checked)
                finally:
                    _wrap_updating["busy"] = False

            def _on_wrap_state_changed(state: int) -> None:
                if _wrap_updating["busy"]:
                    return
                # User action: apply wrap for selection.
                try:
                    st = Qt.CheckState(state)
                except Exception:
                    st = state
                desired = (st == Qt.CheckState.Checked)
                applied = False
                try:
                    applied = bool(viewer.set_wrap_text_for_selection(bool(desired)))
                except Exception:
                    applied = False

                if not applied:
                    try:
                        QMessageBox.information(self, "Wrap Text", "Please select a cell, row, or column first.")
                    except Exception:
                        pass
                # Always refresh checkbox state from the actual selection after attempting.
                QTimer.singleShot(0, _update_wrap_checkbox_from_selection)

            try:
                viewer.table.itemSelectionChanged.connect(_update_wrap_checkbox_from_selection)
            except Exception:
                pass

            try:
                wrap_cb.stateChanged.connect(_on_wrap_state_changed)
            except Exception:
                pass

            # Initial sync.
            QTimer.singleShot(0, _update_wrap_checkbox_from_selection)

        # Wire up the color selector to apply-to-selection behavior (Excel-like).
        if color_group is not None and red_btn is not None and orange_btn is not None and none_btn is not None:
            try:
                # Ensure no persistent paint mode is left enabled.
                try:
                    viewer.set_click_paint_fill_rgb(None)
                except Exception:
                    pass

                def _apply_fill_from_swatch() -> None:
                    checked_id = -1
                    try:
                        checked_id = int(color_group.checkedId())
                    except Exception:
                        checked_id = -1

                    if checked_id == 1:
                        rgb = red_btn.swatch_rgb() or ""
                    elif checked_id == 2:
                        rgb = orange_btn.swatch_rgb() or ""
                    elif checked_id == 3:
                        rgb = none_btn.swatch_rgb() or ""
                    else:
                        return

                    # Apply immediately to selection (or current cell).
                    applied = False
                    try:
                        applied = bool(viewer.apply_fill_to_selection(rgb if rgb else None))
                    except Exception:
                        applied = False

                    if not applied:
                        try:
                            QMessageBox.information(self, "Cell color", "Please select a cell (or range) first.")
                        except Exception:
                            pass

                    # Always clear the swatch selection so no mode stays active.
                    try:
                        if _clear_color_selection is not None:
                            QTimer.singleShot(0, _clear_color_selection)
                    except Exception:
                        pass
                    try:
                        viewer.set_click_paint_fill_rgb(None)
                    except Exception:
                        pass

                # Clicking a swatch applies the color and immediately clears the selection.
                color_group.buttonClicked.connect(lambda _b=None: _apply_fill_from_swatch())

                def _refresh_after_color_pick() -> None:
                    try:
                        if _persist_swatch_colors is not None:
                            _persist_swatch_colors()
                    except Exception:
                        pass

                red_btn.colorChanged.connect(_refresh_after_color_pick)
                orange_btn.colorChanged.connect(_refresh_after_color_pick)
                none_btn.colorChanged.connect(_refresh_after_color_pick)
            except Exception:
                pass

        def _apply_starting_scale() -> None:
            if use_start_cb.isChecked():
                viewer.set_effective_scale(float(start_spin.value()) / 100.0)
            else:
                viewer.reset_auto_scale()

        def _update_scale_label(scale: float) -> None:
            try:
                pct = int(round(float(scale) * 100.0))
            except Exception:
                pct = 100
            scale_label.setText(f"Scale: {pct}%")

        def _persist_settings() -> None:
            self._settings.setValue(f"forms/{form_key}/use_starting_scale", use_start_cb.isChecked())
            self._settings.setValue(f"forms/{form_key}/starting_scale_pct", int(start_spin.value()))

        viewer.scaleChanged.connect(_update_scale_label)
        _update_scale_label(viewer.effective_scale())
        use_start_cb.toggled.connect(lambda _checked: (_persist_settings(), _apply_starting_scale()))
        start_spin.valueChanged.connect(
            lambda _v: (_persist_settings(), _apply_starting_scale()) if use_start_cb.isChecked() else _persist_settings()
        )

        # Form 3 font scaling (fonts only; does not change row/column sizes).
        if str(form_key) == "3" and font_scale_spin is not None and hasattr(viewer, "set_font_scale_multiplier"):

            def _apply_form3_font_scale() -> None:
                try:
                    pt = int(font_scale_spin.value())
                except Exception:
                    pt = 10
                pt = max(6, min(24, int(pt)))
                try:
                    self._settings.setValue("forms/3/font_point_size", int(pt))
                except Exception:
                    pass

                # Map point-size request to a multiplier against the viewer's base font.
                try:
                    base_pt = float(viewer.table.font().pointSizeF())
                    if base_pt <= 0:
                        base_pt = 10.0
                except Exception:
                    base_pt = 10.0
                try:
                    viewer.set_font_scale_multiplier(float(pt) / float(base_pt))
                except Exception:
                    pass

                # Keep pop-out control (if present) in sync.
                try:
                    spin2 = getattr(self, "_form3_font_scale_spin_pop", None)
                    if spin2 is not None and int(spin2.value()) != int(pt):
                        spin2.blockSignals(True)
                        spin2.setValue(int(pt))
                        spin2.blockSignals(False)
                except Exception:
                    pass

            font_scale_spin.valueChanged.connect(lambda _v: _apply_form3_font_scale())
            _apply_form3_font_scale()

        # Apply initial per-form starting scale (applies after first render).
        _apply_starting_scale()
        layout.addWidget(viewer)

        return container, viewer

    def pop_out_form(self, form_key: str):
        viewer = self._form_viewers.get(form_key)
        if viewer is None:
            return

        win = QMainWindow(self)
        title_map = {"1": "Form 1", "2": "Form 2", "2c": "Form 2 Cont.", "3": "Form 3"}
        win.setWindowTitle(title_map.get(form_key, "Form"))

        container = QWidget()
        layout = QVBoxLayout(container)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(4)

        header = QWidget()
        header_layout = QHBoxLayout(header)
        header_layout.setContentsMargins(6, 4, 6, 0)

        start_defaults = {"1": 170, "2": 123, "2c": 191, "3": 88}
        start_default = int(start_defaults.get(form_key, 100))
        saved_use = self._settings.value(f"forms/{form_key}/use_starting_scale", True, type=bool)
        saved_scale = self._settings.value(f"forms/{form_key}/starting_scale_pct", start_default, type=int)

        use_start_cb = QCheckBox("Use starting scale")
        use_start_cb.setChecked(bool(saved_use))
        header_layout.addWidget(use_start_cb)

        header_layout.addWidget(QLabel("Starting scale:"))
        start_spin = QSpinBox()
        start_spin.setRange(50, 200)
        start_spin.setValue(max(50, min(200, int(saved_scale))))
        start_spin.setSuffix("%")
        start_spin.setMaximumWidth(90)
        header_layout.addWidget(start_spin)

        scale_label = QLabel("Scale: 100%")
        header_layout.addWidget(scale_label)

        include_thread_cb = None
        if form_key == "3":
            include_thread_cb = QCheckBox("Auto add Go/No Go + Minor Dia")
            include_thread_cb.setChecked(bool(getattr(self, "_form3_include_thread_extras", False)))
            header_layout.addWidget(include_thread_cb)

            # Form 3 GD&T controls removed per request (always installed-font mode, font family "GDT").

            # Pop-out Form 3: font size control (fonts only).
            try:
                header_layout.addWidget(QLabel("Font size:"))
                saved_pt = self._settings.value("forms/3/font_point_size", None)
                if saved_pt is None:
                    # Back-compat with older % setting.
                    pct = self._settings.value("forms/3/font_scale_pct", 100, type=int)
                    try:
                        saved_pt = int(round(10.0 * (float(pct) / 100.0)))
                    except Exception:
                        saved_pt = 10
                self._form3_font_scale_spin_pop = QSpinBox()
                self._form3_font_scale_spin_pop.setRange(6, 24)
                self._form3_font_scale_spin_pop.setValue(max(6, min(24, int(saved_pt) if saved_pt is not None else 10)))
                self._form3_font_scale_spin_pop.setSuffix(" pt")
                self._form3_font_scale_spin_pop.setMaximumWidth(80)
                header_layout.addWidget(self._form3_font_scale_spin_pop)
            except Exception:
                self._form3_font_scale_spin_pop = None

        # Cell highlight color selectors for Forms 1/2/2c/3 (pop-out).
        color_group = None
        red_btn = orange_btn = none_btn = None
        _clear_color_selection = None
        _persist_swatch_colors = None

        if form_key in ("1", "2", "2c", "3"):
            header_layout.addWidget(QLabel("Cell color:"))
            sw1 = str(self._settings.value("forms/3/paint_swatch1_rgb", "FFC7CE", type=str) or "").strip()
            sw2 = str(self._settings.value("forms/3/paint_swatch2_rgb", "FFEB9C", type=str) or "").strip()
            sw3 = str(self._settings.value("forms/3/paint_swatch3_rgb", "", type=str) or "").strip()
            red_btn = _ColorSwatchCheckBox("", sw1 or None, win)
            orange_btn = _ColorSwatchCheckBox("", sw2 or None, win)
            none_btn = _ColorSwatchCheckBox("", sw3 or None, win)

            red_btn.setToolTip("Color 1 (right-click to change)")
            orange_btn.setToolTip("Color 2 (right-click to change)")
            none_btn.setToolTip("Clear fill (right-click to change into a 3rd color)")

            color_group = QButtonGroup(win)
            color_group.setExclusive(True)
            color_group.addButton(red_btn, 1)
            color_group.addButton(orange_btn, 2)
            color_group.addButton(none_btn, 3)

            def _clear_color_selection() -> None:
                try:
                    color_group.setExclusive(False)
                    red_btn.setChecked(False)
                    orange_btn.setChecked(False)
                    none_btn.setChecked(False)
                finally:
                    color_group.setExclusive(True)

            _clear_color_selection()

            def _persist_swatch_colors() -> None:
                try:
                    self._settings.setValue("forms/3/paint_swatch1_rgb", red_btn.swatch_rgb() or "")
                    self._settings.setValue("forms/3/paint_swatch2_rgb", orange_btn.swatch_rgb() or "")
                    self._settings.setValue("forms/3/paint_swatch3_rgb", none_btn.swatch_rgb() or "")
                except Exception:
                    pass

            header_layout.addWidget(red_btn)
            header_layout.addWidget(orange_btn)
            header_layout.addWidget(none_btn)

            if include_thread_cb is not None:
                def _persist_and_refresh_form3() -> None:
                    try:
                        self._form3_include_thread_extras = include_thread_cb.isChecked()
                        self._settings.setValue(
                            "forms/3/include_thread_extras",
                            self._form3_include_thread_extras,
                        )
                    except Exception:
                        pass
                    # Re-render both the embedded and pop-out views.
                    self._refresh_form3_view()
                    try:
                        pop_viewer.set_overrides({})
                        pop_viewer.render()
                    except Exception:
                        pass

                include_thread_cb.toggled.connect(lambda _checked: _persist_and_refresh_form3())

            # Mirror the embedded Form 3 bubble-follow checkbox in the pop-out header.
            try:
                self._form3_follow_find_cb_pop = QCheckBox("Click on Bubble to find on drawing")
                saved_follow = bool(getattr(self, "_form3_follow_find_enabled", False))
                self._form3_follow_find_cb_pop.setChecked(bool(saved_follow))

                def _persist_follow_pop(on: bool) -> None:
                    try:
                        self._form3_follow_find_enabled = bool(on)
                    except Exception:
                        pass
                    try:
                        self._settings.setValue("forms/3/follow_find", bool(on))
                    except Exception:
                        pass
                    # Keep embedded checkbox (if present) in sync.
                    try:
                        cb1 = getattr(self, "_form3_follow_find_cb", None)
                        if cb1 is not None and bool(cb1.isChecked()) != bool(on):
                            cb1.blockSignals(True)
                            cb1.setChecked(bool(on))
                            cb1.blockSignals(False)
                    except Exception:
                        pass

                self._form3_follow_find_cb_pop.toggled.connect(lambda on: _persist_follow_pop(bool(on)))
                header_layout.addWidget(self._form3_follow_find_cb_pop)
            except Exception:
                pass

        header_layout.addStretch()
        layout.addWidget(header)

        pop_viewer = ExcelSheetViewer()
        try:
            pop_viewer.form_key = str(form_key)
        except Exception:
            pass
        pop_viewer.set_persistence(self._settings, f"tables/forms/{form_key}")
        pop_viewer.set_fit_mode("width")
        # Excel-like selection (outline only, no filled selection background).
        try:
            pop_viewer.set_selection_outline_only(True)
        except Exception:
            pass
        if form_key == "3":
            pop_viewer.set_hidden_columns([23, 24, 25])
            pop_viewer.set_auto_fit_row_height_columns([7])
            try:
                pop_viewer.set_custom_undo_handler(self._on_form3_undo_requested)
            except Exception:
                pass

            # Form 3 pop-out: apply font-only scale and keep it in sync.
            try:
                spin_pop = getattr(self, "_form3_font_scale_spin_pop", None)
                if spin_pop is not None and hasattr(pop_viewer, "set_font_scale_multiplier"):

                    def _apply_form3_font_scale_pop() -> None:
                        try:
                            pt = int(spin_pop.value())
                        except Exception:
                            pt = 10
                        pt = max(6, min(24, int(pt)))
                        try:
                            self._settings.setValue("forms/3/font_point_size", int(pt))
                        except Exception:
                            pass
                        try:
                            base_pt = float(pop_viewer.table.font().pointSizeF())
                            if base_pt <= 0:
                                base_pt = 10.0
                        except Exception:
                            base_pt = 10.0
                        try:
                            pop_viewer.set_font_scale_multiplier(float(pt) / float(base_pt))
                        except Exception:
                            pass

                        # Keep embedded control (if present) in sync.
                        try:
                            spin_emb = getattr(self, "_form3_font_scale_spin", None)
                            if spin_emb is not None and int(spin_emb.value()) != int(pt):
                                spin_emb.blockSignals(True)
                                spin_emb.setValue(int(pt))
                                spin_emb.blockSignals(False)
                        except Exception:
                            pass

                    spin_pop.valueChanged.connect(lambda _v: _apply_form3_font_scale_pop())
                    _apply_form3_font_scale_pop()
            except Exception:
                pass

            # Form 3 pop-out: click/scroll bubble follow (driven by Form 3 checkbox).
            try:
                def _follow_enabled() -> bool:
                    try:
                        return bool(getattr(self, "_form3_follow_find_enabled", False))
                    except Exception:
                        return False

                def _select_bubble_on_drawing(n: int) -> None:
                    try:
                        nn = int(n)
                    except Exception:
                        return
                    if nn <= 0:
                        return

                    # If the Drawing Viewer is popped out, do NOT activate it;
                    # keep keyboard focus on Form 3 so Up/Down works.
                    try:
                        dv = getattr(self, "drawing_viewer_tab", None)
                        pv = getattr(dv, "_pdf_viewer", None) if dv is not None else None
                        popped_out = bool(dv is not None and (dv.windowFlags() & Qt.Window))
                    except Exception:
                        dv = None
                        pv = None
                        popped_out = False

                    if popped_out and pv is not None and hasattr(pv, "select_bubble_number"):
                        try:
                            pv.select_bubble_number(int(nn), center=True)
                        except Exception:
                            pass
                        return

                    try:
                        self._focus_drawing_and_select_bubble(int(nn))
                    except Exception:
                        pass

                def _bubble_number_at_row(ws, row_1based: int) -> int | None:
                    try:
                        v = ws.cell(row=int(row_1based), column=5).value
                        n = int(v)
                        return int(n) if n > 0 else None
                    except Exception:
                        return None

                def _on_form3_bubble_click(r0: int, c0: int) -> None:
                    if not _follow_enabled():
                        return
                    if int(c0) + 1 != 5:
                        return
                    ws = getattr(pop_viewer, "_ws", None)
                    if ws is None:
                        return
                    n = _bubble_number_at_row(ws, int(r0) + 1)
                    if n is None:
                        return
                    _select_bubble_on_drawing(int(n))

                    # If drawing viewer is popped out, keep focus here.
                    try:
                        dv = getattr(self, "drawing_viewer_tab", None)
                        popped_out = bool(dv is not None and (dv.windowFlags() & Qt.Window))
                    except Exception:
                        popped_out = False
                    if popped_out:
                        try:
                            tbl = pop_viewer.table
                            try:
                                tbl.clearSelection()
                            except Exception:
                                pass
                            tbl.setCurrentCell(int(r0), 4)
                            try:
                                tbl.setRangeSelected(QTableWidgetSelectionRange(int(r0), 4, int(r0), 4), True)
                            except Exception:
                                pass
                            try:
                                it = tbl.item(int(r0), 4)
                                if it is not None:
                                    off = max(0, int(r0) - 5)
                                    it2 = tbl.item(int(off), 4)
                                    if it2 is not None:
                                        tbl.scrollToItem(it2, QAbstractItemView.ScrollHint.PositionAtTop)
                            except Exception:
                                pass
                            pop_viewer.table.setFocus(Qt.FocusReason.OtherFocusReason)
                        except Exception:
                            pass

                def _wheel_nav(row_1based: int, col_1based: int, delta_y: int) -> bool:
                    if not _follow_enabled():
                        return False
                    if int(col_1based) != 5:
                        return False
                    if int(delta_y) == 0:
                        return False

                    ws = getattr(pop_viewer, "_ws", None)
                    tbl = getattr(pop_viewer, "table", None)
                    if ws is None or tbl is None:
                        return False

                    step = 1 if int(delta_y) < 0 else -1
                    rr = int(row_1based) + int(step)
                    if rr < 6:
                        rr = 6

                    max_row = 0
                    try:
                        max_row = int(getattr(ws, "max_row", 0) or 0)
                    except Exception:
                        max_row = 0
                    if max_row <= 0:
                        max_row = int(rr) + 500

                    tries = 0
                    while 1 <= rr <= max_row and tries < 2000:
                        tries += 1
                        n = _bubble_number_at_row(ws, rr)
                        if n is not None:
                            try:
                                r0 = int(rr) - 1
                                c0 = 4
                                try:
                                    tbl.clearSelection()
                                except Exception:
                                    pass
                                tbl.setCurrentCell(int(r0), int(c0))
                                try:
                                    tbl.setRangeSelected(QTableWidgetSelectionRange(int(r0), int(c0), int(r0), int(c0)), True)
                                except Exception:
                                    pass
                                try:
                                    it = tbl.item(int(r0), int(c0))
                                    if it is not None:
                                        off = max(0, int(r0) - 5)
                                        it2 = tbl.item(int(off), int(c0))
                                        if it2 is not None:
                                            tbl.scrollToItem(it2, QAbstractItemView.ScrollHint.PositionAtTop)
                                except Exception:
                                    pass
                            except Exception:
                                pass
                            _select_bubble_on_drawing(int(n))

                            # If drawing viewer is popped out, keep focus here.
                            try:
                                dv = getattr(self, "drawing_viewer_tab", None)
                                popped_out = bool(dv is not None and (dv.windowFlags() & Qt.Window))
                            except Exception:
                                popped_out = False
                            if popped_out:
                                try:
                                    tbl.setFocus(Qt.FocusReason.OtherFocusReason)
                                except Exception:
                                    pass
                            return True
                        rr += int(step)

                    return False

                def _key_nav(row_1based: int, col_1based: int, direction: int) -> bool:
                    if not _follow_enabled():
                        return False
                    if int(col_1based) != 5:
                        return False
                    try:
                        direction = int(direction)
                    except Exception:
                        direction = 0
                    if direction not in (-1, 1):
                        return False

                    ws = getattr(pop_viewer, "_ws", None)
                    tbl = getattr(pop_viewer, "table", None)
                    if ws is None or tbl is None:
                        return False

                    rr = int(row_1based) + int(direction)
                    if rr < 6:
                        rr = 6

                    max_row = 0
                    try:
                        max_row = int(getattr(ws, "max_row", 0) or 0)
                    except Exception:
                        max_row = 0
                    if max_row <= 0:
                        max_row = int(rr) + 500

                    tries = 0
                    while 1 <= rr <= max_row and tries < 2000:
                        tries += 1
                        n = _bubble_number_at_row(ws, rr)
                        if n is not None:
                            try:
                                r0 = int(rr) - 1
                                c0 = 4
                                try:
                                    tbl.clearSelection()
                                except Exception:
                                    pass
                                tbl.setCurrentCell(int(r0), int(c0))
                                try:
                                    tbl.setRangeSelected(QTableWidgetSelectionRange(int(r0), int(c0), int(r0), int(c0)), True)
                                except Exception:
                                    pass
                                try:
                                    it = tbl.item(int(r0), int(c0))
                                    if it is not None:
                                        off = max(0, int(r0) - 5)
                                        it2 = tbl.item(int(off), int(c0))
                                        if it2 is not None:
                                            tbl.scrollToItem(it2, QAbstractItemView.ScrollHint.PositionAtTop)
                                except Exception:
                                    pass
                            except Exception:
                                pass
                            _select_bubble_on_drawing(int(n))

                            # If drawing viewer is popped out, keep focus here.
                            try:
                                dv = getattr(self, "drawing_viewer_tab", None)
                                popped_out = bool(dv is not None and (dv.windowFlags() & Qt.Window))
                            except Exception:
                                popped_out = False
                            if popped_out:
                                try:
                                    tbl.setFocus(Qt.FocusReason.OtherFocusReason)
                                except Exception:
                                    pass
                            return True
                        rr += int(direction)

                    return False

                try:
                    pop_viewer.table.cellClicked.connect(_on_form3_bubble_click)
                except Exception:
                    pass
                try:
                    pop_viewer.set_wheel_navigation_handler(_wheel_nav)
                except Exception:
                    pass
                try:
                    pop_viewer.set_key_navigation_handler(_key_nav)
                except Exception:
                    pass
            except Exception:
                pass

        # Wire up the color selector to apply-to-selection behavior (Excel-like).
        if color_group is not None and red_btn is not None and orange_btn is not None and none_btn is not None:
            try:
                try:
                    pop_viewer.set_click_paint_fill_rgb(None)
                except Exception:
                    pass

                def _apply_fill_from_swatch() -> None:
                    checked_id = -1
                    try:
                        checked_id = int(color_group.checkedId())
                    except Exception:
                        checked_id = -1

                    if checked_id == 1:
                        rgb = red_btn.swatch_rgb() or ""
                    elif checked_id == 2:
                        rgb = orange_btn.swatch_rgb() or ""
                    elif checked_id == 3:
                        rgb = none_btn.swatch_rgb() or ""
                    else:
                        return

                    applied = False
                    try:
                        applied = bool(pop_viewer.apply_fill_to_selection(rgb if rgb else None))
                    except Exception:
                        applied = False

                    if not applied:
                        try:
                            QMessageBox.information(self, "Cell color", "Please select a cell (or range) first.")
                        except Exception:
                            pass

                    try:
                        if _clear_color_selection is not None:
                            QTimer.singleShot(0, _clear_color_selection)
                    except Exception:
                        pass
                    try:
                        pop_viewer.set_click_paint_fill_rgb(None)
                    except Exception:
                        pass

                color_group.buttonClicked.connect(lambda _b=None: _apply_fill_from_swatch())

                def _refresh_after_color_pick() -> None:
                    try:
                        if _persist_swatch_colors is not None:
                            _persist_swatch_colors()
                    except Exception:
                        pass

                red_btn.colorChanged.connect(_refresh_after_color_pick)
                orange_btn.colorChanged.connect(_refresh_after_color_pick)
                none_btn.colorChanged.connect(_refresh_after_color_pick)
            except Exception:
                pass

        def _update_scale_label(scale: float) -> None:
            try:
                pct = int(round(float(scale) * 100.0))
            except Exception:
                pct = 100
            scale_label.setText(f"Scale: {pct}%")

        def _apply_starting_scale() -> None:
            if use_start_cb.isChecked():
                pop_viewer.set_effective_scale(float(start_spin.value()) / 100.0)
            else:
                pop_viewer.reset_auto_scale()

        def _persist_settings() -> None:
            self._settings.setValue(f"forms/{form_key}/use_starting_scale", use_start_cb.isChecked())
            self._settings.setValue(f"forms/{form_key}/starting_scale_pct", int(start_spin.value()))

        pop_viewer.scaleChanged.connect(_update_scale_label)
        _update_scale_label(pop_viewer.effective_scale())
        use_start_cb.toggled.connect(lambda _checked: (_persist_settings(), _apply_starting_scale()))
        start_spin.valueChanged.connect(
            lambda _v: (_persist_settings(), _apply_starting_scale()) if use_start_cb.isChecked() else _persist_settings()
        )
        _apply_starting_scale()

        if self._template_wb is not None and self._form_sheet_names.get(form_key):
            ws = self._template_wb[self._form_sheet_names[form_key]]
            pop_viewer.set_worksheet(ws)
            pop_viewer.set_overrides({})
            pop_viewer.render()

        layout.addWidget(pop_viewer)
        win.setCentralWidget(container)
        win.resize(1200, 900)
        win.show()

        # Prevent garbage collection
        self._popout_windows.append(win)

    def _normalize_sheet_name(self, s: str) -> str:
        return re.sub(r"\s+", " ", str(s or "").strip().lower())

    def _detect_form_sheets(self, wb) -> None:
        self._form_sheet_names = {"1": None, "2": None, "2c": None, "3": None}
        for name in wb.sheetnames:
            n = self._normalize_sheet_name(name)
            if "form 1" in n or n.endswith("form1") or "form1" in n:
                self._form_sheet_names["1"] = name
            # Must detect "Form 2 Cont" before "Form 2".
            # Templates often use punctuation like "Form 2 - Cont.".
            elif ("form 2" in n and "cont" in n) or "form2cont" in n or n.endswith("form2cont"):
                self._form_sheet_names["2c"] = name
            elif "form 2" in n or n.endswith("form2") or "form2" in n:
                self._form_sheet_names["2"] = name
            elif "form 3" in n or n.endswith("form3") or "form3" in n:
                self._form_sheet_names["3"] = name

        # Fallback: assign remaining keys in sheet order
        remaining_keys = [k for k in ("1", "2", "2c", "3") if not self._form_sheet_names.get(k)]
        used = {v for v in self._form_sheet_names.values() if v}
        for name in wb.sheetnames:
            if not remaining_keys:
                break
            if name in used:
                continue
            self._form_sheet_names[remaining_keys.pop(0)] = name

    def load_template(self):
        if not self.template_path or not os.path.exists(self.template_path):
            self._template_wb = None
            for v in self._form_viewers.values():
                if v is not None:
                    v.set_worksheet(None)
                    v.set_overrides({})
                    v.render()
            return

        try:
            self._template_wb = openpyxl.load_workbook(self.template_path)
        except Exception as e:
            QMessageBox.warning(self, "Template Error", f"Failed to load template:\n{e}")
            self._template_wb = None
            return

        try:
            self._form3_undo_stack = []
        except Exception:
            pass

        self._settings.setValue("paths/template", self.template_path)

        self._detect_form_sheets(self._template_wb)

        # Customer list (Customer -> Supplier Code)
        # This also merges any legacy hidden sheets on first creation.
        self._supplier_master_sheet(self._template_wb)
        self._apply_supplier_master_validations()
        # Supplier directory list (Company -> Address)
        self._ensure_supplier_directory_sheet(self._template_wb)

        # Calibrated equipment list
        self._calibrated_equipment_sheet(self._template_wb)

        # Apply persisted lists (so adds/deletes carry across runs without Save XLSX).
        self._apply_persistent_lists_to_workbook()

        # Load tabs from workbook after persistence merge.
        self._load_suppliers_tab_from_workbook()
        self._load_supplier_directory_tab_from_workbook()
        self._load_calibrated_equipment_tab_from_workbook()

        for form_key in ("1", "2", "2c", "3"):
            viewer = self._form_viewers.get(form_key)
            sheet_name = self._form_sheet_names.get(form_key)
            if viewer is None or not sheet_name:
                continue
            ws = self._template_wb[sheet_name]
            if form_key == "1":
                self._ensure_form1_reason_dropdown(ws)
                self._ensure_supplier_directory_dropdown(ws, cell_range="E15:E500")

                def _set_default_if_blank(cell_addr: str, default_value: str) -> None:
                    try:
                        c = ws[cell_addr]
                        v = c.value
                        if v is None:
                            c.value = default_value
                            return
                        if isinstance(v, str) and not v.strip():
                            c.value = default_value
                            return
                    except Exception:
                        pass

                # Requested Form 1 placeholders.
                _set_default_if_blank("B9", "Job #")
                _set_default_if_blank("D7", "A/NA/Stamp #")
                _set_default_if_blank("E9", "XXXXXXXXXX/xxxxx")
                _set_default_if_blank("D9", "10033672")
            elif form_key == "2":
                self._ensure_supplier_directory_dropdown(ws, cell_range="F5:F500")

                # Requested default: Wrap Text for Column D on Form 2.
                try:
                    from openpyxl.styles import Alignment
                except Exception:
                    Alignment = None

                if Alignment is not None:
                    # Limit rows for performance; templates can have a lot of formatting.
                    try:
                        max_r = int(getattr(ws, "max_row", 0) or 0)
                    except Exception:
                        max_r = 0
                    max_r = max(max_r, 200)
                    max_r = min(max_r, 1000)
                    col = 4  # D
                    for rr in range(1, int(max_r) + 1):
                        try:
                            cell = ws.cell(row=int(rr), column=int(col))
                            cur = getattr(cell, "alignment", None)
                            if cur is not None:
                                try:
                                    cell.alignment = cur.copy(wrapText=True)
                                except Exception:
                                    cell.alignment = Alignment(wrapText=True)
                            else:
                                cell.alignment = Alignment(wrapText=True)
                        except Exception:
                            continue
            viewer.set_worksheet(ws)
            # Persist any existing Calypso data into Form 3 worksheet.
            if form_key == "3" and self.characteristics:
                self._write_form3_to_worksheet(ws)
            viewer.set_overrides({})
            viewer.render()

        try:
            self._refresh_drawing_viewer_default_save_basename()
        except Exception:
            pass

        # If the drawing bubbles were loaded before (or during) template load,
        # the initial bubbles_changed may have fired while Form 3 was not ready.
        self._sync_bubbles_to_form3(set(getattr(self, "_last_bubbled_numbers", set()) or set()))

    def _ensure_supplier_directory_sheet(self, wb):
        """Ensure a hidden supplier directory sheet exists.

        Sheet layout:
          A: Company
                    B: Address 1
                    C: Address 2
                    D: City
                    E: State
                    F: Zip Code
                    G: Full Address (computed; used for dropdown stored value)
                    H: Marker
        """

        sheet_name = "__as9102_supplier_directory"
        created = False
        if sheet_name in wb.sheetnames:
            ls = wb[sheet_name]
        else:
            ls = wb.create_sheet(sheet_name)
            created = True
        try:
            ls.sheet_state = "hidden"
        except Exception:
            pass

        def _write_structured_rows(rows: list[tuple[str, str, str, str, str, str]]) -> None:
            clear_to = max(len(rows) + 20, 50)
            for rr in range(1, clear_to + 1):
                for cc in range(1, 9):
                    ls.cell(row=rr, column=cc).value = None
            for i, (company, addr1, addr2, city, state, zipc) in enumerate(rows, start=1):
                addr1 = _clean_company_prefix(company, addr1)
                state = str(state or "").strip().upper()
                ls.cell(row=i, column=1).value = company
                ls.cell(row=i, column=2).value = addr1
                ls.cell(row=i, column=3).value = addr2
                ls.cell(row=i, column=4).value = city
                ls.cell(row=i, column=5).value = state
                ls.cell(row=i, column=6).value = zipc
                ls.cell(row=i, column=7).value = _build_full_address_with_company(company, addr1, addr2, city, state, zipc)

        def _migrate_legacy_ab_to_structured() -> None:
            # If we only have A/B populated, split B into columns and compute H.
            max_scan = min(max(getattr(ls, "max_row", 0) or 0, 50), 5000)
            legacy_rows: list[tuple[str, str, str, str, str, str]] = []
            for rr in range(1, max_scan + 1):
                comp = ls.cell(row=rr, column=1).value
                addr = ls.cell(row=rr, column=2).value
                comp_s = str(comp).strip() if comp is not None else ""
                addr_s = str(addr).strip() if addr is not None else ""
                if not (comp_s or addr_s):
                    continue
                addr_s = _clean_company_prefix(comp_s, addr_s)
                a1, a2, a3, ct, st, z = _split_address_lines(addr_s)
                if a3:
                    a2 = f"{a2}, {a3}" if a2 else a3
                legacy_rows.append((comp_s, a1, a2, ct, st, z))

            if not legacy_rows:
                return
            _write_structured_rows(legacy_rows)

        # Seed only if empty/new. Try to find an existing Company/Address table in the workbook.
        a1 = ls.cell(row=1, column=1).value
        b1 = ls.cell(row=1, column=2).value
        is_empty = (a1 is None or str(a1).strip() == "") and (b1 is None or str(b1).strip() == "")
        if created or is_empty:
            rows: list[tuple[str, str, str, str, str, str]] = []

            def _norm(s: str) -> str:
                return re.sub(r"\s+", " ", str(s or "").strip().lower())

            # Look for adjacent headers: Company | Address
            for nm in wb.sheetnames:
                if nm.startswith("__as9102_"):
                    continue
                try:
                    ws_src = wb[nm]
                except Exception:
                    continue

                found = None
                for r in range(1, min(getattr(ws_src, "max_row", 0) or 0, 300) + 1):
                    for c in range(1, min(getattr(ws_src, "max_column", 0) or 0, 50) + 1):
                        v = ws_src.cell(row=r, column=c).value
                        v2 = ws_src.cell(row=r, column=c + 1).value
                        if _norm(v) == "company" and _norm(v2) == "address":
                            found = (r, c)
                            break
                    if found:
                        break

                if not found:
                    continue

                hr, hc = found
                for rr in range(hr + 1, min(hr + 2000, (getattr(ws_src, "max_row", 0) or 0) + 1)):
                    company = ws_src.cell(row=rr, column=hc).value
                    addr = ws_src.cell(row=rr, column=hc + 1).value
                    if (company is None or str(company).strip() == "") and (addr is None or str(addr).strip() == ""):
                        break
                    company_s = str(company).strip() if company is not None else ""
                    addr_s = str(addr).strip() if addr is not None else ""
                    if company_s and addr_s:
                        addr_s = _clean_company_prefix(company_s, addr_s)
                        a1, a2, a3, ct, st, z = _split_address_lines(addr_s)
                        if a3:
                            a2 = f"{a2}, {a3}" if a2 else a3
                        rows.append((company_s, a1, a2, ct, st, z))
                if rows:
                    break

            # Write discovered rows (if any).
            if rows:
                _write_structured_rows(rows)

        # Always try to migrate legacy A/B into structured columns.
        _migrate_legacy_ab_to_structured()

        # One-time reset+seed of the screenshot defaults.
        # User request: clear existing table first, then import all addresses.
        # Marker lives in column H.
        marker_cell = ls.cell(row=1, column=8)
        marker = str(marker_cell.value or "").strip()
        if marker != "seeded_v4_reset":
            # Clear persisted supplier directory so it doesn't repopulate old rows.
            try:
                self._settings.remove("lists/supplier_directory_rows")
            except Exception:
                try:
                    self._settings.setValue("lists/supplier_directory_rows", "")
                except Exception:
                    pass

            # Overwrite the sheet with defaults (no merge).
            _write_structured_rows(list(DEFAULT_SUPPLIER_DIRECTORY_SEED))

            marker_cell.value = "seeded_v4_reset"

        # Always keep the computed Full Address (col G) in sync.
        # This is the stored value written into Form 1/2 cells when selecting a Company.
        try:
            max_scan = min(max(getattr(ls, "max_row", 0) or 0, 50), 5000)
            for rr in range(1, max_scan + 1):
                company = ls.cell(row=rr, column=1).value
                addr1 = ls.cell(row=rr, column=2).value
                addr2 = ls.cell(row=rr, column=3).value
                city = ls.cell(row=rr, column=4).value
                state = ls.cell(row=rr, column=5).value
                zipc = ls.cell(row=rr, column=6).value
                comp_s = str(company).strip() if company is not None else ""
                a1_s = str(addr1).strip() if addr1 is not None else ""
                a2_s = str(addr2).strip() if addr2 is not None else ""
                ct_s = str(city).strip() if city is not None else ""
                st_s = str(state).strip().upper() if state is not None else ""
                z_s = str(zipc).strip() if zipc is not None else ""
                if not (comp_s or a1_s or a2_s or ct_s or st_s or z_s):
                    continue
                ls.cell(row=rr, column=7).value = _build_full_address_with_company(comp_s, a1_s, a2_s, ct_s, st_s, z_s)
        except Exception:
            pass

        # Compute last row based on full-address column (G).
        last_row = 0
        max_scan = min(max(getattr(ls, "max_row", 0) or 0, 50), 5000)
        for rr in range(1, max_scan + 1):
            v = ls.cell(row=rr, column=7).value
            if v is None or str(v).strip() == "":
                continue
            last_row = rr
        if last_row <= 0:
            last_row = 1

        return sheet_name, last_row

    def _ensure_supplier_directory_dropdown(self, ws, cell_range: str) -> None:
        """Apply a mapped dropdown for supplier directory entries.

        Dropdown shows Company, writes Company + Address into the cell.
        """

        wb = ws.parent
        sheet_name, last_row = self._ensure_supplier_directory_sheet(wb)
        # Validation must reference stored values, which are the computed full addresses.
        rng = f"={sheet_name}!$G$1:$G${last_row}"

        dvs = getattr(ws, "data_validations", None)
        if dvs is None:
            ws.data_validations = openpyxl.worksheet.datavalidation.DataValidationList()
            dvs = ws.data_validations

        # If a list validation already exists for the first cell in range, update it.
        try:
            first_cell = str(cell_range).split(":", 1)[0]
        except Exception:
            first_cell = None

        if first_cell:
            for dv in list(getattr(dvs, "dataValidation", []) or []):
                if getattr(dv, "type", None) != "list":
                    continue
                try:
                    if first_cell in str(getattr(dv, "sqref", "")):
                        try:
                            dv.formula1 = rng
                        except Exception:
                            pass
                        return
                except Exception:
                    continue

        dv = DataValidation(type="list", formula1=rng, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(cell_range)

    def _ensure_form1_reason_dropdown(self, ws) -> None:
        """Ensure Form 1 has a dropdown for 'Reason for Full/Partial FAI'.

        The option list is long, so we store it on a hidden sheet and use a
        range-based data validation list.
        """

        # Values provided by the user (from screenshot)
        reasons = [
            "Change in Design - New Part Number",
            "Change in Design - New Revision",
            "Correct Previous FAIR",
            "Corrective Action for Defect History",
            "Create Net-Inspect Record for Legacy FAIR",
            "Delta FAI",
            "FAIR is a pass through from a supplier FAIR",
            "First Production Run",
            "Lapse in Production",
            "Last Article Inspection Report",
            "Natural or man-made event effecting production",
            "New Supplier",
            "Partial - By Similarity",
            "Process Change - Inspection Method",
            "Process Change - Manufacturing Location",
            "Process Change - Manufacturing Process",
            "Process Change - New Equipment/Equipment Move",
            "Process Change - Tooling Change",
            "Supplier Move",
        ]

        def _norm(s: str) -> str:
            return re.sub(r"\s+", " ", str(s or "").strip().lower())

        # Locate the header cell containing the label
        label_cell = None
        needle = "reason for full/partial fai:"
        for (r, c), cell in getattr(ws, "_cells", {}).items():
            v = getattr(cell, "value", None)
            if v and needle in _norm(v):
                label_cell = (int(r), int(c))
                break
        if not label_cell:
            return

        # The entry cell is directly under the label (same column, next row)
        r, c = label_cell
        target = ws.cell(row=r + 1, column=c)

        # Create / update hidden list sheet
        wb = ws.parent
        list_sheet_name = "__as9102_lists"
        if list_sheet_name in wb.sheetnames:
            ls = wb[list_sheet_name]
        else:
            ls = wb.create_sheet(list_sheet_name)
        try:
            ls.sheet_state = "hidden"
        except Exception:
            pass

        # Write the list (overwrite A1..An)
        for i, val in enumerate(reasons, start=1):
            ls.cell(row=i, column=1).value = val

        # Build a range-based list validation
        rng = f"={list_sheet_name}!$A$1:$A${len(reasons)}"

        # Avoid duplicating a validation if one already applies to the target cell
        dvs = getattr(ws, "data_validations", None)
        if dvs is None:
            ws.data_validations = openpyxl.worksheet.datavalidation.DataValidationList()
            dvs = ws.data_validations

        target_coord = target.coordinate
        for dv in list(getattr(dvs, "dataValidation", []) or []):
            if getattr(dv, "type", None) != "list":
                continue
            try:
                if target_coord in str(getattr(dv, "sqref", "")):
                    return
            except Exception:
                continue

        dv = DataValidation(type="list", formula1=rng, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(target_coord)

    def _ensure_form1_supplier_code_dropdown(self, ws) -> None:
        """Ensure Form 1 cell D9 has a dropdown showing customers but storing supplier codes.

        In Excel, the validation list must contain the stored values, so we validate against
        supplier codes. The UI will display customer names and write the code when selected.
        """

        # Target cell is fixed per template request.
        target = ws["D9"]

        # Seed list from screenshot (Customer -> Supplier Code)
        suppliers: list[tuple[str, str]] = [
            ("Raytheon", "10033672"),
            ("DRS", "10V001518"),
            ("SAES Getters", "Camtron Incorporated"),
            ("Vallen", "497735"),
        ]

        wb = ws.parent
        list_sheet_name = "__as9102_suppliers"
        created = False
        if list_sheet_name in wb.sheetnames:
            ls = wb[list_sheet_name]
        else:
            ls = wb.create_sheet(list_sheet_name)
            created = True
        try:
            ls.sheet_state = "hidden"
        except Exception:
            pass

        # Only seed the sheet if it is new/empty; otherwise preserve user edits.
        existing_a1 = ls.cell(row=1, column=1).value
        existing_b1 = ls.cell(row=1, column=2).value
        is_empty = (existing_a1 is None or str(existing_a1).strip() == "") and (existing_b1 is None or str(existing_b1).strip() == "")
        if created or is_empty:
            for i, (customer, code) in enumerate(suppliers, start=1):
                ls.cell(row=i, column=1).value = code
                ls.cell(row=i, column=2).value = customer

        # Compute current list length based on codes column (A).
        last_row = 0
        max_scan = min(max(getattr(ls, "max_row", 0) or 0, 200), 5000)
        for rr in range(1, max_scan + 1):
            v = ls.cell(row=rr, column=1).value
            if v is None or str(v).strip() == "":
                continue
            last_row = rr
        if last_row <= 0:
            last_row = len(suppliers)
        if last_row <= 0:
            last_row = 1

        rng = f"={list_sheet_name}!$A$1:$A${last_row}"

        dvs = getattr(ws, "data_validations", None)
        if dvs is None:
            ws.data_validations = openpyxl.worksheet.datavalidation.DataValidationList()
            dvs = ws.data_validations

        target_coord = target.coordinate
        for dv in list(getattr(dvs, "dataValidation", []) or []):
            if getattr(dv, "type", None) != "list":
                continue
            try:
                if target_coord in str(getattr(dv, "sqref", "")):
                    # Ensure this cell validates against the codes range.
                    try:
                        dv.formula1 = rng
                    except Exception:
                        pass
                    return
            except Exception:
                continue

        dv = DataValidation(type="list", formula1=rng, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(target_coord)

    def _write_form3_to_worksheet(self, ws) -> None:
        """Write Form 3 values into the loaded openpyxl worksheet."""
        if not self.characteristics:
            return

        try:
            dbg_gdt = str(os.environ.get("AS9102_DEBUG_GDT", "") or "").strip().lower() in ("1", "true", "yes", "on")
        except Exception:
            dbg_gdt = False

        if dbg_gdt:
            try:
                logger.debug(
                    "Form3 write start: ws=%s chars=%s max_row=%s max_col=%s",
                    getattr(ws, "title", ""),
                    len(self.characteristics or []),
                    getattr(ws, "max_row", None),
                    getattr(ws, "max_column", None),
                )
            except Exception:
                pass

        # Per request: always use installed-font mode and the "GDT" font.
        gdt_mode = "font"
        gdt_font_family = "GDT"
        enable_gdt_callout = True

        bubbled_numbers: set[int] = set()
        # Prefer the embedded Drawing Viewer tab's PDF viewer.
        try:
            dv = getattr(self, "drawing_viewer_tab", None)
            v = getattr(dv, "_pdf_viewer", None) if dv is not None else None
            if v is not None and hasattr(v, "get_bubbled_numbers"):
                bubbled_numbers = set(v.get_bubbled_numbers() or set())
        except Exception:
            bubbled_numbers = set()
        # Back-compat fallback if a different PDF window is used.
        if not bubbled_numbers:
            try:
                if getattr(self, "pdf_window", None) is not None:
                    v = getattr(self.pdf_window, "_pdf_viewer", None)
                    if v is not None and hasattr(v, "get_bubbled_numbers"):
                        bubbled_numbers = set(v.get_bubbled_numbers() or set())
            except Exception:
                bubbled_numbers = set()

        # Mirror the same header scan used by FaiGenerator.
        # Many templates have a 2-row header (e.g. row with "5. Char. No.", then a subheader row).
        start_row = 6
        header_row = 5
        for r in range(1, 20):
            val = ws.cell(row=r, column=2).value
            if val and "char no." in str(val).lower():
                header_row = r
                start_row = r + 1
                # If the next row still looks like header content (e.g. "Op #"), skip it.
                try:
                    row_text = " ".join(
                        [
                            str(ws.cell(row=start_row, column=c).value or "")
                            for c in range(1, min(ws.max_column, 30) + 1)
                        ]
                    ).lower()
                    if "op #" in row_text or "op#" in row_text or "reference location" in row_text or "bubble" in row_text:
                        start_row = r + 2
                except Exception:
                    pass
                break

        current_row = start_row
        row_num = 0

        # Optional: Unit-of-measure column (template-dependent).
        unit_col: int | None = None
        gdt_col: int | None = None
        tooling_col: int | None = None
        additional_col: int | None = None
        results_col: int | None = None
        bonus_tol_col: int | None = None
        header_row = max(1, int(header_row or max(1, start_row - 1)))
        header_rows_to_scan = [header_row]
        if header_row + 1 <= ws.max_row:
            header_rows_to_scan.append(header_row + 1)

        try:
            for hr in header_rows_to_scan:
                for cc in range(1, ws.max_column + 1):
                    hv = ws.cell(row=hr, column=cc).value
                    if not hv:
                        continue
                    h = str(hv).strip().lower()
                    if "unit" in h and ("measure" in h or "measurement" in h or "uom" in h):
                        unit_col = cc
                        break
                if unit_col is not None:
                    break
            # Fallback: common templates use column J (10)
            if unit_col is None:
                for hr in header_rows_to_scan:
                    hv = ws.cell(row=hr, column=10).value
                    if hv and "unit" in str(hv).lower():
                        unit_col = 10
                        break
        except Exception:
            unit_col = None

        # Optional: GD&T Callout column (template-dependent).
        try:
            for hr in header_rows_to_scan:
                for cc in range(1, ws.max_column + 1):
                    hv = ws.cell(row=hr, column=cc).value
                    if not hv:
                        continue
                    h = str(hv).strip().lower()
                    if ("gdt" in h and "call" in h) or h in ("gdt", "gdt callout", "gdt call out"):
                        gdt_col = cc
                        break
                if gdt_col is not None:
                    break
            # Handle merged header blocks where the visible text isn't in the left-most cell.
            if gdt_col is None:
                try:
                    merged = list(getattr(ws, "merged_cells", []).ranges or [])
                except Exception:
                    merged = []
                for hr in header_rows_to_scan:
                    for mr in merged:
                        try:
                            if not (mr.min_row <= hr <= mr.max_row):
                                continue
                            vals: list[str] = []
                            for cc in range(int(mr.min_col), int(mr.max_col) + 1):
                                v = ws.cell(row=hr, column=cc).value
                                if v is None:
                                    continue
                                s = str(v).strip()
                                if s:
                                    vals.append(s)
                            if not vals:
                                v = ws.cell(row=int(mr.min_row), column=int(mr.min_col)).value
                                if v is not None and str(v).strip():
                                    vals = [str(v).strip()]
                            blob = " ".join(vals).lower()
                            if ("gdt" in blob) and ("call" in blob):
                                gdt_col = int(mr.min_col)
                                break
                        except Exception:
                            continue
                    if gdt_col is not None:
                        break

            # Fallback: common templates place it immediately left of Unit-of-measure.
            if gdt_col is None and unit_col is not None:
                try:
                    if int(unit_col) > 1:
                        gdt_col = int(unit_col) - 1
                except Exception:
                    pass
        except Exception:
            gdt_col = None

        # Hard fallback: if detection failed for any reason, put GD&T immediately left of Unit.
        # This matches the common Form 3 layout (… Specification | GD&T Callout | Unit of measurement …).
        try:
            if gdt_col is None and unit_col is not None and int(unit_col) > 1:
                gdt_col = int(unit_col) - 1
        except Exception:
            pass

        if dbg_gdt:
            try:
                logger.debug(
                    "Form3 columns resolved: start_row=%s header_row=%s unit_col=%s gdt_col=%s mode=%s font=%s",
                    start_row,
                    header_row,
                    unit_col,
                    gdt_col,
                    gdt_mode,
                    gdt_font_family,
                )
            except Exception:
                pass

        # Safety: ensure GD&T column is not the same as Unit column.
        # If mis-detected (merged headers can cause this), force GD&T to be immediately
        # left of Unit so values don't get overwritten by the Unit write.
        try:
            if unit_col is not None and gdt_col is not None and int(gdt_col) == int(unit_col):
                if int(unit_col) > 1:
                    gdt_col = int(unit_col) - 1
        except Exception:
            pass

        def _norm_header(v: object) -> str:
            if v is None:
                return ""
            s = str(v).strip().lower()
            s = re.sub(r"[\s\t\r\n]+", " ", s)
            return s

        def _gdt_symbol_from_text(v: object) -> str:
            """Best-effort Unicode GD&T symbol from Calypso/spec text.

            Uses Unicode symbols where available; falls back to a short token when not.
            """
            t = str(v or "").strip()
            if not t:
                return ""
            tu = t.upper()

            # Order matters: specific phrases first.
            if "PROFILE OF A LINE" in tu or "PROFILE OF LINE" in tu:
                return "⌒"  # U+2312
            if "PROFILE" in tu:
                return "⌓"  # U+2313
            if "TRUE POSITION" in tu or "POSITION" in tu:
                return "⌖"  # U+2316
            if "PERPENDICULAR" in tu:
                return "⊥"  # U+22A5
            if "PARALLEL" in tu:
                return "∥"  # U+2225
            if "ANGULAR" in tu:
                return "∠"  # U+2220
            if "FLATNESS" in tu:
                return "⏥"  # U+23E5
            if "STRAIGHTNESS" in tu:
                return "⏤"  # U+23E4
            if "CIRCULARITY" in tu or "ROUNDNESS" in tu:
                return "○"  # U+25CB
            if "CYLINDRIC" in tu:
                return "⌭"  # U+232D
            if "CONCENTRIC" in tu:
                return "⊙"  # U+2299 (approx)
            if "SYMMETRY" in tu:
                return "≡"  # U+2261 (approx)
            if "CIRCULAR RUNOUT" in tu:
                return "⟲"  # U+27F2 (approx)
            if "TOTAL RUNOUT" in tu or "RUNOUT" in tu:
                return "⟳"  # U+27F3 (approx)

            # Modifiers/feature-control extras sometimes appear standalone.
            if "MMC" in tu:
                return "Ⓜ"  # U+24C2
            if "LMC" in tu:
                return "Ⓛ"  # U+24C1
            if "RFS" in tu:
                return "Ⓢ"  # U+24C8
            if "PROJECTED" in tu and "ZONE" in tu:
                return "Ⓟ"  # U+24C5

            if "SQUARE" in tu:
                return "□"  # U+25A1
            if "CENTERLINE" in tu or "CENTER LINE" in tu:
                return "℄"  # U+2104

            if "COUNTERBORE" in tu:
                return "⌴"  # U+2334
            if "COUNTERSINK" in tu:
                return "⌵"  # U+2335
            if "DEPTH" in tu:
                return "⌷"  # U+2337

            return ""

        def _gdt_font_code_from_text(v: object) -> str:
            """Return the single-letter code (Excel font-mapped) from text."""
            t = str(v or "").strip()
            if not t:
                return ""
            tu = t.upper()
            # Match most-specific first.
            mapping: list[tuple[str, str]] = [
                ("ANGULARITY", "a"),
                ("PERPENDICULAR", "b"),
                ("FLATNESS", "c"),
                ("PROFILE OF A LINE", "k"),
                ("PROFILE", "d"),
                ("CIRCULARITY", "e"),
                ("PARALLEL", "f"),
                ("CYLINDRIC", "g"),
                ("CIRCULAR RUNOUT", "h"),
                ("SYMMETRY", "i"),
                ("TRUE POSITION", "j"),
                ("POSITION", "j"),
                ("LMC", "l"),
                ("MMC", "m"),
                ("SQUARE", "o"),
                ("PROJECTED TOLERANCE ZONE", "p"),
                ("CENTERLINE", "q"),
                ("CONCENTRIC", "r"),
                ("RFS", "s"),
                ("TOTAL RUNOUT", "t"),
                ("STRAIGHTNESS", "u"),
                ("COUNTERBORE", "v"),
                ("COUNTERSINK", "w"),
                ("DEPTH", "x"),
                ("CONICAL TAPER", "y"),
                ("FLAT TAPER", "Z"),
            ]
            for key, code in mapping:
                if key in tu:
                    return code
            return ""

        def _truthy_flag(v: object) -> bool:
            s = str(v or "").strip().lower()
            if not s:
                return False
            return s not in ("0", "0.0", "false", "no", "off")

        def _datum_letter(v: object) -> str:
            """Extract a single datum letter (A/B/C/...) from a Calypso datum string."""
            s = str(v or "").strip()
            if not s:
                return ""
            # Prefer a standalone single letter token.
            m = re.findall(r"\b([A-Za-z])\b", s)
            if m:
                return str(m[-1]).upper()
            # Fallback: any letter at all.
            m2 = re.findall(r"([A-Za-z])", s)
            if m2:
                return str(m2[-1]).upper()
            return ""

        def _mmc_symbol(*, mode: str) -> str:
            return "m" if mode == "font" else "Ⓜ"

        def _build_gdt_callout(
            *,
            mode: str,
            symbol_font_code: str,
            symbol_unicode: str,
            tolerance_text: str,
            mmc_flag: bool,
            datums: list[str],
        ) -> str:
            """Build a callout text similar to the provided Excel formula.

            Font mode uses the template's GD&T font mapping (single-letter codes).
            Unicode mode uses Unicode GD&T symbols.
            """

            tol = str(tolerance_text or "").strip()
            if not tol:
                return ""

            sym = symbol_font_code if mode == "font" else symbol_unicode
            sym = str(sym or "").strip()
            if not sym:
                return ""

            # Match the provided Excel logic:
            # - if BG is i or n -> just return BG
            if mode == "font" and symbol_font_code in ("i", "n"):
                return symbol_font_code

            mmc = _mmc_symbol(mode=mode) if mmc_flag else ""

            parts: list[str] = [f"{sym}|{tol}{mmc}"]
            for d in datums:
                dd = str(d or "").strip().upper()
                if dd:
                    parts.append(dd)
            return "|".join(parts)

        def _is_gdt_callout_row(char_obj: object) -> bool:
            """Return True only for rows that should display a GD&T callout.

            Calypso exports often include helper/component rows like .X/.Z that are
            "basic" components for a GD&T position and should NOT display a callout.
            Also exclude pure size dimensions like Diameter of Cylinder/Circle.
            """
            try:
                idsym = str(getattr(char_obj, "idsymbol", "") or "").strip()
            except Exception:
                idsym = ""
            try:
                typ = str(getattr(char_obj, "type", "") or "").strip()
            except Exception:
                typ = ""
            try:
                cid = str(getattr(char_obj, "id", "") or "").strip()
            except Exception:
                cid = ""

            idsym_l = idsym.lower()
            typ_u = typ.upper()
            cid_u = cid.upper()

            # Exclude X/Z component rows (e.g., True Position ... .X / .Z)
            if idsym_l.endswith(".x") or idsym_l.endswith(".z"):
                return False
            if cid_u.endswith(".X") or cid_u.endswith(".Z"):
                return False
            if ".X" in typ_u or ".Z" in typ_u:
                return False
            if typ_u in ("X VALUE", "Y VALUE", "Z VALUE"):
                return False

            # Exclude diameter size dimensions (not GD&T callouts)
            if idsym_l == "diameter" or typ_u == "DIAMETER":
                return False

            # Most GD&T rows in this export have idsymbol starting with 'gdt'.
            # Keep that as a strong signal.
            if idsym_l.startswith("gdt"):
                return True

            # Fallback: allow explicit GD&T types.
            return any(
                k in typ_u
                for k in (
                    "POSITION",
                    "TRUE POSITION",
                    "FLATNESS",
                    "PERPENDICULAR",
                    "PROFILE",
                    "PARALLEL",
                    "ANGULAR",
                    "STRAIGHT",
                    "CIRCULAR",
                    "CYLINDRIC",
                    "RUNOUT",
                    "SYMMETRY",
                    "CONCENTRIC",
                )
            )

        # Discover Form 3 columns for:
        # - 10. Designed/Qualified Tooling
        # - 12. Additional Data/Comments (first column)
        # - Results (if present; fallback stays at 12 for older templates)
        try:
            for hr in header_rows_to_scan:
                for cc in range(1, ws.max_column + 1):
                    hv = ws.cell(row=hr, column=cc).value
                    if hv is None or str(hv).strip() == "":
                        continue
                    h = _norm_header(hv)

                    if tooling_col is None and ("tooling" in h) and ("designed" in h) and ("qualified" in h):
                        # If merged, prefer left-most column of the merged range.
                        tooling_col = cc
                        try:
                            coord = ws.cell(row=hr, column=cc).coordinate
                            for mr in getattr(ws, "merged_cells", []).ranges:
                                if coord in mr:
                                    tooling_col = mr.min_col
                                    break
                        except Exception:
                            pass

                    if additional_col is None and ("comment" in h) and ("additional" in h or "addtion" in h) and ("data" in h):
                        additional_col = cc
                        try:
                            coord = ws.cell(row=hr, column=cc).coordinate
                            for mr in getattr(ws, "merged_cells", []).ranges:
                                if coord in mr:
                                    additional_col = mr.min_col
                                    break
                        except Exception:
                            pass

                    # Some templates have a separate "Bonus Tolerance" column near Results.
                    # Never treat that as the Results column.
                    if bonus_tol_col is None and ("bonus" in h and "tolerance" in h):
                        bonus_tol_col = cc

                    if results_col is None and ("result" in h or "results" in h or "actual" in h):
                        if not ("bonus" in h and "tolerance" in h):
                            results_col = cc

            # Some templates put header text inside a merged range but not necessarily in
            # the left-most cell; scan merged header blocks to find the true start column.
            try:
                merged = list(getattr(ws, "merged_cells", []).ranges or [])
            except Exception:
                merged = []
            for hr in header_rows_to_scan:
                for mr in merged:
                    try:
                        if not (mr.min_row <= hr <= mr.max_row):
                            continue
                        # Inspect the values across this merged block on the header row.
                        vals: list[str] = []
                        for cc in range(int(mr.min_col), int(mr.max_col) + 1):
                            v = ws.cell(row=hr, column=cc).value
                            if v is None:
                                continue
                            s = str(v).strip()
                            if s:
                                vals.append(s)
                        if not vals:
                            # Fallback to the top-left value for the merged range.
                            v = ws.cell(row=int(mr.min_row), column=int(mr.min_col)).value
                            if v is not None and str(v).strip():
                                vals = [str(v).strip()]
                        blob = " ".join(vals).lower()
                        if tooling_col is None and ("tooling" in blob) and ("designed" in blob) and ("qualified" in blob):
                            tooling_col = int(mr.min_col)
                        if additional_col is None and ("comment" in blob) and ("additional" in blob or "addtion" in blob) and ("data" in blob):
                            additional_col = int(mr.min_col)
                        if bonus_tol_col is None and ("bonus" in blob and "tolerance" in blob):
                            bonus_tol_col = int(mr.min_col)
                        if results_col is None and ("result" in blob or "results" in blob or "actual" in blob):
                            if not ("bonus" in blob and "tolerance" in blob):
                                results_col = int(mr.min_col)
                    except Exception:
                        continue
        except Exception:
            pass

        # If Results accidentally matched Bonus Tolerance, try to find a different Results column.
        try:
            if results_col is not None and bonus_tol_col is not None and int(results_col) == int(bonus_tol_col):
                for hr in header_rows_to_scan:
                    for cc in range(1, ws.max_column + 1):
                        if int(cc) == int(bonus_tol_col):
                            continue
                        hv = ws.cell(row=hr, column=cc).value
                        if hv is None or str(hv).strip() == "":
                            continue
                        h = _norm_header(hv)
                        if ("result" in h or "results" in h or "actual" in h) and not ("bonus" in h and "tolerance" in h):
                            results_col = cc
                            raise StopIteration
        except StopIteration:
            pass
        except Exception:
            pass

        # Extra fallbacks for the common template shown in screenshots:
        # J=Unit (10), K=Results (11), L=Tooling (12), P=Additional Data/Comments (16)
        try:
            # If header detection failed (often due to merged/blank header cells),
            # fall back to the known columns for this template.
            if tooling_col is None and ws.max_column >= 12:
                tooling_col = 12

            if results_col is None and ws.max_column >= 11:
                results_col = 11

            # User template revision: Additional Data/Comments is column Q (17)
            if additional_col is None and ws.max_column >= 17:
                additional_col = 17
        except Exception:
            pass

        # Last-resort fallback.
        if results_col is None:
            results_col = 12

        # Resolve selected machine values once (applied only to Calypso-imported rows).
        tooling_value = ""
        additional_value = ""
        try:
            details = self._selected_calibrated_equipment_details()
            if details:
                machine_id, machine_type, due_date = details
                mid_clean = re.sub(
                    r"^\s*(?:gage\s*id|id)\s*:\s*",
                    "",
                    str(machine_id or ""),
                    flags=re.IGNORECASE,
                ).strip()
                due_clean = re.sub(
                    r"^\s*(?:cal\s*due|due)\s*:\s*",
                    "",
                    str(due_date or ""),
                    flags=re.IGNORECASE,
                ).strip()
                tooling_parts: list[str] = []
                if mid_clean:
                    tooling_parts.append(f"Gage ID: {mid_clean}")
                if due_clean:
                    tooling_parts.append(f"Cal Due: {due_clean}")
                tooling_value = "  ".join(tooling_parts).strip()
                additional_value = str(machine_type or "").strip()
        except Exception:
            tooling_value = ""
            additional_value = ""

        # Excel-like light red/green fills
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

        # Preserve any user-entered Reference Location (col D / 4) per characteristic id
        # so toggling the derived-row checkbox doesn't scramble the values.
        ref_location_by_id: dict[str, str] = {}

        def _parse_char_id_from_desc(v: object) -> str:
            s = str(v or "").strip()
            if not s:
                return ""
            # description_text is written as "{id} {feature_name}"; id is the first token.
            return s.split()[0].strip()

        def _find_table_end_row() -> int:
            # Walk downward until we hit a stretch of empty rows in key columns.
            # This prevents leaving stale rows visible when the derived rows are hidden.
            max_scan = min(max(int(getattr(ws, "max_row", 0) or 0), start_row + 250), start_row + 2000)
            started = False
            empty_run = 0
            last_seen = start_row
            for rr in range(start_row, max_scan + 1):
                c2 = ws.cell(row=rr, column=2).value
                c7 = ws.cell(row=rr, column=7).value
                has_any = (c2 is not None and str(c2).strip() != "") or (c7 is not None and str(c7).strip() != "")
                if has_any:
                    started = True
                    empty_run = 0
                    last_seen = rr
                else:
                    if started:
                        empty_run += 1
                        if empty_run >= 25:
                            break
            return max(start_row, last_seen)

        end_row = _find_table_end_row()

        # Base font for the GDT callout column (used to restore when switching away from font mode).
        base_gdt_font = None
        try:
            if gdt_col is not None and enable_gdt_callout:
                base_gdt_font = ws.cell(row=int(start_row), column=int(gdt_col)).font
        except Exception:
            base_gdt_font = None

        # If the user picked Font mode but didn't choose a font family,
        # prefer the template's existing font for that column.
        try:
            if gdt_mode == "font" and not gdt_font_family and base_gdt_font is not None:
                nm = str(getattr(base_gdt_font, "name", "") or "").strip()
                if nm:
                    gdt_font_family = nm
        except Exception:
            pass

        # Capture existing ref locations before clearing.
        for rr in range(start_row, end_row + 1):
            try:
                desc = ws.cell(row=rr, column=7).value
                char_id = _parse_char_id_from_desc(desc)
                if not char_id:
                    continue
                ref_val = ws.cell(row=rr, column=4).value
                if ref_val is None:
                    continue
                ref_s = str(ref_val).strip()
                if ref_s:
                    ref_location_by_id[char_id] = ref_s
            except Exception:
                continue

        # Clear any accidental calibrated-equipment writes in the header area.
        # These show up when templates have merged blocks to the right of the table.
        try:
            details = self._selected_calibrated_equipment_details()
            if details:
                _mid, _mtype, _due = details
                header_max_row = max(1, int(start_row) - 1)
                # Common problematic area (per screenshot): columns O/Q in top rows.
                for rr in range(1, min(header_max_row, 25) + 1):
                    for cc in (15, 16, 17):  # O, P, Q
                        try:
                            cell = ws.cell(row=rr, column=cc)
                            v = str(cell.value or "").strip()
                            if not v:
                                continue
                            vv = v.lower()
                            if ("gage id:" in vv and "cal due:" in vv) or v == str(_mtype or "").strip():
                                cell.value = None
                        except Exception:
                            continue
        except Exception:
            pass

        # Clear prior table values + fills for the columns we control.
        # Column D (Reference Location) IS cleared and then restored per id.
        cols_to_clear: list[int] = [2, 4, 5, 7, 8, int(results_col or 12)]
        if unit_col is not None:
            cols_to_clear.append(int(unit_col))
        if gdt_col is not None:
            cols_to_clear.append(int(gdt_col))
        if tooling_col is not None:
            cols_to_clear.append(int(tooling_col))
        if additional_col is not None:
            cols_to_clear.append(int(additional_col))
        if bonus_tol_col is not None:
            cols_to_clear.append(int(bonus_tol_col))
        cols_to_clear = sorted(set([c for c in cols_to_clear if c and c > 0]))

        default_fill = PatternFill()  # openpyxl default
        for rr in range(start_row, end_row + 1):
            for cc in cols_to_clear:
                try:
                    cell = ws.cell(row=rr, column=cc)
                    # Never clear Bonus Tolerance values (often formulas); only manage its fill.
                    try:
                        if bonus_tol_col is not None and int(cc) == int(bonus_tol_col):
                            cell.fill = default_fill
                            continue
                    except Exception:
                        pass

                    cell.value = None
                    cell.fill = default_fill
                    if gdt_col is not None and int(cc) == int(gdt_col):
                        # Avoid leaving the GD&T font stuck on when switching to Unicode.
                        if gdt_mode == "unicode" and base_gdt_font is not None:
                            try:
                                cell.font = base_gdt_font
                            except Exception:
                                pass
                except Exception:
                    continue

        for char in self.characteristics:
            # Optionally hide derived thread rows (Go/No-Go, Minor Dia, etc.)
            # when the Form 3 checkbox is unchecked.
            if not bool(getattr(self, "_form3_include_thread_extras", True)):
                try:
                    if str(getattr(char, "source", "calypso") or "calypso").strip().lower() != "calypso":
                        continue
                except Exception:
                    pass

            if not getattr(char, "description", None) or not str(char.description).strip():
                continue
            if "nan" in str(char.description).lower():
                continue

            row_num += 1

            ws.cell(row=current_row, column=2).value = row_num
            # Bubble Number is column E in the template (header row shows
            # D: 6. Reference Location, E: Bubble Number)
            ws.cell(row=current_row, column=5).value = row_num

            # Reference Location is column D (4). Preserve user-entered values.
            try:
                group1_val = getattr(char, "group1", "")
                ref_cell = ws.cell(row=current_row, column=4)
                existing_ref = ""
                try:
                    existing_ref = ref_location_by_id.get(str(getattr(char, "id", "") or "").strip(), "")
                except Exception:
                    existing_ref = ""
                if existing_ref:
                    ref_cell.value = existing_ref
                elif group1_val is not None and str(group1_val).strip():
                    ref_cell.value = str(group1_val).strip()
            except Exception:
                pass

            description_text = f"{getattr(char, 'id', '')}".strip()
            desc_cell = ws.cell(row=current_row, column=7)
            desc_cell.value = description_text
            try:
                desc_cell.alignment = (desc_cell.alignment or Alignment()).copy(wrap_text=True)
            except Exception:
                # If the openpyxl Alignment API differs, fall back to a simple wrap.
                desc_cell.alignment = Alignment(wrap_text=True)
            ws.cell(row=current_row, column=8).value = char.description

            basic_text = " ".join(
                [
                    str(description_text or ""),
                    str(getattr(char, "description", "") or ""),
                    str(getattr(char, "comment", "") or ""),
                    str(getattr(char, "feature_name", "") or ""),
                ]
            )
            is_basic = bool(re.search(r"\bbasic\b", basic_text, flags=re.IGNORECASE))

            # GD&T callout (best-effort) from Calypso/spec text.
            if gdt_col is not None and enable_gdt_callout and not is_basic:
                try:
                    # Only emit callouts for real GD&T rows.
                    if not _is_gdt_callout_row(char):
                        raise RuntimeError("not_gdt_row")

                    # Calypso imports often put the GD&T type in the feature name, while the
                    # spec text is just the numeric requirement (e.g. ".0100 MAX").
                    gdt_source = " ".join(
                        [
                            str(getattr(char, "type", "") or "").strip(),
                            str(getattr(char, "idsymbol", "") or "").strip(),
                            str(getattr(char, "id", "") or "").strip(),
                            str(getattr(char, "feature_name", "") or "").strip(),
                            str(getattr(char, "description", "") or "").strip(),
                        ]
                    ).strip()

                    symbol_code = _gdt_font_code_from_text(gdt_source)
                    symbol_unicode = _gdt_symbol_from_text(gdt_source)
                    tol_text = str(getattr(char, "description", "") or "").strip()
                    # Clean up tolerance text to only include the number (remove "MAX", "MIN", etc.)
                    _m = re.search(r"(\d*\.?\d+)", tol_text)
                    if _m:
                        tol_text = _m.group(1)
                    mmc_flag = _truthy_flag(getattr(char, "mmc", ""))

                    datum_letters: list[str] = []
                    try:
                        for raw in (
                            getattr(char, "datumaid", ""),
                            getattr(char, "datumbid", ""),
                            getattr(char, "datumcid", ""),
                        ):
                            d = _datum_letter(raw)
                            if d and d not in datum_letters:
                                datum_letters.append(d)
                    except Exception:
                        datum_letters = []

                    # Default datum structure to A|B|C when missing (per request).

                        # Snapshot for Ctrl+Z undo.
                        try:
                            print("Form3 delete (multi): calling _push_form3_undo_state")
                        except Exception:
                            pass
                        self._push_form3_undo_state()
                        try:
                            print("Form3 delete: snapshot taken (multi)")
                        except Exception:
                            pass

                    for d in ("A", "B", "C"):
                        if d not in datum_letters:
                            datum_letters.append(d)

                    callout = _build_gdt_callout(
                        mode=gdt_mode,
                        symbol_font_code=symbol_code,
                        symbol_unicode=symbol_unicode,
                        tolerance_text=tol_text,
                        mmc_flag=mmc_flag,
                        datums=datum_letters,
                    )

                    if callout:
                        gcell = ws.cell(row=current_row, column=gdt_col)
                        gcell.value = callout
                        if gdt_mode == "font" and gdt_font_family:
                            try:
                                f = getattr(gcell, "font", None)
                                if f is not None and hasattr(f, "copy"):
                                    gcell.font = f.copy(name=gdt_font_family)
                                else:
                                    gcell.font = Font(name=gdt_font_family)
                            except Exception:
                                try:
                                    gcell.font = Font(name=gdt_font_family)
                                except Exception:
                                    pass

                    # If the GD&T callout includes MMC, highlight Bonus Tolerance yellow (same row).
                    try:
                        mmc_sym = _mmc_symbol(mode=gdt_mode)
                        if bonus_tol_col is not None and callout and (mmc_sym in str(callout)):
                            ws.cell(row=current_row, column=int(bonus_tol_col)).fill = yellow_fill
                    except Exception:
                        pass

                    # Optional debug logging to verify callout generation.
                    if dbg_gdt and row_num <= 5:
                        try:
                            logger.debug(
                                "Form3 GDT row=%s col=%s mode=%s font=%s source=%s => %s",
                                current_row,
                                gdt_col,
                                gdt_mode,
                                gdt_font_family,
                                gdt_source,
                                callout,
                            )
                        except Exception:
                            pass
                except Exception:
                    pass

            if unit_col is not None:
                unit_val = getattr(char, "unit", "")
                if unit_val is not None and str(unit_val).strip():
                    ws.cell(row=current_row, column=unit_col).value = str(unit_val).strip()

            # Results column (template-dependent; default is L/12).
            # Per request:
            # - If the word 'Basic' appears in the description/note text, Results must be 'NA'.
            # - Otherwise, numeric results should always be positive and displayed to 4 decimals.

            # Ensure GD&T callout stays blank (and explicitly clear template formulas).
            if gdt_col is not None and (not enable_gdt_callout or is_basic):
                try:
                    gcell = ws.cell(row=current_row, column=int(gdt_col))
                    gcell.value = None
                    if base_gdt_font is not None:
                        try:
                            gcell.font = base_gdt_font
                        except Exception:
                            pass
                except Exception:
                    pass

            result_value: object = char.actual if char.actual is not None else ""
            if is_basic:
                result_value = "NA"
            else:
                try:
                    result_value = round(abs(float(char.actual)), 4)
                except (ValueError, TypeError):
                    pass

            result_cell = ws.cell(row=current_row, column=int(results_col or 12))
            result_cell.value = result_value

            # Ensure numeric results display with exactly 4 decimals.
            try:
                if isinstance(result_value, (int, float)):
                    result_cell.number_format = "0.0000"
            except Exception:
                pass

            # Pass/fail shading
            fill = None
            if getattr(char, "is_attribute", False):
                if is_basic:
                    fill = None
                elif str(char.actual).lower() == "pass":
                    fill = green_fill
                elif not char.actual:
                    fill = red_fill
                else:
                    fill = red_fill
            else:
                try:
                    if is_basic:
                        raise ValueError("basic")

                    # Per request: evaluate numeric limits in the same absolute-value space
                    # as we display in Form 3 (positive nominal/result).
                    val = abs(float(char.actual))
                    nom = abs(float(char.nominal))
                    up = float(char.upper_tol) if char.upper_tol else 0.0
                    low = float(char.lower_tol) if char.lower_tol else 0.0

                    if abs(up) < 990:
                        limit_high = nom + up
                        limit_low = nom + low
                        if val > limit_high + 1e-6 or val < limit_low - 1e-6:
                            fill = red_fill
                except (ValueError, TypeError):
                    if (not char.actual) and (not is_basic):
                        fill = red_fill

            if fill is not None:
                result_cell.fill = fill

            # Fill Form 3 column 10/12 style fields (tooling/comments) for Calypso rows only.
            is_calypso_row = str(getattr(char, "source", "calypso") or "calypso").strip().lower() == "calypso"
            if is_calypso_row:
                if tooling_col is not None and tooling_value:
                    try:
                        ws.cell(row=current_row, column=tooling_col).value = tooling_value
                    except Exception:
                        pass
                if additional_col is not None and additional_value:
                    try:
                        ws.cell(row=current_row, column=additional_col).value = additional_value
                    except Exception:
                        pass
            else:
                # Highlight derived thread/minor-pin rows in red.
                try:
                    # Do not color Char No (col B / 2); users expect it to remain unfilled.
                    cols_to_mark: list[int] = [4, 5, 7, 8, int(results_col or 12)]
                    if unit_col is not None:
                        cols_to_mark.append(unit_col)
                    if tooling_col is not None:
                        cols_to_mark.append(tooling_col)
                    if additional_col is not None:
                        cols_to_mark.append(additional_col)
                    for cc in sorted(set([c for c in cols_to_mark if c and c > 0])):
                        ws.cell(row=current_row, column=cc).fill = red_fill
                except Exception:
                    pass

            # Bubble status shading (always applied last so it wins):
            # - default red until bubble exists on drawing
            # - green when bubble (or range containing it) exists
            try:
                bcell = ws.cell(row=current_row, column=5)
                bcell.fill = green_fill if int(row_num) in bubbled_numbers else red_fill
            except Exception:
                pass

            current_row += 1

    def _update_form3_bubble_fills(self, bubbled_numbers: set[int]) -> None:
        """Apply red/green fills to Form 3 Bubble Number cells based on drawing bubbles."""

        if self._template_wb is None:
            return

        ws3_name = self._form_sheet_names.get("3")
        if not ws3_name or ws3_name not in self._template_wb.sheetnames:
            return

        ws = self._template_wb[ws3_name]
        viewer = self._form_viewers.get("3")

        # Discover the start row similar to _write_form3_to_worksheet.
        start_row = 6
        header_row = 5
        try:
            for r in range(1, 20):
                val = ws.cell(row=r, column=2).value
                if val and "char no." in str(val).lower():
                    header_row = r
                    start_row = r + 1
                    try:
                        row_text = " ".join(
                            [
                                str(ws.cell(row=start_row, column=c).value or "")
                                for c in range(1, min(ws.max_column, 30) + 1)
                            ]
                        ).lower()
                        if "op #" in row_text or "op#" in row_text or "reference location" in row_text or "bubble" in row_text:
                            start_row = r + 2
                    except Exception:
                        pass
                    break
        except Exception:
            start_row = 6

        # Determine an end row by scanning for a run of empties.
        try:
            max_scan = min(max(int(getattr(ws, "max_row", 0) or 0), start_row + 250), start_row + 2000)
            started = False
            empty_run = 0
            end_row = start_row
            for rr in range(start_row, max_scan + 1):
                c2 = ws.cell(row=rr, column=2).value
                c7 = ws.cell(row=rr, column=7).value
                has_any = (c2 is not None and str(c2).strip() != "") or (c7 is not None and str(c7).strip() != "")
                if has_any:
                    started = True
                    empty_run = 0
                    end_row = rr
                else:
                    if started:
                        empty_run += 1
                        if empty_run >= 25:
                            break
        except Exception:
            end_row = start_row

        red_rgb = "FFC7CE"
        green_rgb = "C6EFCE"

        # Fetch reference location strings if available (mode is selected in Drawing Viewer)
        bubble_zones = {}
        ref_mode = "sheet_zone"
        try:
            dv = getattr(self, "drawing_viewer_tab", None)
            pv = getattr(dv, "_pdf_viewer", None) if dv is not None else None
            if pv is not None and hasattr(pv, "reference_location_mode"):
                ref_mode = str(getattr(pv, "reference_location_mode", "sheet_zone") or "sheet_zone")
            if pv is not None and hasattr(pv, "get_reference_locations"):
                bubble_zones = pv.get_reference_locations(ref_mode)
            elif pv is not None and hasattr(pv, "get_bubble_zones"):
                bubble_zones = pv.get_bubble_zones()
        except Exception:
            pass

        # Update fills based on the bubble number in column 5 (E).
        for rr in range(int(start_row), int(end_row) + 1):
            try:
                v = ws.cell(row=rr, column=5).value
                n = int(v)
            except Exception:
                continue
            
            is_internal = n in (bubbled_numbers or set())
            rgb = green_rgb if is_internal else red_rgb
            
            # Update Bubble Number Cell (Col 5)
            try:
                # Ensure Bubble numbers remain readable: if the cell has an explicit
                # font color (e.g., white), Qt will honor it even on light fills.
                # Clearing the explicit font color lets the viewer pick a contrasting
                # foreground automatically.
                try:
                    c = ws.cell(row=rr, column=5)
                    f = getattr(c, "font", None)
                    if f is not None and getattr(f, "color", None) is not None:
                        try:
                            c.font = f.copy(color=None)
                        except Exception:
                            pass
                except Exception:
                    pass

                if viewer is not None and hasattr(viewer, "_apply_cell_fill"):
                    viewer._apply_cell_fill(rr, 5, rgb, push_undo=False)
                else:
                    ws.cell(row=rr, column=5).fill = PatternFill(start_color=rgb, end_color=rgb, fill_type="solid")
            except Exception:
                pass
                
            # Update Reference Location Cell (Col 4)
            # If internal (Green), update value to Zone and set color Green.
            # If external/missing (Red), set color Red.
            try:
                if str(ref_mode or "").strip().lower() in ("none", "off", "disable", "disabled"):
                    # Clear Reference Location values when mode is None.
                    try:
                        if viewer is not None and hasattr(viewer, "_apply_cell_value"):
                            viewer._apply_cell_value(rr, 4, "", push_undo=False)
                        else:
                            ws.cell(row=rr, column=4).value = ""
                    except Exception:
                        pass

                    if is_internal:
                        if viewer is not None and hasattr(viewer, "_apply_cell_fill"):
                            viewer._apply_cell_fill(rr, 4, green_rgb, push_undo=False)
                        else:
                            ws.cell(row=rr, column=4).fill = PatternFill(start_color=green_rgb, end_color=green_rgb, fill_type="solid")
                    else:
                        if viewer is not None and hasattr(viewer, "_apply_cell_fill"):
                            viewer._apply_cell_fill(rr, 4, red_rgb, push_undo=False)
                        else:
                            ws.cell(row=rr, column=4).fill = PatternFill(start_color=red_rgb, end_color=red_rgb, fill_type="solid")
                    continue

                if is_internal:
                    # Update Value
                    if n in bubble_zones:
                        zone_str = bubble_zones[n]
                        if viewer is not None and hasattr(viewer, "_apply_cell_value"):
                            viewer._apply_cell_value(rr, 4, zone_str, push_undo=False)
                        else:
                            ws.cell(row=rr, column=4).value = zone_str
                    
                    # Update Color (Green)
                    if viewer is not None and hasattr(viewer, "_apply_cell_fill"):
                        viewer._apply_cell_fill(rr, 4, green_rgb, push_undo=False)
                    else:
                        ws.cell(row=rr, column=4).fill = PatternFill(start_color=green_rgb, end_color=green_rgb, fill_type="solid")
                else:
                    # Bubble is missing on the drawing: do not leave a populated Reference Location.
                    try:
                        if viewer is not None and hasattr(viewer, "_apply_cell_value"):
                            viewer._apply_cell_value(rr, 4, "", push_undo=False)
                        else:
                            ws.cell(row=rr, column=4).value = ""
                    except Exception:
                        pass
                    # Update Color (Red)
                    if viewer is not None and hasattr(viewer, "_apply_cell_fill"):
                        viewer._apply_cell_fill(rr, 4, red_rgb, push_undo=False)
                    else:
                        ws.cell(row=rr, column=4).fill = PatternFill(start_color=red_rgb, end_color=red_rgb, fill_type="solid")
            except Exception:
                pass

    def _on_drawing_bubbles_changed(self, bubbled_numbers: object) -> None:
        try:
            s = set(int(x) for x in (bubbled_numbers or set()))
        except Exception:
            s = set()
        try:
            self._last_bubbled_numbers = set(s)
        except Exception:
            pass
        self._update_form3_bubble_fills(s)
        try:
            v3 = self._form_viewers.get("3")
            if v3 is not None:
                tbl = getattr(v3, "table", None)
                if tbl is not None:
                    tbl.viewport().update()
        except Exception:
            pass

    def _on_drawing_saved(self, out_path: str) -> None:
        p = str(out_path or "").strip()
        if not p:
            return
        try:
            self.drawing_pdf_path = p
        except Exception:
            pass

        try:
            if hasattr(self, "drawing_pdf_edit") and self.drawing_pdf_edit is not None:
                self.drawing_pdf_edit.setText(p)
        except Exception:
            pass

        try:
            if hasattr(self, "_settings") and self._settings is not None:
                self._settings.setValue("paths/drawing_pdf", p)
        except Exception:
            pass

        # Keep embedded Drawing Viewer tab (if present) on the latest saved PDF.
        try:
            v = getattr(self, "drawing_viewer_tab", None)
            if v is not None and os.path.exists(p):
                v.load_pdf(p)
                QTimer.singleShot(75, self._sync_bubbles_to_form3)
        except Exception:
            pass

    def _on_insert_notes_to_form3(self, note_text: str, source_dialog=None) -> None:
        raw = str(note_text or "").replace("\u2029", "\n")
        raw = raw.replace("\r\n", "\n").replace("\r", "\n")
        raw = raw.strip()
        if not raw:
            return

        # If the user selected the region header too, strip it.
        try:
            raw = re.sub(r"(?is)^\s*Page\s+\d+\s*/\s*Region\s+\d+\s*:\s*", "", raw).strip()
        except Exception:
            pass

        def _split_numbered_notes(s: str) -> list[str]:
            # Keep newlines for splitting, but collapse other whitespace.
            try:
                s = re.sub(r"[ \t]+", " ", s)
                s = re.sub(r"\n{3,}", "\n\n", s)
            except Exception:
                pass
            s = s.strip()
            if not s:
                return []

            # Find note markers like "1. " (require a trailing space so we don't split decimals).
            try:
                matches = list(re.finditer(r"(?m)(^|\n)(\d{1,3}\.\s)", s))
            except Exception:
                matches = []

            if not matches:
                one = re.sub(r"\s+", " ", s).strip()
                return [one] if one else []

            parts: list[str] = []
            for i, m in enumerate(matches):
                start = int(m.start(2))
                end = int(matches[i + 1].start(2)) if i + 1 < len(matches) else len(s)
                seg = s[start:end].strip()
                seg = re.sub(r"\s+", " ", seg).strip()
                if seg:
                    parts.append(seg)
            return parts

        notes = [n for n in _split_numbered_notes(raw) if str(n or "").strip()]
        if not notes:
            return

        # Optional progress popup (only for requests originating from Extracted Notes).
        loading_dlg = None

        def _focus_form3() -> None:
            try:
                if hasattr(self, "forms_tabs") and self.forms_tabs is not None:
                    idx3 = -1
                    for i in range(int(self.forms_tabs.count())):
                        try:
                            if str(self.forms_tabs.tabText(i)).strip().lower() == "form 3":
                                idx3 = i
                                break
                        except Exception:
                            continue
                    if idx3 != -1:
                        self.forms_tabs.setCurrentIndex(idx3)
                v3 = self._form_viewers.get("3") if hasattr(self, "_form_viewers") else None
                if v3 is not None:
                    tbl = getattr(v3, "table", None)
                    if tbl is not None:
                        tbl.setFocus()
                        try:
                            sb = tbl.horizontalScrollBar()
                            if sb is not None:
                                sb.setValue(int(sb.minimum()))
                        except Exception:
                            pass
            except Exception:
                pass

        def _loading_init() -> None:
            nonlocal loading_dlg
            if source_dialog is None:
                return
            try:
                loading_dlg = QDialog(self)
                loading_dlg.setWindowTitle("Loading")
                loading_dlg.setWindowModality(Qt.WindowModality.ApplicationModal)
                loading_dlg.setMinimumWidth(320)

                lay = QVBoxLayout(loading_dlg)
                lay.addWidget(QLabel("Loading…"))
                bar = QProgressBar()
                bar.setRange(0, 0)  # indeterminate
                lay.addWidget(bar)

                # Disable close button.
                try:
                    loading_dlg.setWindowFlag(Qt.WindowType.WindowCloseButtonHint, False)
                except Exception:
                    pass

                loading_dlg.show()

                # Also disable the insert button in the Extracted Notes window.
                try:
                    if hasattr(source_dialog, "insert_selected_btn"):
                        source_dialog.insert_selected_btn.setEnabled(False)
                except Exception:
                    pass
            except Exception:
                loading_dlg = None

        def _loading_finish_and_close() -> None:
            # Close loading popup + Extracted Notes, then focus Form 3.
            try:
                if loading_dlg is not None:
                    loading_dlg.close()
            except Exception:
                pass

            # Clean notes window state (regions + extracted text) before closing.
            try:
                pv = source_dialog.parent() if source_dialog is not None else None
                if pv is not None:
                    if hasattr(pv, "clear_note_regions"):
                        pv.clear_note_regions()
                    if hasattr(pv, "clear_extracted_notes_dialog"):
                        pv.clear_extracted_notes_dialog()
            except Exception:
                pass
            try:
                if source_dialog is not None and hasattr(source_dialog, "clear_content"):
                    source_dialog.clear_content()
            except Exception:
                pass
            try:
                if source_dialog is not None:
                    source_dialog.close()
            except Exception:
                pass
            _focus_form3()

        _loading_init()

        # If an insert is already in progress, don't interleave operations.
        try:
            if getattr(self, "_form3_insert_in_progress", False):
                QMessageBox.information(self, "Insert Busy", "A Form 3 insert is already in progress. Please wait for it to finish.")
                _loading_finish_and_close()
                return
        except Exception:
            pass

        def _try_set_value(row_1based: int, col_1based: int, value) -> bool:
            try:
                r = int(row_1based)
                c = int(col_1based)
            except Exception:
                return False

            try:
                cell = ws.cell(row=r, column=c)
            except Exception:
                return False

            # If the target is a merged cell, write to the top-left cell of that merged range.
            try:
                if cell is not None and cell.__class__.__name__ == "MergedCell":
                    try:
                        for rng in getattr(ws, "merged_cells", []).ranges:
                            if rng.min_row <= r <= rng.max_row and rng.min_col <= c <= rng.max_col:
                                cell = ws.cell(row=int(rng.min_row), column=int(rng.min_col))
                                break
                    except Exception:
                        pass
            except Exception:
                pass

            try:
                cell.value = value
                return True
            except Exception:
                return False

        if self._template_wb is None:
            try:
                print("Form3 delete: no template wb")
            except Exception:
                pass
            QMessageBox.warning(self, "No Template", "Load an FAI template first so Form 3 exists.")
            return

        ws3_name = self._form_sheet_names.get("3")
        if not ws3_name or ws3_name not in self._template_wb.sheetnames:
            try:
                print(f"Form3 delete: missing sheet name (ws3_name={ws3_name})")
            except Exception:
                pass
            QMessageBox.warning(self, "No Form 3", "Form 3 worksheet is not available in the loaded template.")
            return

        # Snapshot for Ctrl+Z undo.
        try:
            print("Form3 delete: calling _push_form3_undo_state")
        except Exception:
            pass
        self._push_form3_undo_state()
        try:
            print("Form3 delete: snapshot taken (single)")
        except Exception:
            pass

        try:
            ws = self._template_wb[ws3_name]
            try:
                print(f"Form3 delete: using sheet '{ws3_name}'")
            except Exception:
                pass
        except Exception as e:
            try:
                print(f"Form3 delete: failed to access sheet '{ws3_name}': {e}")
            except Exception:
                pass
            return

        # Fill used to mark Results column (L) red for inserted rows.
        try:
            _insert_red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        except Exception:
            _insert_red_fill = None

        def _try_set_fill(row_1based: int, col_1based: int, fill_obj) -> bool:
            if fill_obj is None:
                return False
            try:
                r = int(row_1based)
                c = int(col_1based)
            except Exception:
                return False
            try:
                cell = ws.cell(row=r, column=c)
            except Exception:
                return False
            try:
                if cell is not None and cell.__class__.__name__ == "MergedCell":
                    for rng in getattr(ws, "merged_cells", []).ranges:
                        if rng.min_row <= r <= rng.max_row and rng.min_col <= c <= rng.max_col:
                            cell = ws.cell(row=int(rng.min_row), column=int(rng.min_col))
                            break
            except Exception:
                pass
            try:
                cell.fill = fill_obj
                return True
            except Exception:
                return False

        def _resolve_merged_top_left(row_1based: int, col_1based: int) -> tuple[int, int]:
            """If (row,col) is a merged cell, return the merged range top-left."""
            try:
                r = int(row_1based)
                c = int(col_1based)
            except Exception:
                return int(row_1based), int(col_1based)
            try:
                cell = ws.cell(row=r, column=c)
                if cell is not None and cell.__class__.__name__ == "MergedCell":
                    for rng in getattr(ws, "merged_cells", []).ranges:
                        if rng.min_row <= r <= rng.max_row and rng.min_col <= c <= rng.max_col:
                            return int(rng.min_row), int(rng.min_col)
            except Exception:
                pass
            return int(r), int(c)

        def _viewer_set_cell(viewer_obj, row_1based: int, col_1based: int, value) -> None:
            """Fast-path update for the visible table without a full render()."""
            if viewer_obj is None:
                return
            tbl = getattr(viewer_obj, "table", None)
            if tbl is None:
                return
            try:
                r1, c1 = _resolve_merged_top_left(int(row_1based), int(col_1based))
                r0 = int(r1) - 1
                c0 = int(c1) - 1
                if r0 < 0 or c0 < 0:
                    return

                # Ensure the table is large enough.
                if r0 >= int(tbl.rowCount()):
                    tbl.setRowCount(r0 + 1)
                if c0 >= int(tbl.columnCount()):
                    tbl.setColumnCount(c0 + 1)

                from PySide6.QtWidgets import QTableWidgetItem

                item = tbl.item(r0, c0)
                if item is None:
                    item = QTableWidgetItem("")
                    tbl.setItem(r0, c0, item)
                item.setText("" if value is None else str(value))
            except Exception:
                return

        def _wrapped_height_px(viewer_obj, row_1based: int, col_1based: int, text: str) -> int | None:
            if viewer_obj is None:
                return None
            tbl = getattr(viewer_obj, "table", None)
            if tbl is None:
                return None
            try:
                r0 = int(row_1based) - 1
                c0 = int(col_1based) - 1
            except Exception:
                return None
            if r0 < 0 or c0 < 0:
                return None
            try:
                col_w = int(tbl.columnWidth(int(c0)))
            except Exception:
                col_w = 0
            if col_w <= 0:
                return None

            padding = 8
            avail = max(int(col_w) - padding, 20)
            try:
                it = tbl.item(int(r0), int(c0))
                font = it.font() if it is not None else tbl.font()
            except Exception:
                font = tbl.font()
            try:
                fm = QFontMetrics(font)
                rect = fm.boundingRect(QRect(0, 0, int(avail), 10000), Qt.TextWordWrap, str(text or ""))
                height = int(rect.height())
            except Exception:
                try:
                    height = int(fm.height())
                except Exception:
                    height = 0
            try:
                min_h = int(fm.height()) + 6
            except Exception:
                min_h = 15
            return max(int(height) + 6, int(min_h))

        # Discover the start row similar to _write_form3_to_worksheet.
        start_row = 6
        header_row = 5
        try:
            for r in range(1, 20):
                val = ws.cell(row=r, column=2).value
                if val and "char no." in str(val).lower():
                    header_row = r
                    start_row = r + 1
                    try:
                        row_text = " ".join(
                            [
                                str(ws.cell(row=start_row, column=c).value or "")
                                for c in range(1, min(ws.max_column, 30) + 1)
                            ]
                        ).lower()
                        if "op #" in row_text or "op#" in row_text or "reference location" in row_text or "bubble" in row_text:
                            start_row = r + 2
                    except Exception:
                        pass
                    break
        except Exception:
            start_row = 6

        # Fixed Form 3 mapping per your template:
        # - Char No. is column B
        # - Bubble number is column E
        # - Notes are column G
        char_col = 2
        bubble_col = 5
        notes_col = 7

        # Determine the last "real" row by scanning for a run of empties.
        try:
            max_scan = min(max(int(getattr(ws, "max_row", 0) or 0), start_row + 250), start_row + 2000)
            started = False
            empty_run = 0
            end_row = start_row - 1
            for rr in range(start_row, max_scan + 1):
                c2 = ws.cell(row=rr, column=2).value
                c7 = ws.cell(row=rr, column=7).value
                has_any = (c2 is not None and str(c2).strip() != "") or (c7 is not None and str(c7).strip() != "")
                if has_any:
                    started = True
                    empty_run = 0
                    end_row = rr
                else:
                    if started:
                        empty_run += 1
                        if empty_run >= 25:
                            break
        except Exception:
            end_row = start_row - 1

        target_row = int(end_row) + 1
        if target_row < int(start_row):
            target_row = int(start_row)

        # Determine next available Char No. / Bubble number.
        next_num = 1
        try:
            max_num = 0
            for rr in range(int(start_row), int(end_row) + 1):
                for cc in (int(char_col), int(bubble_col)):
                    try:
                        v = ws.cell(row=rr, column=cc).value
                        n = int(v)
                        if n > max_num:
                            max_num = n
                    except Exception:
                        continue
            next_num = int(max_num) + 1 if max_num > 0 else 1
        except Exception:
            next_num = 1

        # Insert step-by-step (so user can see it happen) rather than in a tight loop.
        viewer = self._form_viewers.get("3")

        state = {
            "ws": ws,
            "viewer": viewer,
            "notes": list(notes),
            "char_col": int(char_col),
            "bubble_col": int(bubble_col),
            "notes_col": int(notes_col),
            "row_cursor": int(target_row),
            "next_num": int(next_num),
            "i": 0,
            "_pre_shown": False,
            "_pre_row": None,
        }

        try:
            self._form3_insert_in_progress = True
        except Exception:
            pass

        # Fixed speed per latest UX: 10ms
        delay_ms = 10

        def _find_next_empty(start_row: int) -> int:
            rr2 = int(start_row)
            try:
                guard = 0
                while guard < 500:
                    guard += 1
                    b_val = ws.cell(row=int(rr2), column=int(char_col)).value
                    e_val = ws.cell(row=int(rr2), column=int(bubble_col)).value
                    g_val = ws.cell(row=int(rr2), column=int(notes_col)).value
                    has_any = (
                        (b_val is not None and str(b_val).strip() != "")
                        or (e_val is not None and str(e_val).strip() != "")
                        or (g_val is not None and str(g_val).strip() != "")
                    )
                    if not has_any:
                        break
                    rr2 += 1
            except Exception:
                pass
            return int(rr2)

        # If delay is 0ms (or negative), run a fast bulk insert.
        # (Not used in normal UX; retained as a safety fallback.)
        if int(delay_ms) <= 0:
            def _bulk() -> None:
                try:
                    rr_cursor = _find_next_empty(int(target_row))

                    # Preview/select insertion row once (4 rows of context above), then insert quickly.
                    try:
                        if viewer is not None:
                            tbl = getattr(viewer, "table", None)
                            if tbl is not None:
                                try:
                                    tbl.setFocus()
                                except Exception:
                                    pass
                                r0 = int(rr_cursor) - 1
                                c0 = int(notes_col) - 1
                                try:
                                    tbl.setCurrentCell(r0, c0)
                                except Exception:
                                    pass
                                try:
                                    tbl.selectRow(r0)
                                except Exception:
                                    pass
                                try:
                                    from PySide6.QtWidgets import QAbstractItemView

                                    anchor_r0 = max(0, r0 - 4)
                                    idx = tbl.model().index(anchor_r0, c0)
                                    if idx.isValid():
                                        tbl.scrollTo(idx, QAbstractItemView.ScrollHint.PositionAtTop)
                                except Exception:
                                    pass
                    except Exception:
                        pass

                    for i, seg in enumerate(list(notes)):
                        s = str(seg or "").strip()
                        if not s:
                            continue
                        n = int(next_num) + int(i)
                        rr = _find_next_empty(int(rr_cursor))

                        _try_set_value(rr, int(char_col), int(n))
                        _try_set_value(rr, int(bubble_col), int(n))
                        _try_set_value(rr, int(notes_col), s)
                        # Column L (12): mark Results cell red for inserted rows.
                        _try_set_fill(rr, 12, _insert_red_fill)

                        # Lightweight on-screen updates.
                        try:
                            if viewer is not None:
                                tbl = getattr(viewer, "table", None)
                                if tbl is not None:
                                    tbl.blockSignals(True)
                                _viewer_set_cell(viewer, rr, int(char_col), int(n))
                                _viewer_set_cell(viewer, rr, int(bubble_col), int(n))
                                _viewer_set_cell(viewer, rr, int(notes_col), s)
                        finally:
                            try:
                                if viewer is not None:
                                    tbl = getattr(viewer, "table", None)
                                    if tbl is not None:
                                        tbl.blockSignals(False)
                            except Exception:
                                pass

                        try:
                            aw = ws.cell(row=int(rr), column=int(notes_col))
                            if aw is not None and aw.__class__.__name__ == "MergedCell":
                                for rng in getattr(ws, "merged_cells", []).ranges:
                                    if rng.min_row <= rr <= rng.max_row and rng.min_col <= int(notes_col) <= rng.max_col:
                                        aw = ws.cell(row=int(rng.min_row), column=int(rng.min_col))
                                        break
                            aw.alignment = Alignment(wrapText=True, vertical="top")

                            # Auto-fit row height using actual column width/font (Bulk path).
                            desired_px = _wrapped_height_px(viewer, int(rr), int(notes_col), str(s or ""))
                            if desired_px is None:
                                desired_px = 22

                            # Persist row height to worksheet (points at base scale).
                            effective = 1.0
                            try:
                                if viewer is not None and hasattr(viewer, "effective_scale"):
                                    effective = float(viewer.effective_scale()) or 1.0
                            except Exception:
                                effective = 1.0
                            base_px = max(int(desired_px / effective), 1)
                            ws.row_dimensions[int(rr)].height = float(base_px * 72 / 96)

                            # Update viewer row height + base sizes.
                            try:
                                if viewer is not None and hasattr(viewer, "set_row_height_pixels"):
                                    viewer.set_row_height_pixels(int(rr), int(desired_px))
                            except Exception:
                                pass

                            # Ensure wrap role is set for the note cell.
                            try:
                                if viewer is not None:
                                    t_tbl = getattr(viewer, "table", None)
                                    if t_tbl is not None:
                                        it = t_tbl.item(int(rr) - 1, int(notes_col) - 1)
                                        if it is not None:
                                            it.setData(viewer.WRAP_ROLE, True)
                            except Exception:
                                pass
                        except Exception:
                            pass

                        try:
                            ecell = ws.cell(row=int(rr), column=int(bubble_col))
                            f = getattr(ecell, "font", None)
                            if f is not None and getattr(f, "color", None) is not None:
                                try:
                                    ecell.font = f.copy(color=None)
                                except Exception:
                                    pass
                        except Exception:
                            pass

                        rr_cursor = int(rr) + 1

                    try:
                        self._update_form3_bubble_fills(set(getattr(self, "_last_bubbled_numbers", set()) or set()))
                    except Exception:
                        pass
                    try:
                        if viewer is not None:
                            viewer.render()
                    except Exception:
                        pass
                    _loading_finish_and_close()
                except Exception as e:
                    QMessageBox.warning(self, "Insert Failed", f"Could not insert notes into Form 3:\n{e}")
                    _loading_finish_and_close()
                finally:
                    try:
                        self._form3_insert_in_progress = False
                    except Exception:
                        pass

            QTimer.singleShot(0, _bulk)
            return

        def _step() -> None:
            try:
                i = int(state.get("i", 0) or 0)
                notes_list = state.get("notes") or []
                if i >= len(notes_list):
                    # Finalize
                    try:
                        self._update_form3_bubble_fills(set(getattr(self, "_last_bubbled_numbers", set()) or set()))
                    except Exception:
                        pass
                    try:
                        if viewer is not None:
                            viewer.render()
                    except Exception:
                        pass
                    try:
                        self._form3_insert_in_progress = False
                    except Exception:
                        pass
                    _loading_finish_and_close()
                    return

                seg = str(notes_list[i] or "").strip()
                n = int(state.get("next_num", 1) or 1) + i
                rr = int(state.get("row_cursor", 1) or 1)

                # First step: show the row we are about to insert into BEFORE writing anything.
                # Keep the view anchored afterward (no scrolling on subsequent rows).
                if i == 0 and not bool(state.get("_pre_shown", False)):
                    rr_pre = _find_next_empty(rr)
                    state["_pre_shown"] = True
                    state["_pre_row"] = int(rr_pre)

                    # Render + scroll/select the target row to preview insertion location.
                    try:
                        if viewer is not None:
                            tbl = getattr(viewer, "table", None)
                            if tbl is not None:
                                try:
                                    tbl.setFocus()
                                except Exception:
                                    pass
                                r0 = int(rr_pre) - 1
                                c0 = int(notes_col) - 1
                                try:
                                    # Ensure the table is large enough for selection.
                                    if r0 >= int(tbl.rowCount()):
                                        tbl.setRowCount(r0 + 1)
                                    if c0 >= int(tbl.columnCount()):
                                        tbl.setColumnCount(c0 + 1)
                                except Exception:
                                    pass
                                try:
                                    tbl.setCurrentCell(r0, c0)
                                except Exception:
                                    pass
                                try:
                                    tbl.selectRow(r0)
                                except Exception:
                                    pass
                                try:
                                    from PySide6.QtWidgets import QAbstractItemView

                                    # Show a little context above the insertion row.
                                    anchor_r0 = max(0, r0 - 4)
                                    idx = tbl.model().index(anchor_r0, c0)
                                    if idx.isValid():
                                        tbl.scrollTo(idx, QAbstractItemView.ScrollHint.PositionAtTop)
                                except Exception:
                                    pass
                    except Exception:
                        pass

                    QTimer.singleShot(int(delay_ms), _step)
                    return

                # Use the precomputed row for the first insert so it doesn't change.
                if i == 0 and state.get("_pre_row"):
                    rr = int(state.get("_pre_row"))
                else:
                    rr = _find_next_empty(rr)

                # Write B/E/G
                _try_set_value(rr, int(char_col), int(n))
                _try_set_value(rr, int(bubble_col), int(n))
                _try_set_value(rr, int(notes_col), seg)
                # Column L (12): mark Results cell red for inserted rows.
                _try_set_fill(rr, 12, _insert_red_fill)

                # Lightweight on-screen updates (avoid full render on every row).
                try:
                    if viewer is not None:
                        tbl = getattr(viewer, "table", None)
                        if tbl is not None:
                            tbl.blockSignals(True)
                        _viewer_set_cell(viewer, rr, int(char_col), int(n))
                        _viewer_set_cell(viewer, rr, int(bubble_col), int(n))
                        _viewer_set_cell(viewer, rr, int(notes_col), seg)
                finally:
                    try:
                        if viewer is not None:
                            tbl = getattr(viewer, "table", None)
                            if tbl is not None:
                                tbl.blockSignals(False)
                    except Exception:
                        pass

                # Wrap notes cell
                try:
                    aw = ws.cell(row=int(rr), column=int(notes_col))
                    if aw is not None and aw.__class__.__name__ == "MergedCell":
                        for rng in getattr(ws, "merged_cells", []).ranges:
                            if rng.min_row <= rr <= rng.max_row and rng.min_col <= int(notes_col) <= rng.max_col:
                                aw = ws.cell(row=int(rng.min_row), column=int(rng.min_col))
                                break
                    aw.alignment = Alignment(wrapText=True, vertical="top")

                    # Auto-fit row height using actual column width/font
                    desired_px = _wrapped_height_px(viewer, int(rr), int(notes_col), str(seg or ""))
                    if desired_px is None:
                        desired_px = 22

                    # Persist row height to worksheet (points at base scale).
                    effective = 1.0
                    try:
                        if viewer is not None and hasattr(viewer, "effective_scale"):
                            effective = float(viewer.effective_scale()) or 1.0
                    except Exception:
                        effective = 1.0
                    base_px = max(int(desired_px / effective), 1)
                    ws.row_dimensions[int(rr)].height = float(base_px * 72 / 96)

                    # Update viewer row height + base sizes.
                    try:
                        if viewer is not None and hasattr(viewer, "set_row_height_pixels"):
                            viewer.set_row_height_pixels(int(rr), int(desired_px))
                    except Exception:
                        pass

                    # Ensure wrap role is set for the note cell.
                    try:
                        if viewer is not None:
                            t_tbl = getattr(viewer, "table", None)
                            if t_tbl is not None:
                                it = t_tbl.item(int(rr) - 1, int(notes_col) - 1)
                                if it is not None:
                                    it.setData(viewer.WRAP_ROLE, True)
                    except Exception:
                        pass
                except Exception:
                    pass

                # Ensure bubble text stays visible
                try:
                    ecell = ws.cell(row=int(rr), column=int(bubble_col))
                    f = getattr(ecell, "font", None)
                    if f is not None and getattr(f, "color", None) is not None:
                        try:
                            ecell.font = f.copy(color=None)
                        except Exception:
                            pass
                except Exception:
                    pass

                # Select the row being inserted.
                # Keep the view stable (no auto-scroll) after the initial preview.
                try:
                    if viewer is not None:
                        tbl = getattr(viewer, "table", None)
                        if tbl is not None:
                            try:
                                tbl.setFocus()
                            except Exception:
                                pass
                            try:
                                r0 = int(rr) - 1
                                c0 = int(notes_col) - 1

                                # Ensure the table is large enough for selection.
                                try:
                                    if r0 >= int(tbl.rowCount()):
                                        tbl.setRowCount(r0 + 1)
                                    if c0 >= int(tbl.columnCount()):
                                        tbl.setColumnCount(c0 + 1)
                                except Exception:
                                    pass

                                # Select the inserted row (no scrolling here).
                                try:
                                    tbl.setCurrentCell(r0, c0)
                                except Exception:
                                    pass
                                try:
                                    tbl.selectRow(r0)
                                except Exception:
                                    pass
                            except Exception:
                                pass
                except Exception:
                    pass

                # Next
                state["i"] = i + 1
                state["row_cursor"] = int(rr) + 1
            except Exception as e:
                try:
                    self._form3_insert_in_progress = False
                except Exception:
                    pass
                QMessageBox.warning(self, "Insert Failed", f"Could not insert notes into Form 3:\n{e}")
                _loading_finish_and_close()
                return

            QTimer.singleShot(int(delay_ms), _step)

        QTimer.singleShot(0, _step)

        # (Insertion now completes asynchronously in _step())
        try:
            if hasattr(self, "drawing_pdf_edit") and self.drawing_pdf_edit is not None:
                self.drawing_pdf_edit.setText(p)
        except Exception:
            pass

    def _on_form3_row_insert_requested(self, row_1based: int, where: str) -> None:
        """Insert a row into Form 3 (below row 5), then renumber and sync bubbles."""
        try:
            row_1based = int(row_1based)
        except Exception:
            return
        where = str(where or "").strip().lower()
        if where not in ("above", "below"):
            where = "below"

        if row_1based <= 5:
            return

        if self._template_wb is None:
            try:
                print("Form3 delete (multi): no template wb")
            except Exception:
                pass
            QMessageBox.warning(self, "No Template", "Load an FAI template first so Form 3 exists.")
            return

        ws3_name = self._form_sheet_names.get("3")
        if not ws3_name or ws3_name not in self._template_wb.sheetnames:
            try:
                print(f"Form3 delete (multi): missing sheet name (ws3_name={ws3_name})")
            except Exception:
                pass
            QMessageBox.warning(self, "No Form 3", "Form 3 worksheet is not available in the loaded template.")
            return

        # Snapshot for Ctrl+Z undo.
        try:
            print("Form3 delete (multi): calling _push_form3_undo_state")
        except Exception:
            pass
        self._push_form3_undo_state()
        try:
            print("Form3 delete: snapshot taken (multi)")
        except Exception:
            pass

        try:
            ws = self._template_wb[ws3_name]
            try:
                print(f"Form3 delete (multi): using sheet '{ws3_name}'")
            except Exception:
                pass
        except Exception as e:
            try:
                print(f"Form3 delete (multi): failed to access sheet '{ws3_name}': {e}")
            except Exception:
                pass
            return
        insert_at = int(row_1based) if where == "above" else int(row_1based) + 1
        if insert_at <= 5:
            insert_at = 6

        # Insert row and best-effort copy formatting from an adjacent row.
        try:
            ws.insert_rows(insert_at, 1)
        except Exception as e:
            QMessageBox.warning(self, "Insert Failed", f"Could not insert row into Form 3:\n{e}")
            return

        try:
            if where == "above":
                src_row = insert_at + 1
            else:
                src_row = max(1, insert_at - 1)
            max_col = int(getattr(ws, "max_column", 0) or 0)
            if max_col <= 0:
                max_col = 50
            for cc in range(1, max_col + 1):
                try:
                    src = ws.cell(row=int(src_row), column=int(cc))
                    dst = ws.cell(row=int(insert_at), column=int(cc))
                    # Copy style-related attributes (best effort) but DO NOT copy fill,
                    # since Form 3 uses dynamic red/green fills and copying them causes
                    # confusing "red bleed" on inserted rows.
                    for attr in ("font", "border", "alignment", "number_format", "protection"):
                        try:
                            v = getattr(src, attr, None)
                            if v is not None:
                                setattr(dst, attr, copy.copy(v))
                        except Exception:
                            continue
                except Exception:
                    continue
            try:
                ws.row_dimensions[int(insert_at)].height = ws.row_dimensions[int(src_row)].height
            except Exception:
                pass
        except Exception:
            pass

        # Ensure inserted row starts clean: no Reference Location and no copied fills.
        try:
            ws.cell(row=int(insert_at), column=4).value = ""
        except Exception:
            pass
        try:
            # Clear fills on Char No (B) and Notes (G) if they were copied/merged.
            ws.cell(row=int(insert_at), column=2).fill = PatternFill()
        except Exception:
            pass
        try:
            ws.cell(row=int(insert_at), column=7).fill = PatternFill()
        except Exception:
            pass

        # Refresh Form 3 shading and ensure Reference Location stays blank for missing bubbles.
        try:
            mapping = self._renumber_form3_char_and_bubble(ws)
        except Exception:
            mapping = {}

        # If the inserted row's bubble number falls inside an existing drawing range
        # bubble (e.g. 14-28), it should be treated as "missing" until the user adds
        # its own bubble. Exclude it from any range bubbles.
        inserted_bubble_num = None
        try:
            v = ws.cell(row=int(insert_at), column=5).value
            inserted_bubble_num = int(v)
        except Exception:
            inserted_bubble_num = None

        try:
            self._apply_bubble_number_mapping_to_drawing(mapping)
        except Exception:
            pass

        try:
            if inserted_bubble_num is not None and inserted_bubble_num > 0:
                dv = getattr(self, "drawing_viewer_tab", None)
                pv = getattr(dv, "_pdf_viewer", None) if dv is not None else None
                if pv is not None and hasattr(pv, "exclude_numbers_from_ranges"):
                    pv.exclude_numbers_from_ranges({int(inserted_bubble_num)})
        except Exception:
            pass

        try:
            dv = getattr(self, "drawing_viewer_tab", None)
            pv = getattr(dv, "_pdf_viewer", None) if dv is not None else None
            if pv is not None and hasattr(pv, "set_pending_bubble_number_to_lowest_available"):
                pv.set_pending_bubble_number_to_lowest_available()
        except Exception:
            pass

        try:
            QTimer.singleShot(0, self._sync_bubbles_to_form3)
        except Exception:
            pass

        # Refresh view.
        try:
            v3 = self._form_viewers.get("3")
            if v3 is not None:
                v3.render()
        except Exception:
            pass

        # Ensure Form 3 has focus after delete so Ctrl+Z works.
        try:
            tabs = getattr(self, "forms_tabs", None)
            key_map = getattr(self, "_form_tab_to_key", {}) or {}
            if tabs is not None:
                for i in range(int(tabs.count())):
                    try:
                        w = tabs.widget(int(i))
                    except Exception:
                        w = None
                    if w is not None and str(key_map.get(w, "")) == "3":
                        tabs.setCurrentIndex(int(i))
                        break
        except Exception:
            pass
        try:
            v3 = self._form_viewers.get("3")
            tbl = getattr(v3, "table", None) if v3 is not None else None
            if tbl is not None:
                tbl.setFocus()
        except Exception:
            pass

        
    def _on_form3_row_delete_requested(self, row_1based: int) -> None:
        """Delete a row from Form 3 (below row 5), then renumber and sync bubbles."""
        try:
            print(f"Form3 delete requested: row={row_1based}")
        except Exception:
            pass
        try:
            row_1based = int(row_1based)
        except Exception:
            return
        if row_1based <= 5:
            return

        if self._template_wb is None:
            QMessageBox.warning(self, "No Template", "Load an FAI template first so Form 3 exists.")
            return

        ws3_name = self._form_sheet_names.get("3")
        if not ws3_name or ws3_name not in self._template_wb.sheetnames:
            QMessageBox.warning(self, "No Form 3", "Form 3 worksheet is not available in the loaded template.")
            return

        # Snapshot for Ctrl+Z undo.
        try:
            print("Form3 delete: calling _push_form3_undo_state")
        except Exception:
            pass
        self._push_form3_undo_state()
        try:
            print("Form3 delete: snapshot taken (single)")
        except Exception:
            pass

        ws = self._template_wb[ws3_name]
        try:
            print(f"Form3 delete: using sheet '{ws3_name}'")
        except Exception:
            pass

        # Best-effort: capture the bubble number on the row being deleted so we can
        # remove the corresponding single bubble from the drawing.
        deleted_bubble_num = None
        try:
            cell = ws.cell(row=int(row_1based), column=5)
            # Handle merged cells (rare but possible).
            if cell is not None and cell.__class__.__name__ == "MergedCell":
                for rng in getattr(ws, "merged_cells", []).ranges:
                    if rng.min_row <= int(row_1based) <= rng.max_row and rng.min_col <= 5 <= rng.max_col:
                        cell = ws.cell(row=int(rng.min_row), column=int(rng.min_col))
                        break
            v = getattr(cell, "value", None)
            deleted_bubble_num = int(v)
            if deleted_bubble_num <= 0:
                deleted_bubble_num = None
        except Exception:
            deleted_bubble_num = None

        try:
            ws.delete_rows(int(row_1based), 1)
        except Exception as e:
            QMessageBox.warning(self, "Delete Failed", f"Could not delete row from Form 3:\n{e}")
            return

        # Renumber and sync bubbles.
        try:
            mapping = self._renumber_form3_char_and_bubble(ws)
        except Exception:
            mapping = {}

        # Remove the deleted bubble number (single-bubble items only), then remap the rest.
        try:
            dv = getattr(self, "drawing_viewer_tab", None)
            pv = getattr(dv, "_pdf_viewer", None) if dv is not None else None
            if pv is not None and deleted_bubble_num is not None and hasattr(pv, "delete_bubbles_with_numbers"):
                pv.delete_bubbles_with_numbers({int(deleted_bubble_num)})
        except Exception:
            pass

        try:
            self._apply_bubble_number_mapping_to_drawing(mapping)
        except Exception:
            pass

        try:
            dv = getattr(self, "drawing_viewer_tab", None)
            pv = getattr(dv, "_pdf_viewer", None) if dv is not None else None
            if pv is not None and hasattr(pv, "set_pending_bubble_number_to_lowest_available"):
                pv.set_pending_bubble_number_to_lowest_available()
        except Exception:
            pass

        try:
            QTimer.singleShot(0, self._sync_bubbles_to_form3)
        except Exception:
            pass

        try:
            v3 = self._form_viewers.get("3")
            if v3 is not None:
                v3.render()
        except Exception:
            pass


    def _on_form3_rows_delete_requested(self, rows_1based) -> None:
        """Delete multiple rows from Form 3 (below row 5), then renumber and sync bubbles."""
        try:
            print(f"Form3 delete requested (multi): rows={rows_1based}")
        except Exception:
            pass
        try:
            rows = [int(r) for r in (rows_1based or [])]
        except Exception:
            return

        rows = sorted({int(r) for r in rows if int(r) > 5}, reverse=True)
        if len(rows) <= 1:
            # Defer to the single-row path if we somehow got a single row.
            try:
                if rows:
                    self._on_form3_row_delete_requested(int(rows[0]))
            except Exception:
                pass
            return

        if self._template_wb is None:
            QMessageBox.warning(self, "No Template", "Load an FAI template first so Form 3 exists.")
            return

        ws3_name = self._form_sheet_names.get("3")
        if not ws3_name or ws3_name not in self._template_wb.sheetnames:
            QMessageBox.warning(self, "No Form 3", "Form 3 worksheet is not available in the loaded template.")
            return

        # Snapshot for Ctrl+Z undo.
        try:
            print("Form3 delete (multi): calling _push_form3_undo_state")
        except Exception:
            pass
        self._push_form3_undo_state()
        try:
            print("Form3 delete: snapshot taken (multi)")
        except Exception:
            pass

        ws = self._template_wb[ws3_name]

        def _bubble_num_at_row(row_1based: int) -> int | None:
            try:
                cell = ws.cell(row=int(row_1based), column=5)
                if cell is not None and cell.__class__.__name__ == "MergedCell":
                    for rng in getattr(ws, "merged_cells", []).ranges:
                        if rng.min_row <= int(row_1based) <= rng.max_row and rng.min_col <= 5 <= rng.max_col:
                            cell = ws.cell(row=int(rng.min_row), column=int(rng.min_col))
                            break
                v = getattr(cell, "value", None)
                n = int(v)
                return int(n) if n > 0 else None
            except Exception:
                return None

        # Capture bubble numbers before deletion so we can delete those exact bubbles
        # before applying any renumbering/mapping.
        deleted_nums: set[int] = set()
        for rr in rows:
            n = _bubble_num_at_row(int(rr))
            if n is not None:
                deleted_nums.add(int(n))

        try:
            for rr in rows:
                ws.delete_rows(int(rr), 1)
        except Exception as e:
            QMessageBox.warning(self, "Delete Failed", f"Could not delete selected rows from Form 3:\n{e}")
            return

        try:
            mapping = self._renumber_form3_char_and_bubble(ws)
        except Exception:
            mapping = {}

        # Remove deleted bubbles (single-bubble items only), then remap remaining.
        try:
            dv = getattr(self, "drawing_viewer_tab", None)
            pv = getattr(dv, "_pdf_viewer", None) if dv is not None else None
            if pv is not None and deleted_nums and hasattr(pv, "delete_bubbles_with_numbers"):
                pv.delete_bubbles_with_numbers(set(int(x) for x in deleted_nums))
        except Exception:
            pass

        try:
            self._apply_bubble_number_mapping_to_drawing(mapping)
        except Exception:
            pass

        try:
            dv = getattr(self, "drawing_viewer_tab", None)
            pv = getattr(dv, "_pdf_viewer", None) if dv is not None else None
            if pv is not None and hasattr(pv, "set_pending_bubble_number_to_lowest_available"):
                pv.set_pending_bubble_number_to_lowest_available()
        except Exception:
            pass

        try:
            QTimer.singleShot(0, self._sync_bubbles_to_form3)
        except Exception:
            pass

        try:
            v3 = self._form_viewers.get("3")
            if v3 is not None:
                v3.render()
        except Exception:
            pass


    def _renumber_form3_char_and_bubble(self, ws) -> dict[int, int]:
        """Renumber Form 3 Char No (col B) and Bubble No (col E) sequentially.

        Returns a mapping of old bubble numbers to new bubble numbers.
        """
        char_col = 2
        bubble_col = 5
        notes_col = 7

        # Find start row after the 'Char No.' header.
        start_row = 6
        try:
            for r in range(1, 20):
                v = ws.cell(row=r, column=2).value
                if v and "char no" in str(v).lower():
                    start_row = int(r) + 1
                    break
        except Exception:
            start_row = 6

        # Find last row by scanning for a run of empties.
        end_row = start_row
        try:
            max_scan = min(max(int(getattr(ws, "max_row", 0) or 0), start_row + 250), start_row + 5000)
            started = False
            empty_run = 0
            end_row = start_row - 1
            for rr in range(int(start_row), int(max_scan) + 1):
                b = ws.cell(row=rr, column=int(char_col)).value
                e = ws.cell(row=rr, column=int(bubble_col)).value
                g = ws.cell(row=rr, column=int(notes_col)).value
                has_any = (b is not None and str(b).strip() != "") or (e is not None and str(e).strip() != "") or (g is not None and str(g).strip() != "")
                if has_any:
                    started = True
                    empty_run = 0
                    end_row = rr
                else:
                    if started:
                        empty_run += 1
                        if empty_run >= 25:
                            break
        except Exception:
            end_row = start_row - 1

        if end_row < start_row:
            return {}

        def _set_value_merged_safe(row_1based: int, col_1based: int, value) -> None:
            try:
                r = int(row_1based)
                c = int(col_1based)
            except Exception:
                return
            try:
                cell = ws.cell(row=r, column=c)
            except Exception:
                return
            try:
                if cell is not None and cell.__class__.__name__ == "MergedCell":
                    for rng in getattr(ws, "merged_cells", []).ranges:
                        if rng.min_row <= r <= rng.max_row and rng.min_col <= c <= rng.max_col:
                            cell = ws.cell(row=int(rng.min_row), column=int(rng.min_col))
                            break
            except Exception:
                pass
            try:
                cell.value = value
            except Exception:
                pass

        mapping: dict[int, int] = {}
        new_num = 1
        for rr in range(int(start_row), int(end_row) + 1):
            old = None
            try:
                old = int(ws.cell(row=rr, column=int(bubble_col)).value)
            except Exception:
                old = None
            _set_value_merged_safe(rr, int(char_col), int(new_num))
            _set_value_merged_safe(rr, int(bubble_col), int(new_num))
            if old is not None and old > 0 and int(old) != int(new_num) and int(old) not in mapping:
                mapping[int(old)] = int(new_num)
            new_num += 1

        return mapping

    def _renumber_form3_char_and_bubble_by_description(self, ws) -> dict[int, int]:
        """Renumber Form 3 Char No/Bubble No only for rows with Description/Note text.

        Rows without Description/Note text (col G) are cleared for Char/Bubble.
        Returns a mapping of old bubble numbers to new bubble numbers.
        """
        char_col = 2
        bubble_col = 5
        notes_col = 7

        # Find start row after the 'Char No.' header.
        start_row = 6
        try:
            for r in range(1, 20):
                v = ws.cell(row=r, column=2).value
                if v and "char no" in str(v).lower():
                    start_row = int(r) + 1
                    break
        except Exception:
            start_row = 6

        # Find last row by scanning for a run of empties.
        end_row = start_row
        try:
            max_scan = min(max(int(getattr(ws, "max_row", 0) or 0), start_row + 250), start_row + 5000)
            started = False
            empty_run = 0
            end_row = start_row - 1
            for rr in range(int(start_row), int(max_scan) + 1):
                b = ws.cell(row=rr, column=int(char_col)).value
                e = ws.cell(row=rr, column=int(bubble_col)).value
                g = ws.cell(row=rr, column=int(notes_col)).value
                has_any = (b is not None and str(b).strip() != "") or (e is not None and str(e).strip() != "") or (g is not None and str(g).strip() != "")
                if has_any:
                    started = True
                    empty_run = 0
                    end_row = rr
                else:
                    if started:
                        empty_run += 1
                        if empty_run >= 25:
                            break
        except Exception:
            end_row = start_row - 1

        if end_row < start_row:
            return {}

        def _set_value_merged_safe(row_1based: int, col_1based: int, value) -> None:
            try:
                r = int(row_1based)
                c = int(col_1based)
            except Exception:
                return
            try:
                cell = ws.cell(row=r, column=c)
            except Exception:
                return
            try:
                if cell is not None and cell.__class__.__name__ == "MergedCell":
                    for rng in getattr(ws, "merged_cells", []).ranges:
                        if rng.min_row <= r <= rng.max_row and rng.min_col <= c <= rng.max_col:
                            cell = ws.cell(row=int(rng.min_row), column=int(rng.min_col))
                            break
            except Exception:
                pass
            try:
                cell.value = value
            except Exception:
                pass

        mapping: dict[int, int] = {}
        new_num = 1
        for rr in range(int(start_row), int(end_row) + 1):
            try:
                desc_val = ws.cell(row=rr, column=int(notes_col)).value
            except Exception:
                desc_val = None
            has_desc = bool(desc_val is not None and str(desc_val).strip() != "")

            if not has_desc:
                _set_value_merged_safe(rr, int(char_col), None)
                _set_value_merged_safe(rr, int(bubble_col), None)
                continue

            old = None
            try:
                old = int(ws.cell(row=rr, column=int(bubble_col)).value)
            except Exception:
                old = None
            _set_value_merged_safe(rr, int(char_col), int(new_num))
            _set_value_merged_safe(rr, int(bubble_col), int(new_num))
            if old is not None and old > 0 and int(old) != int(new_num) and int(old) not in mapping:
                mapping[int(old)] = int(new_num)
            new_num += 1

        return mapping

    def _push_form3_undo_state(self) -> None:
        """Snapshot the current workbook for Form 3 undo (Ctrl+Z)."""
        if self._template_wb is None:
            try:
                print("Form3 undo snapshot: no template wb")
            except Exception:
                pass
            return
        ws3_name = self._form_sheet_names.get("3")
        if not ws3_name or ws3_name not in self._template_wb.sheetnames:
            try:
                print(f"Form3 undo snapshot: missing sheet name (ws3_name={ws3_name})")
            except Exception:
                pass
            return
        try:
            buff = io.BytesIO()
            self._template_wb.save(buff)
            data = buff.getvalue()
            if data:
                self._form3_undo_stack.append(("bytes", data))
                if len(self._form3_undo_stack) > int(self._form3_undo_max):
                    self._form3_undo_stack = self._form3_undo_stack[-int(self._form3_undo_max):]
                try:
                    print(f"Form3 undo snapshot saved (bytes, stack={len(self._form3_undo_stack)})")
                except Exception:
                    pass
                return
        except Exception as e:
            try:
                print(f"Form3 undo snapshot bytes failed: {e}")
            except Exception:
                pass

        try:
            wb_copy = copy.deepcopy(self._template_wb)
            self._form3_undo_stack.append(("wb", wb_copy))
            if len(self._form3_undo_stack) > int(self._form3_undo_max):
                self._form3_undo_stack = self._form3_undo_stack[-int(self._form3_undo_max):]
            try:
                print(f"Form3 undo snapshot saved (copy, stack={len(self._form3_undo_stack)})")
            except Exception:
                pass
        except Exception as e:
            try:
                print(f"Form3 undo snapshot failed: {e}")
            except Exception:
                pass

    def _on_form3_undo_requested(self) -> bool:
        """Undo last Form 3 row delete by restoring the prior workbook snapshot."""
        try:
            try:
                print("Form3 undo requested")
            except Exception:
                pass
            logger.debug("Form3 undo requested: stack_size=%s", len(self._form3_undo_stack or []))
        except Exception:
            pass
        if not self._form3_undo_stack:
            try:
                try:
                    print("Form3 undo requested: no snapshot available")
                except Exception:
                    pass
                QMessageBox.information(self, "Undo", "No Form 3 delete undo is available.")
            except Exception:
                pass
            return False
        try:
            item = self._form3_undo_stack.pop()
        except Exception:
            return False

        wb = None
        try:
            kind, payload = item
        except Exception:
            kind, payload = ("bytes", item)

        if kind == "wb":
            wb = payload
        else:
            try:
                buff = io.BytesIO(payload)
                wb = openpyxl.load_workbook(buff)
            except Exception:
                return False

        if wb is None:
            return False

        self._template_wb = wb
        try:
            self._detect_form_sheets(self._template_wb)
        except Exception:
            pass

        # Re-bind worksheets to viewers and re-render.
        try:
            for form_key in ("1", "2", "2c", "3"):
                viewer = self._form_viewers.get(form_key)
                sheet_name = self._form_sheet_names.get(form_key)
                if viewer is None or not sheet_name:
                    continue
                if sheet_name not in self._template_wb.sheetnames:
                    continue
                ws = self._template_wb[sheet_name]
                viewer.set_worksheet(ws)
                viewer.set_overrides({})
                viewer.render()
        except Exception:
            pass

        try:
            self._set_wb_dirty()
        except Exception:
            pass

        try:
            QTimer.singleShot(0, self._sync_bubbles_to_form3)
        except Exception:
            pass

        try:
            try:
                print("Form3 undo applied")
            except Exception:
                pass
            logger.debug("Form3 undo applied")
        except Exception:
            pass

        return True

    def _on_form3_renumber_char_bubble_requested(self) -> None:
        """Renumber Form 3 Char/Bubble based on Description/Note text."""
        if self._template_wb is None:
            QMessageBox.warning(self, "No Template", "Load an FAI template first so Form 3 exists.")
            return

        ws3_name = self._form_sheet_names.get("3")
        if not ws3_name or ws3_name not in self._template_wb.sheetnames:
            QMessageBox.warning(self, "No Form 3", "Form 3 worksheet is not available in the loaded template.")
            return

        ws = self._template_wb[ws3_name]

        try:
            mapping = self._renumber_form3_char_and_bubble_by_description(ws)
        except Exception:
            mapping = {}

        try:
            self._apply_bubble_number_mapping_to_drawing(mapping)
        except Exception:
            pass

        try:
            if hasattr(self, "_set_wb_dirty"):
                self._set_wb_dirty()
        except Exception:
            pass

        try:
            QTimer.singleShot(0, self._sync_bubbles_to_form3)
        except Exception:
            pass

        try:
            v3 = self._form_viewers.get("3")
            if v3 is not None:
                v3.render()
        except Exception:
            pass

    def _on_fit_rows_requested(self, form_key: str) -> None:
        """Auto-fit rows using wrapped text in columns A–T for a form."""
        applied = False
        key = str(form_key or "").strip()

        def _apply_to_viewer(viewer) -> bool:
            if viewer is None:
                return False
            try:
                if hasattr(viewer, "fit_rows_to_wrapped_text"):
                    viewer.fit_rows_to_wrapped_text(1, 20)
                    return True
            except Exception:
                return False
            return False

        try:
            applied = bool(_apply_to_viewer(self._form_viewers.get(key))) or applied
        except Exception:
            pass

        # Also apply to any pop-out viewers matching the form key.
        try:
            for win in list(getattr(self, "_popout_windows", []) or []):
                try:
                    viewer = win.findChild(ExcelSheetViewer)
                except Exception:
                    viewer = None
                if viewer is None:
                    continue
                try:
                    if str(getattr(viewer, "form_key", "")) != key:
                        continue
                except Exception:
                    continue
                applied = bool(_apply_to_viewer(viewer)) or applied
        except Exception:
            pass

        if not applied:
            try:
                QMessageBox.information(self, "Fit Rows", f"Form {key} is not available yet.")
            except Exception:
                pass

    def _apply_bubble_number_mapping_to_drawing(self, mapping: dict[int, int]) -> None:
        if not mapping:
            return
        try:
            dvw = getattr(self, "drawing_viewer_tab", None)
            pv = getattr(dvw, "_pdf_viewer", None) if dvw is not None else None
            if pv is None:
                return
            if hasattr(pv, "apply_bubble_number_mapping"):
                pv.apply_bubble_number_mapping(mapping)
        except Exception:
            return

    def _focus_drawing_and_select_bubble(self, bubble_number: int) -> None:
        """Focus the Drawing Viewer and select the given bubble number if present."""
        try:
            n = int(bubble_number)
        except Exception:
            return
        if n <= 0:
            return

        dv = getattr(self, "drawing_viewer_tab", None)
        pv = getattr(dv, "_pdf_viewer", None) if dv is not None else None
        if dv is None or pv is None:
            return

        # If the drawing viewer is docked, switch tabs to it.
        try:
            tab = getattr(self, "_drawing_tab_widget", None)
            tabs = getattr(self, "forms_tabs", None)
            if tab is not None and tabs is not None:
                idx = tabs.indexOf(tab)
                if idx != -1:
                    tabs.setCurrentIndex(idx)
        except Exception:
            pass

        # If it's popped out, bring it forward.
        try:
            if dv is not None and dv.windowFlags() & Qt.Window:
                dv.show()
                dv.raise_()
                dv.activateWindow()
        except Exception:
            pass

        try:
            if hasattr(pv, "select_bubble_number"):
                pv.select_bubble_number(int(n), center=True)
        except Exception:
            pass

    def _on_drawing_bubble_scroller_selected(self, start: int, end: int) -> None:
        """Highlight Form 3 bubble cell(s) when the Drawing Viewer scroller moves.

        Requirement: only do this when the Drawing Viewer is popped out.
        """
        try:
            dv = getattr(self, "drawing_viewer_tab", None)
            popped_out = bool(dv is not None and (dv.windowFlags() & Qt.Window))
        except Exception:
            popped_out = False
        if not popped_out:
            return

        # Bring the main window to Form 3 so the user can see the highlight.
        try:
            self.show()
            self.raise_()
            self.activateWindow()
        except Exception:
            pass
        try:
            tabs = getattr(self, "forms_tabs", None)
            key_map = getattr(self, "_form_tab_to_key", {}) or {}
            if tabs is not None:
                for i in range(int(tabs.count())):
                    try:
                        w = tabs.widget(int(i))
                    except Exception:
                        w = None
                    if w is not None and str(key_map.get(w, "")) == "3":
                        tabs.setCurrentIndex(int(i))
                        break
        except Exception:
            pass

        try:
            s = int(start)
            e = int(end)
        except Exception:
            return
        if s <= 0:
            return
        if e < s:
            e = s

        self._highlight_form3_bubble_range(int(s), int(e))

    def _highlight_form3_bubble_range(self, start: int, end: int) -> None:
        """Select bubble-number cell(s) on Form 3 for [start,end] and scroll with an offset."""
        try:
            viewer = self._form_viewers.get("3") if hasattr(self, "_form_viewers") else None
        except Exception:
            viewer = None
        if viewer is None:
            return
        tbl = getattr(viewer, "table", None)
        ws = getattr(viewer, "_ws", None)
        if tbl is None or ws is None:
            return

        try:
            s = int(start)
            e = int(end)
        except Exception:
            return
        if s <= 0:
            return
        if e < s:
            e = s

        try:
            tbl.clearSelection()
        except Exception:
            pass

        matched_rows0: list[int] = []
        try:
            max_row = int(getattr(ws, "max_row", 0) or 0)
        except Exception:
            max_row = 0
        if max_row <= 0:
            max_row = 5000

        for rr in range(6, int(max_row) + 1):
            try:
                v = ws.cell(row=int(rr), column=5).value
                n = int(v)
            except Exception:
                continue
            if int(s) <= int(n) <= int(e):
                r0 = int(rr) - 1
                matched_rows0.append(int(r0))

        if not matched_rows0:
            return

        min_r0 = int(min(matched_rows0))
        max_r0 = int(max(matched_rows0))

        # IMPORTANT: setCurrentCell() can collapse selection to a single cell.
        # Set the current cell *first*, then apply the selection range.
        try:
            tbl.setCurrentCell(int(min_r0), 4)
        except Exception:
            pass

        # Select the whole range in the Bubble column so it highlights as one block.
        try:
            tbl.setRangeSelected(QTableWidgetSelectionRange(int(min_r0), 4, int(max_r0), 4), True)
        except Exception:
            # Fallback: select each row individually.
            for r0 in matched_rows0:
                try:
                    tbl.setRangeSelected(QTableWidgetSelectionRange(int(r0), 4, int(r0), 4), True)
                except Exception:
                    pass

        # Scroll so the selected bubble row is ~5 rows below the top.
        try:
            target_row0 = max(0, int(min_r0) - 5)
            it = tbl.item(int(target_row0), 4)
            if it is not None:
                tbl.scrollToItem(it, QAbstractItemView.ScrollHint.PositionAtTop)
        except Exception:
            pass

    def detach_pdf(self):
        if not self.drawing_pdf_path:
            QMessageBox.warning(self, "No Drawing", "Please load a Drawing PDF first.")
            return

        default_basename = ""
        try:
            default_basename = self._build_bubbled_drawing_basename()
        except Exception:
            default_basename = ""

        self.pdf_window = DrawingViewerWindow(pdf_path=self.drawing_pdf_path, default_save_basename=default_basename)
        try:
            v = getattr(self.pdf_window, "_pdf_viewer", None)
            if v is not None and hasattr(v, "bubbles_changed"):
                v.bubbles_changed.connect(self._on_drawing_bubbles_changed)
            if v is not None and hasattr(v, "drawing_saved"):
                v.drawing_saved.connect(self._on_drawing_saved)
            if v is not None and hasattr(v, "insert_notes_to_form3_requested"):
                v.insert_notes_to_form3_requested.connect(self._on_insert_notes_to_form3)
        except Exception:
            pass
        self.pdf_window.showMaximized()

    def _safe_filename_component(self, s: str) -> str:
        s = (s or "").strip()
        if not s:
            return ""
        s = re.sub(r"[\\/:*?\"<>|]", "_", s)
        s = re.sub(r"\s+", " ", s).strip()
        s = s.strip(" _.-")
        return s

    def _get_form1_fields_for_filenames(self) -> tuple[str, str, str, bool]:
        """Return (part_number, revision_level, lot_value, include_lot_label).

        include_lot_label=True when Form 1 B9 is a Job/Lot field label even if it has no value.
        """
        part = ""
        rev = ""
        lot = ""
        include_lot_label = False

        def _norm(s: object) -> str:
            return re.sub(r"\s+", " ", str(s or "")).strip().lower()

        def _looks_like_label(s: object, keywords: list[str]) -> bool:
            t = _norm(s)
            if not t:
                return False
            hits = 0
            for kw in keywords:
                if kw and kw in t:
                    hits += 1
            return hits >= max(1, int(len(keywords)))

        def _read_value_to_right(ws, row: int, col: int, *, max_off: int = 20) -> str:
            """Best-effort read of the nearest non-empty value cell to the right."""
            try:
                max_col = int(getattr(ws, "max_column", 0) or 0)
            except Exception:
                max_col = 0
            if max_col <= 0:
                max_col = col + max_off

            start_col = int(col)
            try:
                mr = cell_in_merged_range(ws, int(row), int(col))
                if mr is not None:
                    start_col = int(getattr(mr, "max_col", start_col))
            except Exception:
                start_col = int(col)

            for off in range(1, int(max_off) + 1):
                c2 = start_col + off
                if c2 > max_col:
                    break
                try:
                    tr, tc = merged_top_left(ws, int(row), int(c2))
                except Exception:
                    tr, tc = int(row), int(c2)
                try:
                    v = ws.cell(row=int(tr), column=int(tc)).value
                except Exception:
                    v = None
                if v is None:
                    continue
                s = str(v).strip()
                if not s:
                    continue
                return s
            return ""

        def _find_field_value(ws, field_no: int, keywords: list[str]) -> str:
            """Find a label like '1. Part Number' and return the entry value.

            Many AS9102 templates place the entry cell directly below the label.
            """
            try:
                max_row = min(int(getattr(ws, "max_row", 0) or 0), 60)
                max_col = min(int(getattr(ws, "max_column", 0) or 0), 30)
            except Exception:
                max_row, max_col = 60, 30

            needle_no = f"{int(field_no)}."
            best = None  # (hits, rr, cc)
            for rr in range(1, max_row + 1):
                for cc in range(1, max_col + 1):
                    try:
                        v = ws.cell(row=rr, column=cc).value
                    except Exception:
                        v = None
                    t = _norm(v)
                    if not t:
                        continue
                    if needle_no not in t:
                        continue
                    hits = sum(1 for kw in keywords if kw in t)
                    if hits <= 0:
                        continue
                    if best is None or hits > best[0]:
                        best = (hits, rr, cc)

            if best is None:
                return ""

            _hits, rr, cc = best
            
            def _is_next_field_label(s: str) -> bool:
                # Returns True if 's' looks like "8. Additional Changes" or "10." etc.
                s = s.strip()
                if not s:
                    return False
                # Check for "Digit(s). Text" or just "Digit(s)."
                return bool(re.match(r"^\d+\.", s))

            # Prefer the entry cell below the label.
            try:
                tr, tc = merged_top_left(ws, rr + 1, cc)
                below = str(ws.cell(row=int(tr), column=int(tc)).value or "").strip()
            except Exception:
                below = ""
            
            # Check if 'below' is actually the label for the next field (e.g. "8. ...")
            if below and not _looks_like_label(below, keywords) and not _is_next_field_label(below):
                return below

            right = _read_value_to_right(ws, rr, cc)
            if right and not _looks_like_label(right, keywords) and not _is_next_field_label(right):
                return right

            return ""

        def _lot_from_job_field(raw: object) -> tuple[str, bool]:
            s = str(raw or "").strip()
            if not s:
                return ("", False)
            s_norm = re.sub(r"\s+", " ", s).strip()
            low = s_norm.lower()
            # Placeholder-only values should not appear in filenames.
            if low in ("job#", "job #", "job number", "job no", "job no."):
                return ("", True)
            # If the user left the label and typed a value after it, strip the label.
            m = re.match(r"^job\s*#?\s*(.*)$", s_norm, flags=re.IGNORECASE)
            if m:
                rest = str(m.group(1) or "").strip()
                # If they typed only the label, treat as label-only.
                if not rest:
                    return ("", True)
                return (rest, True)
            return (s_norm, False)

        try:
            if self._template_wb is None:
                return ("", "", "")
            form1_name = self._form_sheet_names.get("1")
            if not form1_name or form1_name not in self._template_wb.sheetnames:
                return ("", "", "")
            ws = self._template_wb[form1_name]

            # Priority 1: Check B5 (Value)
            try:
                v5 = str(ws["B5"].value or "").strip()
                if v5 and not _looks_like_label(v5, ["part", "number"]):
                    part = v5
            except Exception:
                pass

            # Priority 2: Check B4 (Label/Value), stripped of leading number.
            if not part:
                try:
                    v4 = str(ws["B4"].value or "").strip()
                    if v4:
                        # "1. Part Number" -> "Part Number"
                        part = re.sub(r"^\d+[\.\)]?\s*", "", v4).strip()
                except Exception:
                    pass

            # Priority 3: Scan widely
            if not part:
                part = _find_field_value(ws, 1, ["part", "number"]).strip()
            
            if not part:
                # Fallback to prior fixed-cell heuristics (legacy path, rarely hit now)
                part_cell = ws["B5"].value
                if _looks_like_label(part_cell, ["part", "number"]):
                    part = _read_value_to_right(ws, 5, 2)
                else:
                    part = str(part_cell or "").strip()
                if not part:
                    part_cell2 = ws["B4"].value
                    if _looks_like_label(part_cell2, ["part", "number"]):
                        part = _read_value_to_right(ws, 4, 2)
                    else:
                        part = str(part_cell2 or "").strip()

            # Lot comes from the Job # entry cell. Do not read adjacent cells (C9 etc).
            job_raw = ""
            try:
                max_row = min(int(getattr(ws, "max_row", 0) or 0), 60)
                max_col = min(int(getattr(ws, "max_column", 0) or 0), 30)
            except Exception:
                max_row, max_col = 60, 30
            for rr in range(1, max_row + 1):
                for cc in range(1, max_col + 1):
                    try:
                        v = ws.cell(row=rr, column=cc).value
                    except Exception:
                        v = None
                    t = _norm(v)
                    if not t:
                        continue
                    if t in ("job#", "job #") or t.startswith("job #") or t.startswith("job#"):
                        job_raw = str(v or "").strip()
                        break
                if job_raw:
                    break
            if job_raw:
                lot, include_lot_label = _lot_from_job_field(job_raw)
            else:
                # Last-resort fallback (older templates).
                lot, include_lot_label = _lot_from_job_field(ws["B9"].value)

            # Revision extraction (User request: B7 priority)
            rev = ""
            try:
                # 1. Try B7 directly
                v7 = str(ws["B7"].value or "").strip()
                if v7 and not _looks_like_label(v7, ["revision", "level"]):
                    rev = v7
                
                # 2. Try scanning for Field 7
                if not rev:
                    rev = _find_field_value(ws, 7, ["revision"]).strip()
            except Exception:
                pass
        except Exception:
            return ("", "", "", False)

        return (
            self._safe_filename_component(part),
            self._safe_filename_component(rev),
            self._safe_filename_component(lot),
            bool(include_lot_label),
        )

    def _build_report_default_filename(self) -> str:
        part, rev, lot, include_lot_label = self._get_form1_fields_for_filenames()
        pieces: list[str] = []
        pieces.append(part or "Part Number")
        
        # Add Revision (User request: Rev <val> or just Rev)
        if rev:
            pieces.append(f"Rev {rev}")
        else:
            pieces.append("Rev")

        if include_lot_label:
            pieces.append(f"Lot {lot}".strip())
        pieces.append(datetime.date.today().strftime("%m-%d-%y"))

        base = " ".join(pieces).strip() or "FAI Report"
        if not base.lower().endswith(".xlsx"):
            base += ".xlsx"
        return base

    def _build_bubbled_drawing_basename(self) -> str:
        part, rev, lot, include_lot_label = self._get_form1_fields_for_filenames()
        pieces: list[str] = ["Bubbled Drawing"]
        pieces.append(part or "Part Number")

        # Add Revision
        if rev:
            pieces.append(f"Rev {rev}")
        else:
            pieces.append("Rev")

        if include_lot_label:
            pieces.append(f"Lot {lot}".strip())
        pieces.append(datetime.date.today().strftime("%m-%d-%y"))
        return self._safe_filename_component(" ".join(pieces).strip())

    def _default_output_dir_from_drawing_pdf(self) -> str:
        try:
            p = str(getattr(self, "drawing_pdf_path", "") or "").strip()
            if p:
                return os.path.dirname(p)
        except Exception:
            pass
        return ""

    def _refresh_drawing_viewer_default_save_basename(self) -> None:
        base = ""
        try:
            base = self._build_bubbled_drawing_basename()
        except Exception:
            base = ""
        if not base:
            return

        for viewer_attr in ("drawing_viewer_tab", "pdf_window"):
            try:
                w = getattr(self, viewer_attr, None)
                v = getattr(w, "_pdf_viewer", None) if w is not None else None
                if v is not None and hasattr(v, "default_save_basename"):
                    v.default_save_basename = str(base)
            except Exception:
                pass

    def create_file_row(self, line_edit, btn):
        widget = QWidget()
        layout = QHBoxLayout(widget)
        layout.setContentsMargins(0,0,0,0)
        layout.addWidget(line_edit)
        layout.addWidget(btn)
        return widget

    def browse_chr(self):
        start_dir = ""
        try:
            cur = str(self.chr_path_edit.text() or "").strip()
            if cur:
                start_dir = os.path.dirname(cur)
        except Exception:
            start_dir = ""
        if not start_dir:
            try:
                start_dir = str(self._settings.value("last_dir/chr", "") or "").strip()
            except Exception:
                start_dir = ""
        if not start_dir:
            try:
                last_chr = str(self._settings.value("paths/chr", "") or "").strip()
                if last_chr:
                    start_dir = os.path.dirname(last_chr)
            except Exception:
                start_dir = ""

        path, _ = QFileDialog.getOpenFileName(self, "Open Calypso File", start_dir, "Text Files (*.txt);;All Files (*)")
        if path:
            try:
                self._settings.setValue("last_dir/chr", os.path.dirname(path))
            except Exception:
                pass
            self.load_chr(path)

            # After browsing a Calypso file, prompt for the machine selection.
            try:
                self._refresh_calibrated_equipment_combo(preserve_selection=True)
                QTimer.singleShot(0, self.calibrated_equipment_combo.showPopup)
            except Exception:
                pass

    def browse_template(self):
        start_dir = ""
        try:
            cur = str(self.template_path_edit.text() or "").strip()
            if cur:
                start_dir = os.path.dirname(cur)
        except Exception:
            start_dir = ""
        if not start_dir:
            try:
                start_dir = str(self._settings.value("last_dir/template", "") or "").strip()
            except Exception:
                start_dir = ""
        if not start_dir:
            try:
                last_template = str(self._settings.value("paths/template", "") or "").strip()
                if last_template:
                    start_dir = os.path.dirname(last_template)
            except Exception:
                start_dir = ""

        path, _ = QFileDialog.getOpenFileName(self, "Open FAI Template", start_dir, "Excel Files (*.xlsx);;All Files (*)")
        if path:
            try:
                self._settings.setValue("last_dir/template", os.path.dirname(path))
            except Exception:
                pass
            self.template_path = path
            self.template_path_edit.setText(path)
            self._settings.setValue("paths/template", path)
            self.load_template()

    def browse_drawing_pdf(self):
        start_dir = ""
        try:
            cur = str(self.drawing_pdf_edit.text() or "").strip()
            if cur:
                start_dir = os.path.dirname(cur)
        except Exception:
            start_dir = ""
        if not start_dir:
            try:
                start_dir = str(self._settings.value("last_dir/drawing_pdf", "") or "").strip()
            except Exception:
                start_dir = ""
        if not start_dir:
            try:
                last_drawing = str(self._settings.value("paths/drawing_pdf", "") or "").strip()
                if last_drawing:
                    start_dir = os.path.dirname(last_drawing)
            except Exception:
                start_dir = ""

        path, _ = QFileDialog.getOpenFileName(self, "Open Drawing PDF", start_dir, "PDF Files (*.pdf);;All Files (*)")
        if path:
            self._set_drawing_pdf_path(path, load_viewer=True)

    def _set_drawing_pdf_path(self, path: str, *, load_viewer: bool = True) -> None:
        p = str(path or "").strip()
        if not p:
            return

        debug_pdf = False
        try:
            debug_pdf = str(os.environ.get("AS9102_DEBUG_PDF", "") or "").strip() not in ("", "0", "false", "False")
        except Exception:
            debug_pdf = False

        try:
            self.drawing_pdf_path = p
        except Exception:
            pass

        try:
            if hasattr(self, "drawing_pdf_edit") and self.drawing_pdf_edit is not None:
                if str(self.drawing_pdf_edit.text() or "").strip() != p:
                    self.drawing_pdf_edit.setText(p)
        except Exception:
            pass

        try:
            if hasattr(self, "_settings") and self._settings is not None:
                self._settings.setValue("paths/drawing_pdf", p)
                try:
                    self._settings.setValue("last_dir/drawing_pdf", os.path.dirname(p))
                except Exception:
                    pass
        except Exception:
            pass

        if not load_viewer:
            return

        if not os.path.exists(p):
            if debug_pdf:
                print(f"[AS9102_DEBUG_PDF] Drawing path does not exist: {p}", flush=True)
            return

        # Reload both the embedded viewer tab and any detached viewer window.
        for viewer_attr in ("drawing_viewer_tab", "pdf_window"):
            try:
                v = getattr(self, viewer_attr, None)
                if v is None:
                    continue
                if hasattr(v, "load_pdf"):
                    v.load_pdf(p)
            except Exception as e:
                if debug_pdf:
                    print(f"[AS9102_DEBUG_PDF] Failed to load PDF into {viewer_attr}: {e}", flush=True)

        try:
            self._refresh_drawing_viewer_default_save_basename()
        except Exception:
            pass

        QTimer.singleShot(75, self._sync_bubbles_to_form3)

    def _on_drawing_pdf_edit_committed(self) -> None:
        try:
            p = str(self.drawing_pdf_edit.text() or "").strip()
        except Exception:
            p = ""
        if not p:
            return
        self._set_drawing_pdf_path(p, load_viewer=True)

    def load_chr(self, path):
        self.chr_path_edit.setText(path)
        self._settings.setValue("paths/chr", path)
        self.characteristics = self.parser.parse_file(path)

        # If a machine is already selected and a template is loaded, apply it now.
        if self._template_wb is not None:
            try:
                self._apply_selected_calibrated_equipment_to_workbook()
            except Exception:
                pass

        # Re-render Form 3 with new data (if template is loaded)
        if self._template_wb is not None and self._form_sheet_names.get("3") and self._form_viewers.get("3"):
            ws = self._template_wb[self._form_sheet_names["3"]]
            self._write_form3_to_worksheet(ws)
            self._form_viewers["3"].set_overrides({})
            self._form_viewers["3"].render()
            # The Form 3 table may have just gained bubble numbers; re-sync fills.
            QTimer.singleShot(0, self._sync_bubbles_to_form3)

    def generate_report(self):
        if not self.template_path:
            QMessageBox.warning(self, "No Template", "Please select an FAI Template.")
            return

        default_dir = self._default_output_dir_from_drawing_pdf()
        default_name = ""
        try:
            default_name = self._build_report_default_filename()
        except Exception:
            default_name = ""
        default_path = os.path.join(default_dir, default_name) if default_dir and default_name else ""

        output_path, _ = QFileDialog.getSaveFileName(self, "Save FAI Report", default_path, "Excel Files (*.xlsx)")
        if not output_path:
            return
            
        # Prefer saving the in-memory workbook (includes UI edits).
        if self._template_wb is not None:
            # Ensure calibrated equipment mapping is written before saving.
            try:
                self._apply_selected_calibrated_equipment_to_workbook()
            except Exception:
                pass

            try:
                self._template_wb.save(output_path)
                self._wb_dirty = False
                
                # Post-processing: Remove internal helper sheets for Net-Inspect compatibility
                try:
                    wb_clean = openpyxl.load_workbook(output_path)
                    sheets_to_remove = [s for s in wb_clean.sheetnames if s.startswith("__as9102_")]
                    if sheets_to_remove:
                        for s in sheets_to_remove:
                            del wb_clean[s]
                        wb_clean.save(output_path)
                except Exception as e:
                    print(f"Warning: Failed to clean up helper sheets: {e}")

            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to save workbook:\n{e}")
                return
            QMessageBox.information(self, "Success", f"Workbook saved to:\n{output_path}")
            return

        # Fallback to generator-based output.
        notes = ""
        if self.pdf_path:
            try:
                extractor = PdfTextExtractor()
                notes = extractor.extract_text(self.pdf_path)
            except Exception as e:
                print(f"Error extracting notes: {e}")

        generator = FaiGenerator(self.template_path)
        success = generator.generate_report(self.characteristics, output_path, notes)

        if success:
            self._wb_dirty = False
            QMessageBox.information(self, "Success", f"Report saved to:\n{output_path}")
        else:
            QMessageBox.critical(self, "Error", "Failed to generate report. Check console for details.")

    def save_workbook(self):
        if self._template_wb is None or not self.template_path:
            QMessageBox.warning(self, "No Template", "Please load an FAI Template first.")
            return

        output_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save XLSX",
            "",
            "Excel Files (*.xlsx)",
        )
        if not output_path:
            return

        try:
            self._template_wb.save(output_path)
            self._wb_dirty = False
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save workbook:\n{e}")
            return
        QMessageBox.information(self, "Saved", f"Saved to:\n{output_path}")

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.accept()
        else:
            event.ignore()

    def dropEvent(self, event):
        files = [u.toLocalFile() for u in event.mimeData().urls()]
        for f in files:
            if f.lower().endswith('.txt'):
                self.load_chr(f)
            elif f.lower().endswith('.xlsx'):
                self.template_path = f
                self.template_path_edit.setText(f)
                self._settings.setValue("paths/template", f)
                try:
                    self._settings.setValue("last_dir/template", os.path.dirname(f))
                except Exception:
                    pass
                self.load_template()
            elif f.lower().endswith('.pdf'):
                self._set_drawing_pdf_path(f, load_viewer=True)

    def _set_wb_dirty(self) -> None:
        self._wb_dirty = True

    def closeEvent(self, event):
        """Handle application close, ensuring child windows like the Drawing Viewer are closed properly and unsaved data is handled."""
        
        # 1. Ask to save Form (Excel) changes
        if getattr(self, "_wb_dirty", False):
            mb = QMessageBox(self)
            mb.setIcon(QMessageBox.Question)
            mb.setWindowTitle("Unsaved Form Changes")
            mb.setText("You have unsaved changes in the FAI Forms. Generate Report before closing?")
            mb.setStandardButtons(QMessageBox.Save | QMessageBox.Discard | QMessageBox.Cancel)
            mb.setDefaultButton(QMessageBox.Save)
            choice = mb.exec()
            
            if choice == QMessageBox.Save:
                self.generate_report()
                # If still dirty, user cancelled save or it failed
                if getattr(self, "_wb_dirty", False):
                    event.ignore()
                    return
            elif choice == QMessageBox.Cancel:
                event.ignore()
                return
            # On Discard, continue to close

        # 2. Ask to save Drawing Viewer changes
        try:
            dv = getattr(self, "drawing_viewer_tab", None)
            if dv is not None:
                is_popped_out = bool(dv.windowFlags() & Qt.Window)
                
                if is_popped_out:
                    # If popped out, we must close the secondary window.
                    # Its closeEvent will handle the prompt.
                    
                    # Disable auto-dock callback temporarily to force close instead of redocking
                    old_cb = getattr(dv, "_dock_back_callback", None)
                    try:
                        dv._dock_back_callback = None
                        if not dv.close():
                            # Close cancelled by user
                            event.ignore()
                            dv._dock_back_callback = old_cb
                            return
                    except Exception:
                        pass
                else:
                    # If docked, the widget doesn't get a closeEvent automatically.
                    # We must manually check dirty state and ask.
                    pv = getattr(dv, "_pdf_viewer", None)
                    if pv is not None and hasattr(pv, "can_close"):
                        if not pv.can_close():
                            event.ignore()
                            return

        except Exception:
            pass
        
        super().closeEvent(event)

    def _set_theme(self, mode: str) -> None:
        """Switch between Light/Dark themes."""
        self._settings.setValue("theme", mode)
        app = QApplication.instance()
        if mode == "Light":
            # Reset to standard palette (usually light)
            # Setting "Windows" or "Fusion" without custom palette usually gives light theme
            app.setStyle("WindowsVista") # or 'Windows' or default
            app.setPalette(app.style().standardPalette())
            app.setStyleSheet("")
        else:
            # Mode is Dark
            try:
                import qdarktheme
                app.setStyleSheet(qdarktheme.load_stylesheet())
            except ImportError:
                self._set_manual_dark_theme()

        # Always ensure combo-box popups are readable (some themes make them transparent,
        # and stale widget palettes can leave white-on-white text after toggling).
        try:
            self._apply_combobox_readability_fix()
        except Exception:
            pass

        # App-wide button polish (no hard-coded colors).
        try:
            self._apply_button_polish_fix()
        except Exception:
            pass

    def _apply_combobox_readability_fix(self) -> None:
        app = QApplication.instance()
        if app is None:
            return

        fix = (
            "\n"
            "QComboBox { background-color: palette(base); color: palette(text); }\n"
            "QComboBox QAbstractItemView {\n"
            "  background-color: palette(base);\n"
            "  color: palette(text);\n"
            "  selection-background-color: palette(highlight);\n"
            "  selection-color: palette(highlighted-text);\n"
            "}\n"
        )

        try:
            current = app.styleSheet() or ""
        except Exception:
            current = ""
        if fix.strip() in current:
            return
        try:
            app.setStyleSheet(current + fix)
        except Exception:
            pass

    def _apply_button_polish_fix(self) -> None:
        """Apply a small, palette-driven button polish stylesheet."""
        app = QApplication.instance()
        if app is None:
            return

        fix = (
            "\n"
            "QPushButton, QToolButton {\n"
            "  padding: 4px 10px;\n"
            "  min-height: 24px;\n"
            "  border-radius: 6px;\n"
            "}\n"
            "QPushButton {\n"
            "  border: 1px solid palette(mid);\n"
            "}\n"
            "QPushButton:hover:!disabled {\n"
            "  border-color: palette(highlight);\n"
            "}\n"
            "QPushButton:pressed {\n"
            "  border-color: palette(shadow);\n"
            "}\n"
            "QPushButton:focus {\n"
            "  border-color: palette(highlight);\n"
            "}\n"
            "QToolButton {\n"
            "  border: 1px solid transparent;\n"
            "}\n"
            "QToolButton:hover:!disabled {\n"
            "  border-color: palette(mid);\n"
            "}\n"
            "QToolButton:pressed {\n"
            "  border-color: palette(shadow);\n"
            "}\n"
        )

        try:
            current = app.styleSheet() or ""
        except Exception:
            current = ""
        if fix.strip() in current:
            return
        try:
            app.setStyleSheet(current + fix)
        except Exception:
            pass

    def _set_manual_dark_theme(self):
        """Fallback dark theme using QPalette."""
        dark_palette = QPalette()
        dark_palette.setColor(QPalette.Window, QColor(53, 53, 53))
        dark_palette.setColor(QPalette.WindowText, Qt.white)
        dark_palette.setColor(QPalette.Base, QColor(25, 25, 25))
        dark_palette.setColor(QPalette.AlternateBase, QColor(53, 53, 53))
        dark_palette.setColor(QPalette.ToolTipBase, Qt.white)
        dark_palette.setColor(QPalette.ToolTipText, Qt.white)
        dark_palette.setColor(QPalette.Text, Qt.white)
        dark_palette.setColor(QPalette.Button, QColor(53, 53, 53))
        dark_palette.setColor(QPalette.ButtonText, Qt.white)
        dark_palette.setColor(QPalette.BrightText, Qt.red)
        dark_palette.setColor(QPalette.Link, QColor(42, 130, 218))
        dark_palette.setColor(QPalette.Highlight, QColor(42, 130, 218))
        dark_palette.setColor(QPalette.HighlightedText, Qt.black)
        
        app = QApplication.instance()
        app.setPalette(dark_palette)
        app.setStyle("Fusion")


def main():
    app = QApplication(sys.argv)
    window = MainWindow()
    window.showMaximized()
    sys.exit(app.exec())

if __name__ == "__main__":
    main()
